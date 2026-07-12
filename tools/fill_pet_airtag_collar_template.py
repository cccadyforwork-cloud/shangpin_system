from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName


BASE_DIR = Path("/Users/cc/Desktop/宠物皮质airtag项圈")
TEMPLATE = BASE_DIR / "ANIMAL_COLLAR.xlsm"
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT = OUTPUT_DIR / "宠物项圈_v1.xlsm"
DESKTOP_OUTPUT = Path("/Users/cc/Desktop/宠物项圈_v1.xlsm")


ROWS = [
    {
        "sku": "CA-Pet-XQ-Airtag-Pink",
        "color": "Pink",
        "color_map": "Pink",
        "source_color": "宠物项圈粉红色",
        "title_color": "Pink",
        "highlight": "Pink, M",
        "image": "https://cbu01.alicdn.com/img/ibank/O1CN01F175nh1wGkVcDF0NI_!!2215585976281-0-cib.jpg",
    },
    {
        "sku": "CA-Pet-XQ-Airtag-Blue",
        "color": "Sky Blue",
        "color_map": "Blue",
        "source_color": "宠物项圈天蓝色",
        "title_color": "Sky Blue",
        "highlight": "Sky Blue, M",
        "image": "https://cbu01.alicdn.com/img/ibank/O1CN01NIxR7L1wGkVUwekM2_!!2215585976281-0-cib.jpg",
    },
    {
        "sku": "CA-Pet-XQ-Airtag-Black",
        "color": "Black",
        "color_map": "Black",
        "source_color": "宠物项圈黑色",
        "title_color": "Black",
        "highlight": "Black, M",
        "image": "https://cbu01.alicdn.com/img/ibank/O1CN01HeAZa81wGkVQG10WB_!!2215585976281-0-cib.jpg",
    },
]


COMMON_BULLETS = [
    "Built-in holder helps keep an Apple AirTag attached to the pet collar during daily walks, travel, and outdoor activity.",
    "Soft PU leather style collar has a smooth surface for comfortable daily wear for cats and small to medium dogs.",
    "Adjustable M size collar measures about 16.93 x 1.77 inches and uses a metal buckle for a secure fit.",
    "Integrated anti-lost tracker holder helps protect the AirTag from everyday movement while keeping it easy to access.",
    "Designed for pet tracking support only; Apple AirTag device is not included with this collar.",
]


def build_col_map(ws):
    cols = {}
    seen = {}
    for col in range(1, ws.max_column + 1):
        label = ws.cell(4, col).value
        if not label:
            continue
        seen[label] = seen.get(label, 0) + 1
        key = label if seen[label] == 1 else f"{label}.{seen[label] - 1}"
        cols[key] = col
    return cols


def setv(ws, row, cols, label, value):
    col = cols.get(label)
    if col and value not in (None, ""):
        ws.cell(row, col).value = value


def clearv(ws, row, cols, label):
    col = cols.get(label)
    if col:
        ws.cell(row, col).value = None


def copy_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


def allow_generic_brand(wb):
    dropdown = wb["Dropdown Lists"]
    dropdown["G5"] = "Generic"
    name = "ANIMAL_COLLARbrandmarketplace_idATVPDKIKX0DERlanguage_tagen_US1.value"
    wb.defined_names[name] = DefinedName(
        name=name,
        attr_text="'Dropdown Lists'!$G$4:$G$5",
    )


def fill_row(ws, row, cols, item):
    sku = item["sku"]
    title = (
        "AirTag Pet Collar, Soft PU Leather Adjustable Tracking Dog and Cat Collar "
        f"with Holder and Metal Buckle, M, {item['title_color']}, AirTag Not Included"
    )
    description = (
        "This AirTag pet collar combines a soft PU leather style strap with an integrated holder designed for Apple AirTag. "
        "The adjustable M size collar is made for daily walking, travel, and outdoor activity for cats and small to medium dogs. "
        "The metal buckle helps secure the fit, while the holder keeps the tracker attached during normal pet movement. "
        "Apple AirTag device is not included."
    )

    setv(ws, row, cols, "SKU", sku)
    setv(ws, row, cols, "Product Type", "ANIMAL_COLLAR")
    setv(ws, row, cols, "Listing Action", "Create or Replace (Full Update)")
    clearv(ws, row, cols, "Parentage Level")
    clearv(ws, row, cols, "Parent SKU")
    clearv(ws, row, cols, "Variation Theme Name")

    setv(ws, row, cols, "Item Name", title)
    setv(ws, row, cols, "Item Highlight", item["highlight"])
    setv(ws, row, cols, "Brand Name", "Generic")
    setv(ws, row, cols, "Item Type Keyword", "宠物用品 > 狗用品 > 狗牵引及项圈 > 狗项圈 > 狗普通项圈 (pet-collars)")
    setv(ws, row, cols, "Package Level", "Unit")
    setv(ws, row, cols, "Target Audience Keyword", "Dogs")
    setv(ws, row, cols, "Target Audience Keyword.1", "Cats")
    setv(ws, row, cols, "Model Number", sku)
    setv(ws, row, cols, "Model Name", "AirTag Pet Collar")
    setv(ws, row, cols, "Manufacturer", "Generic")
    setv(ws, row, cols, "Main Image URL", item["image"])
    setv(ws, row, cols, "Swatch Image URL", item["image"])

    setv(ws, row, cols, "Product Description", description)
    for idx, bullet in enumerate(COMMON_BULLETS, start=1):
        label = "Bullet Point" if idx == 1 else f"Bullet Point.{idx - 1}"
        text = bullet if idx != 5 else f"{item['title_color']} collar is designed for pet tracking support only; Apple AirTag device is not included."
        setv(ws, row, cols, label, text)
    setv(ws, row, cols, "Generic Keyword", "airtag pet collar; airtag dog collar; airtag cat collar; tracking collar; pu leather pet collar; anti lost collar")
    setv(ws, row, cols, "Style", "AirTag Holder Collar")
    setv(ws, row, cols, "Material", "Faux Leather")
    setv(ws, row, cols, "Material.1", "Polyurethane")
    setv(ws, row, cols, "Material.2", "Metal")
    setv(ws, row, cols, "Number of Items", 1)
    setv(ws, row, cols, "Item Package Quantity", 1)
    setv(ws, row, cols, "Color Map", item["color_map"])
    setv(ws, row, cols, "Color", item["color"])
    setv(ws, row, cols, "Size", "M")
    setv(ws, row, cols, "Part Number", sku)
    setv(ws, row, cols, "Care Instructions", "Spot Clean")
    setv(ws, row, cols, "Pattern", "Solid")
    setv(ws, row, cols, "Unit Count", 1)
    setv(ws, row, cols, "Unit Count Type", "Count")
    setv(ws, row, cols, "Included Components", "1 x Pet Collar")
    setv(ws, row, cols, "Specific Uses for Product", "Outdoor")
    setv(ws, row, cols, "Specific Uses for Product.1", "Active")
    setv(ws, row, cols, "Closure Type", "Buckle")

    # 1688 M size is 43 cm x 4.5 cm.
    setv(ws, row, cols, "Item Length Longer Side", 16.93)
    setv(ws, row, cols, "Item Length Unit", "Inches")
    setv(ws, row, cols, "Item Width Shorter Side", 1.77)
    setv(ws, row, cols, "Item Width Unit", "Inches")
    setv(ws, row, cols, "Item Width", 1.77)
    setv(ws, row, cols, "Width Unit", "Inches")
    setv(ws, row, cols, "Item Length", 16.93)
    setv(ws, row, cols, "Item Length Unit.1", "Inches")
    setv(ws, row, cols, "Dog Breed Size", "Medium")
    setv(ws, row, cols, "Dog Breed Size.1", "Small")
    setv(ws, row, cols, "Number of Packs", 1)
    setv(ws, row, cols, "Pet Type", "Dog")
    setv(ws, row, cols, "Pet Type.1", "Cat")
    setv(ws, row, cols, "Animal Collar Type", "Basic Animal Collar")
    setv(ws, row, cols, "Item Weight", 0.106)
    setv(ws, row, cols, "Item Weight Unit", "Pounds")

    setv(ws, row, cols, "Item Condition", "New")
    setv(ws, row, cols, "List Price", 6.99)
    setv(ws, row, cols, "Product Tax Code", "A_GEN_NOTAX")
    setv(ws, row, cols, "Main Image Location", item["image"])
    setv(ws, row, cols, "Your Price USD (Low-Cost Store, US)", 6.99)
    setv(ws, row, cols, "Item Package Length", 3.15)
    setv(ws, row, cols, "Package Length Unit", "Inches")
    setv(ws, row, cols, "Item Package Width", 3.54)
    setv(ws, row, cols, "Package Width Unit", "Inches")
    setv(ws, row, cols, "Item Package Height", 1.38)
    setv(ws, row, cols, "Package Height Unit", "Inches")
    setv(ws, row, cols, "Package Weight", 0.106)
    setv(ws, row, cols, "Package Weight Unit", "Pounds")
    setv(ws, row, cols, "Number of Boxes", 1)
    setv(ws, row, cols, "Country of Origin", "China")
    setv(ws, row, cols, "Are batteries required?", "No")
    setv(ws, row, cols, "Are batteries included?", "No")
    setv(ws, row, cols, "Dangerous Goods Regulations", "Not Applicable")
    setv(ws, row, cols, "Directions", "Adjust collar to fit the pet comfortably before use. Check the fit regularly and remove if damaged. AirTag is not included.")


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(TEMPLATE, keep_vba=True)
    allow_generic_brand(wb)
    ws = wb["Template"]
    cols = build_col_map(ws)

    for row in range(7, 10):
        copy_row_style(ws, 6, row)

    for offset, item in enumerate(ROWS):
        fill_row(ws, 7 + offset, cols, item)

    wb.save(OUTPUT)
    wb.save(DESKTOP_OUTPUT)
    print(OUTPUT)
    print(DESKTOP_OUTPUT)


if __name__ == "__main__":
    main()
