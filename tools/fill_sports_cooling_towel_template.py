from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName


BASE_DIR = Path("/Users/cc/Desktop/运动毛巾资料")
TEMPLATE = BASE_DIR / "TOWEL.xlsm"
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT = OUTPUT_DIR / "运动毛巾_TOWEL_无品牌走单填写稿_修正版.xlsm"
DESKTOP_OUTPUT = Path("/Users/cc/Desktop/运动毛巾_TOWEL_无品牌走单填写稿_修正版.xlsm")


COMMON_BULLETS = [
    "Double-sided cooling towel made with soft microfiber fabric for sports, gym, yoga, running, camping, and hot weather activities.",
    "Quick-dry and breathable fabric helps absorb sweat while staying lightweight and comfortable around the neck, face, or shoulders.",
    "Cooling towel can be wet, wrung out, and snapped to help create a refreshing cooling feel during workouts or outdoor use.",
    "Compact 30 x 113 cm towel is easy to fold, carry, and pack in a gym bag, backpack, or travel pouch.",
    "Reusable sports towel is suitable for men and women and can be machine washed for everyday training and outdoor activities.",
]

SKU_ROWS = [
    {
        "sku": "CA-Towel-green",
        "source_name": "ML100森林绿",
        "color": "Green",
        "highlight": "Forest Green",
        "item_name": "Cooling Sports Towel, Double-Sided Microfiber Quick Dry Ice Towel for Gym Yoga Running Camping, Green",
        "image": "https://cbu01.alicdn.com/img/ibank/O1CN01niFjcd2BitPD0fnnR_!!2081488373-0-cib.jpg",
    },
    {
        "sku": "CA-Towel-grey",
        "source_name": "ML100冰川灰",
        "color": "Gray",
        "highlight": "Glacier Gray",
        "item_name": "Cooling Sports Towel, Double-Sided Microfiber Quick Dry Ice Towel for Gym Yoga Running Camping, Gray",
        "image": "https://cbu01.alicdn.com/img/ibank/O1CN01BNv7pX2BitPDKHkj5_!!2081488373-0-cib.jpg",
    },
    {
        "sku": "CA-Towel-pink",
        "source_name": "ML100樱花粉",
        "color": "Pink",
        "highlight": "Sakura Pink",
        "item_name": "Cooling Sports Towel, Double-Sided Microfiber Quick Dry Ice Towel for Gym Yoga Running Camping, Pink",
        "image": "https://cbu01.alicdn.com/img/ibank/O1CN01kxKPdd2BitPD9DEQL_!!2081488373-0-cib.jpg",
    },
]


def col_map(ws):
    cols = {}
    seen = {}
    for col in range(1, ws.max_column + 1):
        label = ws.cell(4, col).value
        if not label:
            continue
        seen[label] = seen.get(label, 0) + 1
        key = label if seen[label] == 1 else f"{label}.{seen[label] - 1}"
        cols[key] = col
    return cols


def setv(ws, row, cols, label, value):
    col = cols.get(label)
    if col and value is not None:
        ws.cell(row, col).value = value


def copy_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


def allow_generic_brand(wb):
    dropdown = wb["Dropdown Lists"]
    dropdown["G5"] = "Generic"
    name = "TOWELbrandmarketplace_idATVPDKIKX0DERlanguage_tagen_US1.value"
    wb.defined_names[name] = DefinedName(
        name=name,
        attr_text="'Dropdown Lists'!$G$4:$G$5",
    )


def fill_common(ws, row, cols, item):
    setv(ws, row, cols, "SKU", item["sku"])
    setv(ws, row, cols, "Product Type", "TOWEL")
    setv(ws, row, cols, "Listing Action", "Create or Replace (Full Update)")
    setv(ws, row, cols, "Item Name", item["item_name"])
    setv(ws, row, cols, "Item Highlight", item["highlight"])
    setv(ws, row, cols, "Brand Name", "Generic")
    setv(ws, row, cols, "Item Type Keyword", "yoga-towels")
    setv(ws, row, cols, "Package Level", "Unit")
    setv(ws, row, cols, "Target Audience", "Men")
    setv(ws, row, cols, "Target Audience.1", "Women")
    setv(ws, row, cols, "Model Number", item["sku"])
    setv(ws, row, cols, "Model Name", "Double-Sided Cooling Sports Towel")
    setv(ws, row, cols, "Manufacturer", "Generic")
    setv(
        ws,
        row,
        cols,
        "Product Description",
        "A double-sided microfiber cooling sports towel designed for gym, yoga, running, camping, golf, hiking, travel, and other hot weather activities. The soft breathable fabric absorbs sweat, dries quickly, and can be used around the neck, face, head, or shoulders for portable cooling comfort.",
    )
    for idx, bullet in enumerate(COMMON_BULLETS, start=1):
        setv(ws, row, cols, "Bullet Point" if idx == 1 else f"Bullet Point.{idx - 1}", bullet)
    setv(ws, row, cols, "Generic Keyword", "cooling towel; sports towel; microfiber towel; gym towel; yoga towel; quick dry towel")
    for idx, feature in enumerate(["Quick Dry", "Breathable", "Lightweight", "Reversible", "Super Absorbent"], start=1):
        setv(ws, row, cols, "Special Features" if idx == 1 else f"Special Features.{idx - 1}", feature)
    setv(ws, row, cols, "Style", "Cooling Sports Towel")
    setv(ws, row, cols, "Material", "Polyester")
    setv(ws, row, cols, "Material.1", "Nylon")
    setv(ws, row, cols, "Fabric Type", "Microfiber")
    setv(ws, row, cols, "Number of Items", 1)
    setv(ws, row, cols, "Item Package Quantity", 1)
    setv(ws, row, cols, "Color", item["color"])
    setv(ws, row, cols, "Size", "Small")
    setv(ws, row, cols, "Part Number", item["sku"])
    setv(ws, row, cols, "Theme", "Fitness")
    setv(ws, row, cols, "Weave Type", "Waffle")
    setv(ws, row, cols, "Care Instructions", "Machine Wash")
    setv(ws, row, cols, "Pattern", "Solid")
    setv(ws, row, cols, "Unit Count", 1)
    setv(ws, row, cols, "Unit Count Type", "Count")
    setv(ws, row, cols, "Towel Form Type", "Cooling Towel")
    setv(ws, row, cols, "Item Weight", 0.152)
    setv(ws, row, cols, "Item Weight Unit", "Pounds")
    setv(ws, row, cols, "Item Condition", "New")
    setv(ws, row, cols, "List Price", 3.99)
    setv(ws, row, cols, "Product Tax Code", "A_GEN_NOTAX")
    setv(ws, row, cols, "Main Image Location", item["image"])
    setv(ws, row, cols, "Your Price USD (Low-Cost Store, US)", 3.99)
    setv(ws, row, cols, "Country of Origin", "China")
    setv(ws, row, cols, "Are batteries required?", "No")
    setv(ws, row, cols, "Are batteries included?", "No")
    setv(ws, row, cols, "Dangerous Goods Regulations", "Not Applicable")
    setv(ws, row, cols, "Product Compliance Certificate", "Not Applicable")


def fill_dimensions(ws, row, cols):
    # Final price table controls package dimensions and shipping weight.
    setv(ws, row, cols, "Item Package Length", 3.94)
    setv(ws, row, cols, "Package Length Unit", "Inches")
    setv(ws, row, cols, "Item Package Width", 5.12)
    setv(ws, row, cols, "Package Width Unit", "Inches")
    setv(ws, row, cols, "Item Package Height", 1.57)
    setv(ws, row, cols, "Package Height Unit", "Inches")
    setv(ws, row, cols, "Package Weight", 0.152)
    setv(ws, row, cols, "Package Weight Unit", "Pounds")
    setv(ws, row, cols, "Item Display Weight", 0.152)
    setv(ws, row, cols, "Item Display Weight Unit", "Pounds")

    # 1688 specification is 30 cm x 113 cm for the unfolded towel.
    setv(ws, row, cols, "Item Length Longer Edge", 44.49)
    setv(ws, row, cols, "Item Length Unit", "Inches")
    setv(ws, row, cols, "Item Width Shorter Edge", 11.81)
    setv(ws, row, cols, "Item Width Unit", "Inches")


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(TEMPLATE, keep_vba=True)
    allow_generic_brand(wb)
    ws = wb["Template"]
    cols = col_map(ws)

    for row in range(7, 10):
        copy_row_style(ws, 6, row)

    for offset, item in enumerate(SKU_ROWS):
        row = 7 + offset
        fill_common(ws, row, cols, item)
        fill_dimensions(ws, row, cols)

    wb.save(OUTPUT)
    wb.save(DESKTOP_OUTPUT)
    print(OUTPUT)
    print(DESKTOP_OUTPUT)


if __name__ == "__main__":
    main()
