from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName


BASE_DIR = Path("/Users/cc/Desktop/宠物玩具拓展资料")
TEMPLATE = BASE_DIR / "PET_TOY.xlsm"
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT = OUTPUT_DIR / "黄麻猫玩具_PET_TOY_无品牌走单填写稿.xlsm"


COMMON_BULLETS = [
    "Natural jute and silvervine chew toy helps support chewing, batting, and pouncing play for indoor cats.",
    "Lightweight toy is easy for kittens and adult cats to chase, carry, and bat across the floor.",
    "Textured jute surface encourages chewing and may help clean teeth during supervised play.",
    "Designed for self-play enrichment to help reduce boredom for indoor cats.",
    "Small assorted character shape makes the toy suitable for daily interactive cat play.",
]

SKU_ROWS = [
    {
        "sku": "CA-Petchewtoy-fly",
        "source_name": "木天蓼薄荷蝴蝶",
        "title_style": "Butterfly",
        "item_name": "Natural Jute Silvervine Cat Chew Toy, Butterfly Shape Catnip Interactive Dental Toy for Indoor Cats",
        "highlight": "Butterfly Shape",
        "subject": "Butterfly",
        "shape": "Butterfly",
        "materials": ["Jute", "Silvervine", "Catnip"],
        "price": 4.99,
        "image": "https://cbu01.alicdn.com/img/ibank/O1CN01XLN3QP1LUouAOA2kr_!!2978681303-0-cib.jpg",
    },
    {
        "sku": "CA-Petchewtoy-mouse",
        "source_name": "老鼠+球套装",
        "title_style": "Mouse",
        "item_name": "Natural Jute Silvervine Cat Chew Toy, Mouse and Ball Interactive Dental Toy for Indoor Cats",
        "highlight": "Mouse and Ball",
        "subject": "Mouse",
        "shape": "Mouse",
        "materials": ["Jute", "Silvervine"],
        "price": 3.99,
        "image": "https://cbu01.alicdn.com/img/ibank/O1CN01YzC8GH1LUp2A6SNR5_!!2978681303-0-cib.jpg",
    },
    {
        "sku": "CA-Petchewtoy-bird",
        "source_name": "木天蓼鹦鹉",
        "title_style": "Parrot",
        "item_name": "Natural Jute Silvervine Cat Chew Toy, Parrot Shape Interactive Dental Toy for Indoor Cats",
        "highlight": "Parrot Shape",
        "subject": "Parrot",
        "shape": "Parrot",
        "materials": ["Jute", "Silvervine"],
        "price": 3.99,
        "image": "https://cbu01.alicdn.com/img/ibank/O1CN01Pt2wDy1LUox2k1Ie6_!!2978681303-0-cib.jpg",
    },
    {
        "sku": "CA-Petchewtoy-ball",
        "source_name": "木天蓼串串球",
        "title_style": "Ball",
        "item_name": "Natural Jute Silvervine Cat Chew Toy, String Ball Interactive Dental Toy for Indoor Cats",
        "highlight": "String Ball",
        "subject": "Ball",
        "shape": "Ball",
        "materials": ["Jute", "Silvervine"],
        "price": 3.99,
        "image": "https://cbu01.alicdn.com/img/ibank/O1CN01wGtP8n1LUouAO9m8a_!!2978681303-0-cib.jpg",
    },
]


def col_map(ws):
    cols = {}
    seen = {}
    for col in range(1, ws.max_column + 1):
        label = ws.cell(4, col).value
        if not label:
            continue
        seen[label] = seen.get(label, 0) + 1
        key = label if seen[label] == 1 else f"{label}.{seen[label] - 1}"
        cols[key] = col
    return cols


def setv(ws, row, cols, label, value):
    col = cols.get(label)
    if col and value is not None:
        ws.cell(row, col).value = value


def copy_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


def fill_common(ws, row, cols, item_name, highlight):
    setv(ws, row, cols, "Product Type", "PET_TOY")
    setv(ws, row, cols, "Listing Action", "Create or Replace (Full Update)")
    setv(ws, row, cols, "Item Name", item_name)
    setv(ws, row, cols, "Item Highlight", highlight)
    setv(ws, row, cols, "Brand Name", "Generic")
    setv(ws, row, cols, "Item Type Keyword", "pet-chew-toys")
    setv(ws, row, cols, "Target Audience Keyword", "Cats")
    setv(ws, row, cols, "Model Name", "Natural Jute Silvervine Cat Chew Toy")
    setv(ws, row, cols, "Manufacturer", "Generic")
    setv(ws, row, cols, "Product Description", "A small natural jute and silvervine cat chew toy designed for indoor cats to chew, bat, chase, and pounce. The textured surface supports dental chewing play, while the lightweight character shape encourages daily enrichment and self-play.")
    for idx, bullet in enumerate(COMMON_BULLETS, start=1):
        setv(ws, row, cols, "Bullet Point" if idx == 1 else f"Bullet Point.{idx - 1}", bullet)
    setv(ws, row, cols, "Generic Keyword", "cat chew toy; jute cat toy; silvervine cat toy; dental cat toy; indoor cat toy")
    setv(ws, row, cols, "Special Features", "Interactive")
    setv(ws, row, cols, "Special Features.1", "Chewable")
    setv(ws, row, cols, "Special Features.2", "Lightweight")
    setv(ws, row, cols, "Style", "Natural Chew Toy")
    setv(ws, row, cols, "Number of Items", 1)
    setv(ws, row, cols, "Item Package Quantity", 1)
    setv(ws, row, cols, "Color", "Natural")
    setv(ws, row, cols, "Size", "Small")
    setv(ws, row, cols, "Theme", "Animals")
    setv(ws, row, cols, "Breed Recommendation", "All Breed Sizes")
    setv(ws, row, cols, "Unit Count", 1)
    setv(ws, row, cols, "Unit Count Type", "Count")
    setv(ws, row, cols, "Included Components", "Cat chew toy")
    setv(ws, row, cols, "Specific Uses for Product", "Chewing")
    setv(ws, row, cols, "Recommended Uses For Product", "Indoor cat play")
    setv(ws, row, cols, "Pet Toy Type", "Chew Toy")
    setv(ws, row, cols, "Pet Type", "Cat")
    setv(ws, row, cols, "Item Weight", 0.055)
    setv(ws, row, cols, "Item Weight Unit", "Pounds")
    setv(ws, row, cols, "Item Condition", "New")
    setv(ws, row, cols, "Product Tax Code", "A_GEN_NOTAX")
    setv(ws, row, cols, "Country of Origin", "China")
    setv(ws, row, cols, "Are batteries required?", "No")
    setv(ws, row, cols, "Are batteries included?", "No")
    setv(ws, row, cols, "Dangerous Goods Regulations", "Not Applicable")
    setv(ws, row, cols, "Contains Liquid Contents?", "No")
    setv(ws, row, cols, "Number of Boxes", 1)
    setv(ws, row, cols, "Directions", "Use under supervision. Replace the toy if it becomes damaged.")


def fill_dimensions(ws, row, cols):
    for label in ["Item Display Length", "Length longer horizontal edge", "Item Package Length"]:
        setv(ws, row, cols, label, 3.94)
    for label in ["Item Display Width", "Width shorter horizontal edge", "Item Package Width"]:
        setv(ws, row, cols, label, 3.54)
    for label in ["Item Display Height", "Height base to top", "Item Package Height"]:
        setv(ws, row, cols, label, 1.38)
    for label in ["Item Display Length Unit", "Length Unit", "Package Length Unit"]:
        setv(ws, row, cols, label, "Inches")
    for label in ["Item Display Width Unit", "Width Unit", "Package Width Unit"]:
        setv(ws, row, cols, label, "Inches")
    for label in ["Item Display Height Unit", "Height Unit", "Package Height Unit"]:
        setv(ws, row, cols, label, "Inches")
    setv(ws, row, cols, "Package Weight", 0.055)
    setv(ws, row, cols, "Package Weight Unit", "Pounds")


def allow_generic_brand(wb):
    dropdown = wb["Dropdown Lists"]
    dropdown["G5"] = "Generic"
    name = "PET_TOYbrandmarketplace_idATVPDKIKX0DERlanguage_tagen_US1.value"
    wb.defined_names[name] = DefinedName(
        name=name,
        attr_text="'Dropdown Lists'!$G$4:$G$5",
    )


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(TEMPLATE, keep_vba=True)
    allow_generic_brand(wb)
    ws = wb["Template"]
    cols = col_map(ws)

    for row in range(7, 11):
        copy_row_style(ws, 6, row)

    for offset, item in enumerate(SKU_ROWS):
        row = 7 + offset
        setv(ws, row, cols, "SKU", item["sku"])
        fill_common(ws, row, cols, item["item_name"], item["highlight"])
        for index, material in enumerate(item["materials"], start=1):
            setv(ws, row, cols, "Material" if index == 1 else f"Material.{index - 1}", material)
        setv(ws, row, cols, "Subject Character", item["subject"])
        setv(ws, row, cols, "Item Shape", item["shape"])
        setv(ws, row, cols, "Part Number", item["sku"])
        setv(ws, row, cols, "Model Number", item["sku"])
        setv(ws, row, cols, "Main Image Location", item["image"])
        setv(ws, row, cols, "List Price", item["price"])
        setv(ws, row, cols, "Your Price USD (Low-Cost Store, US)", item["price"])
        fill_dimensions(ws, row, cols)

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
