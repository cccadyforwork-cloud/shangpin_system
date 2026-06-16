from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


INTAKE_HEADERS = [
    "project_name",
    "product_name",
    "route",
    "brand",
    "manufacturer",
    "category",
    "product_type",
    "item_type_keyword",
    "sku",
    "color",
    "size",
    "set_count",
    "material",
    "accessories",
    "cost",
    "target_price",
    "list_price",
    "haul_price",
    "package_length_in",
    "package_width_in",
    "package_height_in",
    "package_weight_lb",
    "country_of_origin",
    "batteries_required",
    "dangerous_goods",
    "main_image_url",
    "supplier_link",
    "competitor_links",
    "title",
    "bullet_1",
    "bullet_2",
    "bullet_3",
    "bullet_4",
    "bullet_5",
    "description",
    "notes"
]


REQUIRED_CORE_FIELDS = [
    "project_name",
    "product_name",
    "route",
    "brand",
    "manufacturer",
    "sku",
    "title",
    "list_price",
    "package_length_in",
    "package_width_in",
    "package_height_in",
    "package_weight_lb",
    "country_of_origin",
    "batteries_required",
    "dangerous_goods"
]


def create_intake_workbook(path):
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "产品资料"
    ws.append(INTAKE_HEADERS)
    ws.append([
        "示例项目",
        "示例收纳袋",
        "Haul Generic",
        "Generic",
        "Generic",
        "home",
        "EXERCISE_BLOCK",
        "storage-bags",
        "DEMO-STORAGE-BAG-001",
        "Black",
        "Large",
        1,
        "Oxford cloth",
        "1 storage bag",
        3.2,
        9.99,
        12.99,
        9.99,
        8,
        6,
        1.5,
        0.6,
        "China",
        "No",
        "No",
        "https://example.com/main.jpg",
        "https://example.com/supplier",
        "https://example.com/competitor",
        "Large Storage Bag for Travel and Closet Organization, Black",
        "Designed for travel, closet organization, and daily storage.",
        "Made with lightweight Oxford cloth for everyday use.",
        "Large opening helps keep clothing and accessories organized.",
        "Includes one black storage bag for simple packing.",
        "Portable design fits luggage, shelves, and dorm rooms.",
        "This storage bag keeps clothing and small essentials organized for travel, closets, dorms, and daily home use.",
        "示例行可以删除"
    ])

    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, header in enumerate(INTAKE_HEADERS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(header) + 4, 14), 28)

    notes = wb.create_sheet("填写说明")
    notes.append(["字段", "说明"])
    descriptions = {
        "route": "填写 Haul Generic 或 Brand。",
        "brand": "Generic 路线建议填写 Generic；品牌路线填写真实品牌。",
        "manufacturer": "Generic 路线建议和 Brand 逻辑一致；品牌路线与包装、图片、文案统一。",
        "target_price": "内部目标售价，用于判断 Haul 价格风险。",
        "haul_price": "Haul 上传价，若走 Haul 路线建议填写。",
        "package_length_in/package_width_in/package_height_in": "单位为 inches。",
        "package_weight_lb": "单位为 pounds。",
        "batteries_required/dangerous_goods": "建议填写 Yes 或 No。",
        "competitor_links": "多个链接可以用逗号或换行分隔。"
    }
    for field, desc in descriptions.items():
        notes.append([field, desc])
    notes.column_dimensions["A"].width = 32
    notes.column_dimensions["B"].width = 90
    for cell in notes[1]:
        cell.font = Font(bold=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_intake_workbook(path, rows):
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "产品资料"
    ws.append(INTAKE_HEADERS)

    for row in rows:
        ws.append([row.get(header, "") for header in INTAKE_HEADERS])

    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, header in enumerate(INTAKE_HEADERS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(header) + 4, 14), 30)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def read_intake_rows(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["产品资料"] if "产品资料" in wb.sheetnames else wb.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    rows = []
    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(value not in (None, "") for value in values):
            continue
        row = {headers[idx]: values[idx] if idx < len(values) else None for idx in range(len(headers))}
        row["_row_number"] = row_number
        rows.append(row)
    return rows
