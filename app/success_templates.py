import json
import warnings
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

from .paths import DATA_DIR, OUTPUTS_DIR
from .template_sheet import find_template_sheet
from .template_writer import FIELD_MAP


warnings.filterwarnings("ignore", message="Data Validation extension is not supported.*")

SUCCESS_TEMPLATES_DIR = DATA_DIR / "success_templates"
RULES_JSON = DATA_DIR / "success_template_rules.json"
RULES_REPORT = OUTPUTS_DIR / "成功模板规则提炼报告.md"

IMAGE_FIELD_KEYWORDS = ("image", "media_location", "swatch")
SYSTEM_FIELDS = {
    "contribution_sku#1.value",
    "product_type#1.value",
}
MIN_ALWAYS_FILLED_RATIO = 0.95
MIN_OFTEN_FILLED_RATIO = 0.6
MAX_EXAMPLE_VALUES = 8
MAX_CONSECUTIVE_BLANK_SKU_ROWS = 50
MAX_EXAMPLE_TEXT_LENGTH = 90


def learn_success_templates(source_dir=None, json_path=None, report_path=None):
    source_dir = Path(source_dir) if source_dir else SUCCESS_TEMPLATES_DIR
    json_path = Path(json_path) if json_path else RULES_JSON
    report_path = Path(report_path) if report_path else RULES_REPORT

    workbook_records = []
    by_product_type = defaultdict(_empty_product_type_stats)
    skipped = []

    for path in sorted(source_dir.glob("*"), key=lambda item: item.name):
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            continue
        if path.name.startswith("~$") or path.name.startswith(".~"):
            continue
        try:
            record = _analyze_workbook(path)
        except Exception as exc:
            skipped.append({
                "file": path.name,
                "reason": f"{type(exc).__name__}: {exc}",
            })
            continue
        workbook_records.append(record)
        for product_type in record["product_types"]:
            stats = by_product_type[product_type]
            stats["product_type"] = product_type
            stats["template_files"].append(record["file"])
            stats["template_count"] += 1
            stats["sku_count"] += record["sku_count"]
            stats["field_count_total"] += record["field_count"]
            for field_name, field_stats in record["fields"].items():
                target = stats["fields"][field_name]
                target["filled_rows"] += field_stats["filled_rows"]
                target["blank_rows"] += field_stats["blank_rows"]
                target["files"].add(record["file"])
                target["values"].update(field_stats["values"])

    product_rules = {}
    for product_type, stats in sorted(by_product_type.items()):
        product_rules[product_type] = _build_product_rules(stats)

    rules = {
        "source_dir": str(source_dir),
        "template_count": len(workbook_records),
        "skipped_count": len(skipped),
        "product_type_count": len(product_rules),
        "product_types": product_rules,
        "skipped": skipped,
    }

    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(rules, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(_build_report(rules), encoding="utf-8")
    return rules, json_path, report_path


def _empty_product_type_stats():
    return {
        "product_type": "",
        "template_count": 0,
        "sku_count": 0,
        "field_count_total": 0,
        "template_files": [],
        "fields": defaultdict(_empty_field_stats),
    }


def _empty_field_stats():
    return {
        "filled_rows": 0,
        "blank_rows": 0,
        "files": set(),
        "values": Counter(),
    }


def _analyze_workbook(path):
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Data Validation extension is not supported.*")
        wb = load_workbook(path, data_only=True, read_only=True)
    ws = find_template_sheet(wb)
    if ws is None:
        raise ValueError("没有 Template/模板 页")

    field_names = []
    for row_index, row_values in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_index == 5:
            field_names = [_clean_value(value) for value in row_values]
            break

    field_indexes = {
        field_name: index
        for index, field_name in enumerate(field_names)
        if field_name
    }
    sku_index = field_indexes.get("contribution_sku#1.value")
    product_type_index = field_indexes.get("product_type#1.value")
    if sku_index is None or product_type_index is None:
        raise ValueError("缺少 SKU 或 Product Type 字段")

    sku_count = 0
    product_types = set()
    field_stats = {
        field_name: _empty_field_stats_for_workbook()
        for field_name in field_indexes
        if not _should_ignore_field(field_name)
    }

    blank_streak = 0
    for row_index, row_values in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
        sku = _value_at(row_values, sku_index)
        if not sku:
            blank_streak += 1
            if sku_count and blank_streak >= MAX_CONSECUTIVE_BLANK_SKU_ROWS:
                break
            continue
        blank_streak = 0
        sku_count += 1
        product_type = _value_at(row_values, product_type_index)
        if product_type:
            product_types.add(product_type)
        for field_name, index in field_indexes.items():
            if field_name not in field_stats:
                continue
            value = _value_at(row_values, index)
            if value:
                field_stats[field_name]["filled_rows"] += 1
                field_stats[field_name]["values"][value] += 1
            else:
                field_stats[field_name]["blank_rows"] += 1

    fields = {
        field_name: stats
        for field_name, stats in field_stats.items()
        if stats["filled_rows"]
    }
    return {
        "file": path.name,
        "sku_count": sku_count,
        "product_types": sorted(product_types) or ["UNKNOWN"],
        "field_count": len(field_indexes),
        "fields": fields,
    }


def _empty_field_stats_for_workbook():
    return {
        "filled_rows": 0,
        "blank_rows": 0,
        "values": Counter(),
    }


def _value_at(row_values, index):
    if index >= len(row_values):
        return ""
    return _clean_value(row_values[index])


def _should_ignore_field(field_name):
    lowered = field_name.lower()
    return any(keyword in lowered for keyword in IMAGE_FIELD_KEYWORDS)


def _clean_value(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "nan"}:
        return ""
    return text


def _build_product_rules(stats):
    sku_count = stats["sku_count"]
    auto_fields = set(FIELD_MAP)
    fields = []
    fixed_defaults = []
    always_filled = []
    often_filled_unmapped = []

    for field_name, field_stats in stats["fields"].items():
        if field_name in SYSTEM_FIELDS:
            continue
        filled_rows = field_stats["filled_rows"]
        fill_ratio = filled_rows / sku_count if sku_count else 0
        values = field_stats["values"].most_common(MAX_EXAMPLE_VALUES)
        unique_count = len(field_stats["values"])
        top_value = values[0][0] if values else ""
        top_count = values[0][1] if values else 0
        top_ratio = top_count / filled_rows if filled_rows else 0
        mapped = field_name in auto_fields
        item = {
            "field_name": field_name,
            "filled_rows": filled_rows,
            "fill_ratio": round(fill_ratio, 4),
            "template_files": sorted(field_stats["files"]),
            "mapped_by_auto_fill": mapped,
            "unique_value_count": unique_count,
            "example_values": [value for value, _count in values],
        }
        fields.append(item)
        if fill_ratio >= MIN_ALWAYS_FILLED_RATIO:
            always_filled.append(item)
        if fill_ratio >= MIN_OFTEN_FILLED_RATIO and not mapped:
            often_filled_unmapped.append(item)
        if fill_ratio >= MIN_ALWAYS_FILLED_RATIO and top_ratio >= MIN_ALWAYS_FILLED_RATIO and unique_count <= 3:
            fixed_defaults.append({
                "field_name": field_name,
                "value": top_value,
                "filled_rows": filled_rows,
                "fill_ratio": round(fill_ratio, 4),
                "mapped_by_auto_fill": mapped,
            })

    return {
        "template_count": stats["template_count"],
        "sku_count": sku_count,
        "template_files": sorted(stats["template_files"]),
        "average_field_count": round(stats["field_count_total"] / stats["template_count"], 1) if stats["template_count"] else 0,
        "always_filled_fields": _sort_fields(always_filled),
        "fixed_default_fields": sorted(fixed_defaults, key=lambda item: item["field_name"]),
        "often_filled_unmapped_fields": _sort_fields(often_filled_unmapped),
        "field_summary": _sort_fields(fields),
    }


def _sort_fields(fields):
    return sorted(fields, key=lambda item: (-item["fill_ratio"], item["field_name"]))


def _build_report(rules):
    lines = [
        "# 成功模板规则提炼报告",
        "",
        f"- 样板文件数：{rules['template_count']}",
        f"- Product Type 数：{rules['product_type_count']}",
        f"- 跳过文件数：{rules['skipped_count']}",
        "- 图片字段：本阶段已排除，不参与规则建议。",
        "",
        "## 总览",
        "",
        "| Product Type | 样板数 | SKU 数 | 固定默认值 | 常填但未自动映射 |",
        "| --- | ---: | ---: | ---: | ---: |",
    ]
    for product_type, data in rules["product_types"].items():
        lines.append(
            f"| `{product_type}` | {data['template_count']} | {data['sku_count']} | "
            f"{len(data['fixed_default_fields'])} | {len(data['often_filled_unmapped_fields'])} |"
        )

    for product_type, data in rules["product_types"].items():
        lines.extend([
            "",
            f"## {product_type}",
            "",
            f"- 样板文件数：{data['template_count']}",
            f"- SKU 数：{data['sku_count']}",
            f"- 平均字段数：{data['average_field_count']}",
            f"- 样板文件：{', '.join(data['template_files'])}",
            "",
            "### 可考虑固化的固定默认值",
            "",
        ])
        fixed_defaults = data["fixed_default_fields"]
        if fixed_defaults:
            for item in fixed_defaults[:20]:
                mapped = "已映射" if item["mapped_by_auto_fill"] else "未映射"
                lines.append(f"- `{item['field_name']}` = `{item['value']}`（{mapped}）")
        else:
            lines.append("- 暂未识别到稳定固定值。")

        lines.extend([
            "",
            "### 成功样板中总是填写的字段",
            "",
        ])
        always = data["always_filled_fields"]
        if always:
            for item in always[:30]:
                mapped = "已映射" if item["mapped_by_auto_fill"] else "未映射"
                examples = _examples_text(item["example_values"])
                lines.append(f"- `{item['field_name']}`（{mapped}，示例：{examples}）")
        else:
            lines.append("- 暂未识别到总是填写的字段。")

        lines.extend([
            "",
            "### 经常填写但当前自动填表未覆盖",
            "",
        ])
        unmapped = data["often_filled_unmapped_fields"]
        if unmapped:
            for item in unmapped[:30]:
                examples = _examples_text(item["example_values"])
                percent = round(item["fill_ratio"] * 100)
                lines.append(f"- `{item['field_name']}`（填写率 {percent}%，示例：{examples}）")
        else:
            lines.append("- 暂无明显缺口。")

    if rules["skipped"]:
        lines.extend(["", "## 跳过文件", ""])
        for item in rules["skipped"]:
            lines.append(f"- `{item['file']}`：{item['reason']}")

    lines.append("")
    return "\n".join(lines)


def _examples_text(values):
    if not values:
        return "无"
    return ", ".join(f"`{_shorten(value)}`" for value in values[:5])


def _shorten(value):
    text = " ".join(str(value).split())
    if len(text) <= MAX_EXAMPLE_TEXT_LENGTH:
        return text
    return text[:MAX_EXAMPLE_TEXT_LENGTH - 3] + "..."
