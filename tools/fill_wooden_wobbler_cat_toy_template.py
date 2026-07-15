from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName


BASE_DIR = Path("/Users/cc/Desktop/猫玩具不倒翁")
TEMPLATE = BASE_DIR / "PET_TOY.xlsm"
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT = OUTPUT_DIR / "木质不倒翁猫玩具_v2.xlsm"
DESKTOP_OUTPUT = Path("/Users/cc/Desktop/木质不倒翁猫玩具_v2.xlsm")


COMMON_BULLETS = [
    "Self-righting wooden tumbler design rocks back upright after batting, encouraging solo play for indoor cats.",
    "Silent, battery-free motion helps cats chase, paw, and pounce without electronic noise or charging.",
    "Compact 6.3 inch height is easy for kittens and adult cats to bat during supervised daily play.",
    "Rounded beechwood-style base and soft fabric body are designed for light interactive cat enrichment.",
    "Good for boredom relief, exercise, and mental stimulation when your cat needs independent play time.",
]

SKU_ROWS = [
    {
        "sku": "CA-Pettoy-Feather",
        "source_name": "不倒翁羽毛款【天然榉木底座-耐抓耐咬】",
        "title": "Wooden Cat Wobble Toy with Feather Teaser, Self-Righting Tumbler Cat Toy for Indoor Cats, Silent Battery-Free Solo Play",
        "highlight": "Feather Teaser",
        "style": "Whimsical",
        "color": "Multicolor",
        "materials": ["Engineered Wood", "Polyester"],
        "subject": "Cat",
        "theme": "Animals",
        "pet_toy_type": "Teaser Wand",
        "price": 5.49,
        "description_tail": "The feather teaser gives cats a light, fluttering target while the weighted wooden base keeps the toy wobbling back upright.",
    },
    {
        "sku": "CA-Pettoy-Sisal",
        "source_name": "不倒翁剑麻款",
        "title": "Wooden Cat Wobble Toy with Sisal Fabric Body, Self-Righting Tumbler Cat Toy for Indoor Cats, Silent Battery-Free Solo Play",
        "highlight": "Sisal Fabric",
        "style": "Rustic",
        "color": "Natural",
        "materials": ["Engineered Wood", "Jute"],
        "subject": "Cat",
        "theme": "Animals",
        "pet_toy_type": "Scratch Toy",
        "price": 5.51,
        "description_tail": "The sisal-style fabric body adds texture for pawing and scratching play while the rounded wooden base creates a wobbling motion.",
    },
]


def col_map(ws):
    cols = {}
    seen = {}
    for col in range(1, ws.max_column + 1):
        label = ws.cell(4, col).value
        if not label:
            continue
        seen[label] = seen.get(label, 0) + 1
        key = label if seen[label] == 1 else f"{label}.{seen[label] - 1}"
        cols[key] = col
    return cols


def setv(ws, row, cols, label, value):
    col = cols.get(label)
    if col and value is not None:
        ws.cell(row, col).value = value


def clearv(ws, row, cols, label):
    col = cols.get(label)
    if col:
        ws.cell(row, col).value = None


def copy_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


def fill_common(ws, row, cols, item):
    setv(ws, row, cols, "SKU", item["sku"])
    setv(ws, row, cols, "Product Type", "PET_TOY")
    setv(ws, row, cols, "Listing Action", "Create or Replace (Full Update)")
    clearv(ws, row, cols, "Parentage Level")
    clearv(ws, row, cols, "Parent SKU")
    clearv(ws, row, cols, "Variation Theme Name")

    setv(ws, row, cols, "Item Name", item["title"])
    setv(ws, row, cols, "Item Highlight", item["highlight"])
    setv(ws, row, cols, "Brand Name", "Generic")
    setv(ws, row, cols, "Item Type Keyword", "pet-interactive-toys")
    setv(ws, row, cols, "Target Audience Keyword", "Cats")
    setv(ws, row, cols, "Model Number", item["sku"])
    setv(ws, row, cols, "Model Name", "Wooden Wobble Cat Toy")
    setv(ws, row, cols, "Manufacturer", "Generic")

    description = (
        "A wooden self-righting wobble cat toy made for indoor cat enrichment. "
        "The rounded base rocks back upright after cats bat it, creating silent physical motion without batteries. "
        f"{item['description_tail']} "
        "Use during supervised play and replace the toy if it becomes damaged."
    )
    setv(ws, row, cols, "Product Description", description)
    for idx, bullet in enumerate(COMMON_BULLETS, start=1):
        setv(ws, row, cols, "Bullet Point" if idx == 1 else f"Bullet Point.{idx - 1}", bullet)
    setv(ws, row, cols, "Generic Keyword", "wooden cat toy; wobble cat toy; tumbler cat toy; interactive cat toy; indoor cat toy")

    setv(ws, row, cols, "Special Features", "Interactive")
    setv(ws, row, cols, "Special Features.1", "Lightweight")
    setv(ws, row, cols, "Style", item["style"])
    for index, material in enumerate(item["materials"], start=1):
        setv(ws, row, cols, "Material" if index == 1 else f"Material.{index - 1}", material)
    setv(ws, row, cols, "Number of Items", 1)
    setv(ws, row, cols, "Item Package Quantity", 1)
    setv(ws, row, cols, "Subject Character", item["subject"])
    setv(ws, row, cols, "Color", item["color"])
    setv(ws, row, cols, "Size", "Small")
    setv(ws, row, cols, "Part Number", item["sku"])
    setv(ws, row, cols, "Item Shape", "Round")
    setv(ws, row, cols, "Theme", item["theme"])
    setv(ws, row, cols, "Breed Recommendation", "All Breed Sizes")
    setv(ws, row, cols, "Pattern", "Animal Print")
    setv(ws, row, cols, "Unit Count", 1)
    setv(ws, row, cols, "Unit Count Type", "Count")
    setv(ws, row, cols, "Included Components", "Wooden wobble cat toy")
    setv(ws, row, cols, "Specific Uses for Product", "Indoor")
    setv(ws, row, cols, "Specific Uses for Product.1", "Active")
    setv(ws, row, cols, "Recommended Uses For Product", "Boredom Buster")
    setv(ws, row, cols, "Recommended Uses For Product.1", "Exercise")
    setv(ws, row, cols, "Indoor Outdoor Usage", "Indoor")
    setv(ws, row, cols, "Pet Toy Type", item["pet_toy_type"])
    setv(ws, row, cols, "Number of Packs", 1)
    setv(ws, row, cols, "Pet Type", "Cat")

    setv(ws, row, cols, "Item Condition", "New")
    setv(ws, row, cols, "List Price", item["price"])
    setv(ws, row, cols, "Your Price USD (Low-Cost Store, US)", item["price"])
    setv(ws, row, cols, "Product Tax Code", "A_GEN_NOTAX")
    setv(ws, row, cols, "Country of Origin", "China")
    setv(ws, row, cols, "Are batteries required?", "No")
    setv(ws, row, cols, "Are batteries included?", "No")
    setv(ws, row, cols, "Dangerous Goods Regulations", "Not Applicable")
    setv(ws, row, cols, "Contains Liquid Contents?", "No")
    setv(ws, row, cols, "Number of Boxes", 1)
    setv(ws, row, cols, "Directions", "Use under supervision. Replace the toy if it becomes damaged.")

    for label in ["Main Image URL", "Swatch Image URL", "Main Image Location"]:
        clearv(ws, row, cols, label)


def fill_dimensions(ws, row, cols):
    length = 2.36
    width = 2.36
    height = 6.3
    weight = 0.082
    for label, value in [
        ("Item Display Height", height),
        ("Item Display Length", length),
        ("Item Display Width", width),
        ("Height base to top", height),
        ("Length longer horizontal edge", length),
        ("Width shorter horizontal edge", width),
        ("Item Package Length", length),
        ("Item Package Width", width),
        ("Item Package Height", height),
    ]:
        setv(ws, row, cols, label, value)
    for label in [
        "Item Display Height Unit",
        "Item Display Length Unit",
        "Item Display Width Unit",
        "Height Unit",
        "Length Unit",
        "Width Unit",
        "Package Length Unit",
        "Package Width Unit",
        "Package Height Unit",
    ]:
        setv(ws, row, cols, label, "Inches")
    setv(ws, row, cols, "Item Weight", weight)
    setv(ws, row, cols, "Item Weight Unit", "Pounds")
    setv(ws, row, cols, "Package Weight", weight)
    setv(ws, row, cols, "Package Weight Unit", "Pounds")


def allow_generic_brand(wb):
    dropdown = wb["Dropdown Lists"]
    dropdown["G5"] = "Generic"
    name = "PET_TOYbrandmarketplace_idATVPDKIKX0DERlanguage_tagen_US1.value"
    wb.defined_names[name] = DefinedName(
        name=name,
        attr_text="'Dropdown Lists'!$G$4:$G$5",
    )


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(TEMPLATE, keep_vba=True)
    allow_generic_brand(wb)
    ws = wb["Template"]
    cols = col_map(ws)

    for row in range(7, 9):
        copy_row_style(ws, 6, row)

    for offset, item in enumerate(SKU_ROWS):
        row = 7 + offset
        fill_common(ws, row, cols, item)
        fill_dimensions(ws, row, cols)

    wb.save(OUTPUT)
    wb.save(DESKTOP_OUTPUT)
    print(OUTPUT)
    print(DESKTOP_OUTPUT)


if __name__ == "__main__":
    main()
