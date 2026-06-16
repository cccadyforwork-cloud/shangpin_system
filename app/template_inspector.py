from pathlib import Path

from openpyxl import load_workbook

from .paths import OUTPUTS_DIR, safe_name


def _text(value):
    if value is None:
        return ""
    return str(value).strip()


def inspect_template(template_path, output_path=None):
    template_path = Path(template_path)
    wb = load_workbook(template_path, data_only=True, read_only=True, keep_vba=template_path.suffix.lower() == ".xlsm")
    findings = {
        "template_path": str(template_path),
        "sheetnames": wb.sheetnames,
        "product_type": "",
        "browse_node": "",
        "required_fields": [],
        "template_columns": []
    }

    if "Data Definitions" in wb.sheetnames:
        ws = wb["Data Definitions"]
        for row in ws.iter_rows(min_row=1, values_only=True):
            values = [_text(value) for value in row]
            if len(values) < 6:
                continue
            field_name = values[1]
            local_label = values[2]
            required = values[5]
            if required.lower() == "required" and field_name:
                findings["required_fields"].append({
                    "field_name": field_name,
                    "local_label": local_label
                })
            if field_name == "product_type#1.value" and not findings["product_type"]:
                example = values[4]
                if example:
                    findings["product_type"] = example

    if "Valid Values" in wb.sheetnames:
        ws = wb["Valid Values"]
        for row in ws.iter_rows(values_only=True):
            values = [_text(value) for value in row]
            for idx, value in enumerate(values):
                if value.startswith("Product Type"):
                    product_type = next((item for item in values[idx + 1:] if item), "")
                    if product_type:
                        findings["product_type"] = product_type
                        break
            if findings["product_type"] and findings["product_type"] != "ACCESSORY":
                break

    if "Template" in wb.sheetnames:
        ws = wb["Template"]
        settings = _text(ws.cell(1, 1).value)
        marker = "ptds="
        if marker in settings:
            import base64
            from urllib.parse import parse_qs

            query = parse_qs(settings)
            encoded_values = query.get("ptds")
            if encoded_values:
                try:
                    findings["product_type"] = base64.b64decode(encoded_values[0]).decode("utf-8")
                except Exception:
                    pass

    if findings["product_type"] == "ACCESSORY" and "EXERCISE_BLOCK" in template_path.stem.upper():
        findings["product_type"] = "EXERCISE_BLOCK"

    if "Browse Data" in wb.sheetnames:
        ws = wb["Browse Data"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                findings["browse_node"] = _text(row[0])
                break

    if "Template" in wb.sheetnames:
        ws = wb["Template"]
        labels = [_text(cell.value) for cell in ws[4]]
        field_names = [_text(cell.value) for cell in ws[5]]
        for idx, field_name in enumerate(field_names, 1):
            if field_name:
                findings["template_columns"].append({
                    "column": idx,
                    "label": labels[idx - 1] if idx - 1 < len(labels) else "",
                    "field_name": field_name
                })

    if output_path is None:
        output_path = OUTPUTS_DIR / f"{safe_name(template_path.stem)}_模板字段扫描报告.md"
    else:
        output_path = Path(output_path)
    _write_report(output_path, findings)
    return findings, output_path


def _write_report(path, findings):
    required = findings["required_fields"]
    lines = [
        f"# {Path(findings['template_path']).name} 模板字段扫描报告",
        "",
        f"- Product Type：{findings['product_type'] or '未识别'}",
        f"- Browse Node / Item Type Keyword：{findings['browse_node'] or '未识别'}",
        f"- 工作表：{', '.join(findings['sheetnames'])}",
        f"- Data Definitions 必填字段数：{len(required)}",
        "",
        "## 必填字段",
        ""
    ]
    if not required:
        lines.append("未从 Data Definitions 识别到 Required 字段。")
    else:
        for item in required:
            lines.append(f"- {item['local_label'] or item['field_name']}：`{item['field_name']}`")

    lines.extend([
        "",
        "## 后续处理",
        "",
        "- 能从产品资料草稿映射的字段，后续可以自动写入 Template 页。",
        "- 条件必填字段需要结合 Valid Values、Conditions List 和报错报告继续补规则。",
        "- 无法稳定识别的字段保留给人工确认。",
        ""
    ])
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")
