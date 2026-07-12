from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName


BASE_DIR = Path("/Users/cc/Desktop/咖啡毛刷")
TEMPLATE = BASE_DIR / "CLEANING_BRUSH.xlsm"
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT = OUTPUT_DIR / "咖啡清洁毛刷_v1.xlsm"
DESKTOP_OUTPUT = Path("/Users/cc/Desktop/咖啡清洁毛刷_v1.xlsm")


ROWS = [
    {
        "sku": "CA-Coffeebrush-Blackwood",
        "wood_cn": "黑檀木",
        "title_color": "Ebony Wood",
        "color": "Dark Brown",
        "handle": "Ebony Wood",
    },
    {
        "sku": "CA-Coffeebrush-Pearwood",
        "wood_cn": "花梨木",
        "title_color": "Pearwood",
        "color": "Brown",
        "handle": "Pearwood",
    },
]


COMMON_BULLETS = [
    "Coffee grinder cleaning brush helps sweep coffee grounds, dust, and residue from grinders, espresso machine parts, portafilters, counters, and coffee bar tools.",
    "Natural pig-hair style bristles are soft enough for daily coffee accessory cleaning while still gathering fine powder from small gaps and corners.",
    "Solid wood handle with brass-tone metal collar gives the brush a classic barista-tool look and a comfortable hand feel for home or cafe use.",
    "Compact 7.09 inch brush is easy to keep near a grinder, espresso station, kitchen drawer, or travel coffee kit without taking much space.",
    "Reusable manual cleaning brush needs no batteries and is suitable for barista workstations, home kitchens, coffee shops, and office coffee areas.",
]


def build_col_maps(ws):
    labels = {}
    fields = {}
    seen = {}
    for col in range(1, ws.max_column + 1):
        label = ws.cell(4, col).value
        field = ws.cell(5, col).value
        if label:
            seen[label] = seen.get(label, 0) + 1
            key = label if seen[label] == 1 else f"{label}.{seen[label] - 1}"
            labels[key] = col
        if field:
            fields[field] = col
    return labels, fields


def setv(ws, row, cols, label, value):
    col = cols.get(label)
    if col and value not in (None, ""):
        ws.cell(row, col).value = value


def clearv(ws, row, cols, label):
    col = cols.get(label)
    if col:
        ws.cell(row, col).value = None


def copy_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


def allow_generic_brand(wb):
    dropdown = wb["Dropdown Lists"]
    dropdown["G5"] = "Generic"
    name = "CLEANING_BRUSHbrandmarketplace_idATVPDKIKX0DERlanguage_tagen_US1.value"
    wb.defined_names[name] = DefinedName(
        name=name,
        attr_text="'Dropdown Lists'!$G$4:$G$5",
    )


def fill_row(ws, row, cols, item):
    sku = item["sku"]
    title_color = item["title_color"]
    title = (
        "Coffee Grinder Cleaning Brush with Wooden Handle, Natural Bristles "
        f"Espresso Machine Barista Tool for Home Kitchen Coffee Bar, {title_color}"
    )
    description = (
        f"This coffee grinder cleaning brush is a compact manual tool for sweeping coffee grounds and fine powder from grinders, espresso machine parts, portafilters, counters, and coffee bar accessories. "
        f"It has a solid wood handle, brass-tone metal collar, and soft natural bristles for everyday home kitchen, cafe, and barista workstation cleaning. "
        f"Handle style: {title_color}."
    )

    setv(ws, row, cols, "SKU", sku)
    setv(ws, row, cols, "Product Type", "CLEANING_BRUSH")
    setv(ws, row, cols, "Listing Action", "Create or Replace (Full Update)")
    clearv(ws, row, cols, "Parentage Level")
    clearv(ws, row, cols, "Parent SKU")
    clearv(ws, row, cols, "Variation Theme Name")

    setv(ws, row, cols, "Item Name", title)
    setv(ws, row, cols, "Item Highlight", f"{title_color}, 7.09 in")
    setv(ws, row, cols, "Brand Name", "Generic")
    setv(ws, row, cols, "Item Type Keyword", "家居、厨具、家装 > 家庭清洁用品 > 清洁刷 (cleaning-brushes)")
    setv(ws, row, cols, "Package Level", "Unit")
    setv(ws, row, cols, "Model Number", sku)
    setv(ws, row, cols, "Model Name", "Coffee Grinder Cleaning Brush")
    setv(ws, row, cols, "Manufacturer", "Generic")

    # Image fields intentionally left blank: available source images contain extra text.
    setv(ws, row, cols, "Product Description", description)
    for idx, bullet in enumerate(COMMON_BULLETS, start=1):
        label = "Bullet Point" if idx == 1 else f"Bullet Point.{idx - 1}"
        setv(ws, row, cols, label, bullet)
    setv(ws, row, cols, "Generic Keyword", "coffee grinder cleaning brush; espresso machine brush; barista brush; coffee powder brush; wooden handle brush; grinder dust brush")

    for idx, feature in enumerate(["Lightweight", "Portable", "Reusable", "Grip Handle"], start=1):
        label = "Special Features" if idx == 1 else f"Special Features.{idx - 1}"
        setv(ws, row, cols, label, feature)

    setv(ws, row, cols, "Style", "Classic")
    setv(ws, row, cols, "Material", "Wood")
    setv(ws, row, cols, "Material.1", "Copper")
    setv(ws, row, cols, "Material.2", "Pig Hair")
    setv(ws, row, cols, "Number of Items", 1)
    setv(ws, row, cols, "Item Package Quantity", 1)
    setv(ws, row, cols, "Color", item["color"])
    setv(ws, row, cols, "Size", "7.09 in")
    setv(ws, row, cols, "Part Number", sku)
    setv(ws, row, cols, "Item Shape", "Round")
    setv(ws, row, cols, "Surface Recommendation", "Metal")
    setv(ws, row, cols, "Surface Recommendation.1", "Wood")
    setv(ws, row, cols, "Handle Material", item["handle"])
    setv(ws, row, cols, "Item Firmness Description", "Soft")
    setv(ws, row, cols, "Pattern", "Solid")
    setv(ws, row, cols, "Unit Count", 1)
    setv(ws, row, cols, "Unit Count Type", "Count")
    setv(ws, row, cols, "Specific Uses for Product", "Counter")
    setv(ws, row, cols, "Specific Uses for Product.1", "Cooker")
    setv(ws, row, cols, "Bristle Material", "Pig Hair")

    setv(ws, row, cols, "Height base to top", 0.59)
    setv(ws, row, cols, "Height Unit", "Inches")
    setv(ws, row, cols, "Length longer horizontal edge", 7.09)
    setv(ws, row, cols, "Length Unit", "Inches")
    setv(ws, row, cols, "Width shorter horizontal edge", 1.57)
    setv(ws, row, cols, "Width Unit", "Inches")
    setv(ws, row, cols, "Item Width", 1.57)
    setv(ws, row, cols, "Width Unit.1", "Inches")
    setv(ws, row, cols, "Item Length", 7.09)
    setv(ws, row, cols, "Item Length Unit", "Inches")
    setv(ws, row, cols, "Number of Packs", 1)
    setv(ws, row, cols, "Item Weight", 0.066)
    setv(ws, row, cols, "Item Weight Unit", "Pounds")

    setv(ws, row, cols, "Item Condition", "New")
    setv(ws, row, cols, "List Price", 2.49)
    setv(ws, row, cols, "Product Tax Code", "A_GEN_NOTAX")
    setv(ws, row, cols, "Your Price USD (Low-Cost Store, US)", 2.49)
    setv(ws, row, cols, "Item Package Length", 7.09)
    setv(ws, row, cols, "Package Length Unit", "Inches")
    setv(ws, row, cols, "Item Package Width", 1.57)
    setv(ws, row, cols, "Package Width Unit", "Inches")
    setv(ws, row, cols, "Item Package Height", 0.59)
    setv(ws, row, cols, "Package Height Unit", "Inches")
    setv(ws, row, cols, "Package Weight", 0.066)
    setv(ws, row, cols, "Package Weight Unit", "Pounds")
    setv(ws, row, cols, "Country of Origin", "China")
    setv(ws, row, cols, "Are batteries required?", "No")
    setv(ws, row, cols, "Are batteries included?", "No")
    setv(ws, row, cols, "Dangerous Goods Regulations", "Not Applicable")
    setv(ws, row, cols, "Product Compliance Certificate", "Not Applicable")
    setv(ws, row, cols, "Compliance - Brush Intended Use", "Cleaning")
    setv(ws, row, cols, "Compliance - Is Motorized", "No")
    setv(ws, row, cols, "Compliance - Is Mechanical", "No")
    setv(ws, row, cols, "Compliance - Bristle Material", "Other")
    setv(ws, row, cols, "Compliance - Is Hand-Operated", "Yes")


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(TEMPLATE, keep_vba=True)
    allow_generic_brand(wb)
    ws = wb["Template"]
    cols, _fields = build_col_maps(ws)

    for row in range(7, 9):
        copy_row_style(ws, 6, row)

    for offset, item in enumerate(ROWS):
        fill_row(ws, 7 + offset, cols, item)

    wb.save(OUTPUT)
    wb.save(DESKTOP_OUTPUT)
    print(OUTPUT)
    print(DESKTOP_OUTPUT)


if __name__ == "__main__":
    main()
