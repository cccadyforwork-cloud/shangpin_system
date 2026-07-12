from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName


BASE_DIR = Path("/Users/cc/Desktop/防水eva包")
TEMPLATE = BASE_DIR / "COSMETIC_CASE.xlsm"
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT = OUTPUT_DIR / "防水洗漱包_v1.xlsm"
DESKTOP_OUTPUT = Path("/Users/cc/Desktop/防水洗漱包_v1.xlsm")


ROWS = [
    {
        "sku": "CA-Washbag-Grey",
        "source_name": "透明灰色✪大号",
        "title_color": "Transparent Gray",
        "color": "Transparent Gray",
        "color_map": "Grey",
        "image": "https://cbu01.alicdn.com/img/ibank/O1CN01SU44DB23QYVulE9w4_!!2218675367250-0-cib.jpg",
    },
    {
        "sku": "CA-Washbag-White",
        "source_name": "透明白色✪大号",
        "title_color": "Transparent White",
        "color": "Transparent White",
        "color_map": "White",
        "image": "https://cbu01.alicdn.com/img/ibank/O1CN01brlVIt23QYVvNzzJ6_!!2218675367250-0-cib.jpg",
    },
]


COMMON_BULLETS = [
    "Waterproof EVA and PVC toiletry bag helps separate damp towels, swimwear, toiletries, cosmetics, and small travel items from dry belongings.",
    "Large portable pouch is useful for travel, gym, swimming, beach days, outdoor activities, bathroom storage, and everyday organizing.",
    "Transparent body makes contents easy to see at a glance, so makeup, shampoo bottles, skincare, and accessories are quicker to find.",
    "Drawstring closure helps keep items gathered inside the pouch while remaining simple to open, close, carry, and pack flat.",
    "Lightweight flexible material is easy to wipe clean and reuse, with a simple solid-color look for men, women, and shared travel use.",
]


def build_col_maps(ws):
    labels = {}
    fields = {}
    seen = {}
    for col in range(1, ws.max_column + 1):
        label = ws.cell(4, col).value
        field = ws.cell(5, col).value
        if label:
            seen[label] = seen.get(label, 0) + 1
            key = label if seen[label] == 1 else f"{label}.{seen[label] - 1}"
            labels[key] = col
        if field:
            fields[field] = col
    return labels, fields


def set_label(ws, row, labels, label, value):
    col = labels.get(label)
    if col and value not in (None, ""):
        ws.cell(row, col).value = value


def clear_label(ws, row, labels, label):
    col = labels.get(label)
    if col:
        ws.cell(row, col).value = None


def copy_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


def allow_generic_brand(wb):
    dropdown = wb["Dropdown Lists"]
    dropdown["G5"] = "Generic"
    name = "COSMETIC_CASEbrandmarketplace_idATVPDKIKX0DERlanguage_tagen_US1.value"
    wb.defined_names[name] = DefinedName(
        name=name,
        attr_text="'Dropdown Lists'!$G$4:$G$5",
    )


def fill_row(ws, row, labels, item):
    sku = item["sku"]
    title_color = item["title_color"]
    title = (
        "Waterproof EVA Toiletry Bag Large, Portable Dry Wet Separation "
        f"Travel Makeup Organizer Pouch for Gym Swimming Beach, {title_color}"
    )
    description = (
        f"This large waterproof toiletry bag is made for travel, gym, swimming, beach, and bathroom organization. "
        f"The flexible EVA and PVC material helps keep damp items separated from dry belongings, while the transparent body makes cosmetics, toiletries, skincare, and small accessories easy to identify. "
        f"The drawstring closure keeps items gathered for simple packing and carrying. Color: {title_color}."
    )

    set_label(ws, row, labels, "SKU", sku)
    set_label(ws, row, labels, "Product Type", "COSMETIC_CASE")
    set_label(ws, row, labels, "Listing Action", "Create or Replace (Full Update)")
    clear_label(ws, row, labels, "Parentage Level")
    clear_label(ws, row, labels, "Parent SKU")
    clear_label(ws, row, labels, "Variation Theme Name")

    set_label(ws, row, labels, "Item Name", title)
    set_label(ws, row, labels, "Item Highlight", f"{title_color}, Large")
    set_label(ws, row, labels, "Brand Name", "Generic")
    set_label(ws, row, labels, "Item Type Keyword", "时尚 > 皮具箱包 > 旅行配件 > 洗漱包 (toiletry-bags)")
    set_label(ws, row, labels, "Target Audience", "Unisex-Adult")
    set_label(ws, row, labels, "Model Number", sku)
    set_label(ws, row, labels, "Model Name", "Waterproof Toiletry Bag")
    set_label(ws, row, labels, "Manufacturer", "Generic")
    set_label(ws, row, labels, "Main Image URL", item["image"])
    set_label(ws, row, labels, "Swatch Image URL", item["image"])

    set_label(ws, row, labels, "Product Description", description)
    for idx, bullet in enumerate(COMMON_BULLETS, start=1):
        label = "Bullet Point" if idx == 1 else f"Bullet Point.{idx - 1}"
        set_label(ws, row, labels, label, bullet)
    set_label(ws, row, labels, "Generic Keyword", "waterproof toiletry bag; dry wet separation bag; travel makeup pouch; gym swim bag; EVA wash bag; cosmetic organizer")

    for idx, feature in enumerate(["Leak Resistant", "Stain Resistant", "Tear Resistant"], start=1):
        label = "Special Features" if idx == 1 else f"Special Features.{idx - 1}"
        set_label(ws, row, labels, label, feature)

    set_label(ws, row, labels, "Style", "Modern")
    set_label(ws, row, labels, "Department Name", "Unisex")
    set_label(ws, row, labels, "Target Gender", "Unisex")
    set_label(ws, row, labels, "Material Type", "Ethylene Vinyl Acetate")
    set_label(ws, row, labels, "Material Type.1", "Polyvinyl Chloride")
    set_label(ws, row, labels, "Number of Items", 1)
    set_label(ws, row, labels, "Item Package Quantity", 1)
    set_label(ws, row, labels, "Water Resistance Level", "Waterproof")
    set_label(ws, row, labels, "Color Map", item["color_map"])
    set_label(ws, row, labels, "Color", item["color"])
    set_label(ws, row, labels, "Size", "Large")
    set_label(ws, row, labels, "Part Number", sku)
    set_label(ws, row, labels, "Item Shape", "Rectangular")
    set_label(ws, row, labels, "Material Features", "Reusable")
    set_label(ws, row, labels, "Form Factor", "Bag")
    set_label(ws, row, labels, "Pattern", "Solid")
    set_label(ws, row, labels, "Unit Count", 1)
    set_label(ws, row, labels, "Unit Count Type", "Count")
    set_label(ws, row, labels, "Recommended Uses For Product", "Travel")
    set_label(ws, row, labels, "Recommended Uses For Product.1", "Outdoors")
    set_label(ws, row, labels, "Recommended Uses For Product.2", "Home")
    set_label(ws, row, labels, "Closure Type", "Drawstring")

    set_label(ws, row, labels, "Height base to top", 1.57)
    set_label(ws, row, labels, "Height Unit", "Inches")
    set_label(ws, row, labels, "Length longer horizontal edge", 6.3)
    set_label(ws, row, labels, "Length Unit", "Inches")
    set_label(ws, row, labels, "Width shorter horizontal edge", 4.72)
    set_label(ws, row, labels, "Width Unit", "Inches")
    set_label(ws, row, labels, "Number of Packs", 1)
    set_label(ws, row, labels, "Item Weight", 0.193)
    set_label(ws, row, labels, "Item Weight Unit", "Pounds")

    set_label(ws, row, labels, "Item Condition", "New")
    set_label(ws, row, labels, "List Price", 4.99)
    set_label(ws, row, labels, "Product Tax Code", "A_GEN_NOTAX")
    set_label(ws, row, labels, "Main Image Location", item["image"])
    set_label(ws, row, labels, "Your Price USD (Low-Cost Store, US)", 4.99)
    set_label(ws, row, labels, "Item Package Length", 6.3)
    set_label(ws, row, labels, "Package Length Unit", "Inches")
    set_label(ws, row, labels, "Item Package Width", 4.72)
    set_label(ws, row, labels, "Package Width Unit", "Inches")
    set_label(ws, row, labels, "Item Package Height", 1.57)
    set_label(ws, row, labels, "Package Height Unit", "Inches")
    set_label(ws, row, labels, "Package Weight", 0.193)
    set_label(ws, row, labels, "Package Weight Unit", "Pounds")
    set_label(ws, row, labels, "Country of Origin", "China")
    set_label(ws, row, labels, "Are batteries required?", "No")
    set_label(ws, row, labels, "Are batteries included?", "No")
    set_label(ws, row, labels, "Dangerous Goods Regulations", "Not Applicable")


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(TEMPLATE, keep_vba=True)
    allow_generic_brand(wb)
    ws = wb["Template"]
    labels, _fields = build_col_maps(ws)

    for row in range(7, 9):
        copy_row_style(ws, 6, row)

    for offset, item in enumerate(ROWS):
        fill_row(ws, 7 + offset, labels, item)

    wb.save(OUTPUT)
    wb.save(DESKTOP_OUTPUT)
    print(OUTPUT)
    print(DESKTOP_OUTPUT)


if __name__ == "__main__":
    main()
