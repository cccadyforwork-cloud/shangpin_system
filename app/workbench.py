import cgi
import base64
import json
import re
import shutil
import subprocess
import threading
import webbrowser
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlparse

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .analyzer import analyze_project
from .auto_fill import auto_fill_project
from .error_learning import learn_reports
from .paths import (
    COMPETITOR_DIR,
    DRAFT_DIRS,
    LEGACY_COMPETITOR_DIR,
    LEGACY_FILLED_TEMPLATE_DIR,
    LEGACY_TEMPLATE_SOURCE_DIR,
    OUTPUTS_DIR,
    PACKAGING_PRICING_DIR,
    PRODUCT_DETAIL_DIR,
    PROJECTS_DIR,
    PROJECT_FOLDERS,
    PURCHASE_DIR,
    ROOT,
    TEMPLATE_DIRS,
    TEMPLATE_SOURCE_DIR,
    ensure_base_dirs,
    safe_name,
)
from .project_manager import create_project, delete_project, list_project_summaries
from .project_status import infer_latest_template, infer_product_name, infer_sku_count, load_project_status, mark_uploaded_success, save_project_status
from .report_parser import extract_processing_summary
from .success_templates import RULES_JSON, RULES_REPORT, SUCCESS_TEMPLATES_DIR, learn_success_templates
from .template_sheet import find_template_sheet
from .template_writer import FIELD_MAP, extract_template_product_type, find_template, prepare_variation_rows
from .validator import validate_intake
from .versioning import next_template_version as shared_next_template_version
from .versioning import version_label as shared_version_label
from .versioning import version_number as shared_version_number
from .versioning import versioned_template_path as shared_versioned_template_path
from .workbook_io import INTAKE_HEADERS, REQUIRED_CORE_FIELDS, read_intake_rows, write_intake_workbook


STATUS_LABELS = {
    "uploaded_success": "已上传",
    "ready_for_upload": "待上传",
    "upload_failed_pending_confirmation": "待确认修正",
    "needs_manual_fix": "待修正",
    "blocked": "卡住",
    "not_started": "未开始",
}

PARENT_OPTIONAL_CORE_FIELDS = {
    "list_price",
    "package_length_in",
    "package_width_in",
    "package_height_in",
    "package_weight_lb",
    "batteries_required",
    "dangerous_goods",
}
CJK_RE = re.compile(r"[\u4e00-\u9fff]")
COPY_FIELDS = ["title", "bullet_1", "bullet_2", "bullet_3", "bullet_4", "bullet_5", "description"]

READABLE_SUFFIXES = {".txt", ".md", ".csv", ".tsv", ".html", ".htm", ".xlsx", ".xlsm", ".pdf"}
LEGACY_DISPLAY_FOLDERS = ["02_原始图片", "03_竞品参考", "04_模板原件", "05_填表版本", "06_处理报告", "07_上架备注"]
FEEDBACK_REPORT_DIR = "06_处理报告"
SOURCE_REVIEW_FOLDERS = {
    "01_采购资料",
    "02_产品包装和定价",
    "03_产品详情页",
    "04_竞品参考",
    "03_竞品参考",
    "07_上架备注",
}
LEGACY_SOURCE_FOLDERS = {"02_原始图片", "03_竞品参考", "04_模板原件", "07_上架备注"}


def run_workbench(host="127.0.0.1", port=8766, open_browser=True):
    ensure_base_dirs()
    server = ThreadingHTTPServer((host, port), _handler())
    url = f"http://{host}:{port}"
    print(f"上品工作台已启动：{url}")
    print("按 Ctrl+C 停止。")
    if open_browser:
        threading.Timer(0.5, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n工作台已停止。")
    finally:
        server.server_close()


def _handler():
    class WorkbenchHandler(BaseHTTPRequestHandler):
        def do_HEAD(self):
            parsed = urlparse(self.path)
            if parsed.path in {"/", "/api/summary", "/api/workbench", "/api/rules", "/api/report"}:
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8" if parsed.path == "/" else "application/json; charset=utf-8")
                self.end_headers()
            elif parsed.path == "/file":
                path = _served_file_path(parsed.query)
                if path.exists() and path.is_file():
                    self.send_response(200)
                    self.send_header("Content-Type", _download_content_type(path))
                    self.send_header("Content-Length", str(path.stat().st_size))
                    self.end_headers()
                else:
                    self.send_response(404)
                    self.end_headers()
            else:
                self.send_response(404)
                self.end_headers()

        def do_GET(self):
            parsed = urlparse(self.path)
            if parsed.path == "/":
                self._send_html(_html())
            elif parsed.path == "/api/summary":
                self._send_json(_summary_payload())
            elif parsed.path == "/api/workbench":
                self._send_json(_workbench_payload())
            elif parsed.path == "/api/rules":
                self._send_json(_rules_payload())
            elif parsed.path == "/api/report":
                self._send_json(_report_payload())
            elif parsed.path == "/file":
                self._send_file(_served_file_path(parsed.query))
            else:
                self._send_json({"error": "not_found"}, status=404)

        def do_POST(self):
            parsed = urlparse(self.path)
            try:
                if parsed.path == "/api/upload-files":
                    result = self._handle_file_upload()
                elif parsed.path == "/api/upload-feedback":
                    result = self._handle_feedback_upload()
                else:
                    payload = self._read_json()
                    if parsed.path == "/api/projects":
                        result = _create_project(payload)
                    elif parsed.path == "/api/rename-project":
                        result = _rename_project(payload)
                    elif parsed.path == "/api/analyze-project":
                        result = _analyze_project(payload)
                    elif parsed.path == "/api/save-intake":
                        result = _save_intake(payload)
                    elif parsed.path == "/api/fill-template":
                        result = _fill_template_version(payload)
                    elif parsed.path == "/api/auto-fill":
                        result = _auto_fill(payload)
                    elif parsed.path == "/api/mark-uploaded":
                        result = _mark_uploaded(payload)
                    elif parsed.path == "/api/confirm-feedback-fix":
                        result = _confirm_feedback_fix(payload)
                    elif parsed.path == "/api/delete-project":
                        result = _delete_project(payload)
                    elif parsed.path == "/api/reveal-file":
                        result = _reveal_file(payload)
                    elif parsed.path == "/api/learn-success":
                        result = _learn_success()
                    else:
                        self._send_json({"error": "not_found"}, status=404)
                        return
                self._send_json(result)
            except Exception as exc:
                self._send_json({"ok": False, "error": str(exc)}, status=500)

        def _handle_file_upload(self):
            form = self._read_multipart_form()
            project_dir = form.getfirst("project_dir", "")
            folder = form.getfirst("folder", "")
            files = _form_files(form)
            return _upload_files(project_dir, folder, files)

        def _handle_feedback_upload(self):
            form = self._read_multipart_form()
            project_dir = form.getfirst("project_dir", "")
            files = _form_files(form)
            note = form.getfirst("note", "")
            return _upload_feedback(project_dir, files, note=note)

        def _read_multipart_form(self):
            content_type = self.headers.get("Content-Type", "")
            if not content_type.startswith("multipart/form-data"):
                raise ValueError("上传请求格式不正确。")
            return cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": content_type,
                    "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                },
            )

        def log_message(self, _format, *_args):
            return

        def _read_json(self):
            length = int(self.headers.get("Content-Length", "0"))
            if not length:
                return {}
            return json.loads(self.rfile.read(length).decode("utf-8"))

        def _send_html(self, text, status=200):
            data = text.encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)

        def _send_json(self, payload, status=200):
            data = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)

        def _send_file(self, path):
            if not path.exists() or not path.is_file():
                self._send_json({"error": "file_not_found"}, status=404)
                return
            self.send_response(200)
            self.send_header("Content-Type", _download_content_type(path))
            self.send_header("Content-Length", str(path.stat().st_size))
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(path.name)}")
            self.end_headers()
            with path.open("rb") as source:
                shutil.copyfileobj(source, self.wfile)

    return WorkbenchHandler


def _summary_payload():
    projects = []
    counts = {}
    for item in list_project_summaries():
        status = item["status"]
        counts[status] = counts.get(status, 0) + 1
        project_dir = item["project_dir"]
        latest_draft = _file_info(project_dir, item.get("latest_draft") or _infer_latest_draft(project_dir))
        latest_template = _file_info(project_dir, item.get("latest_template"))
        source_template = _file_info(project_dir, item.get("source_template"))
        latest_check_report = _file_info(project_dir, item.get("latest_check_report"))
        folders = _project_folder_infos(project_dir)
        health = _workflow_health(project_dir, item, latest_draft, latest_template, latest_check_report, folders)
        projects.append({
            "folder": item["folder"],
            "product_name": item["product_name"],
            "status": status,
            "status_label": STATUS_LABELS.get(status, status),
            "sku_count": item["sku_count"] or "",
            "template_error_count": item["template_error_count"],
            "latest_draft": item.get("latest_draft", ""),
            "latest_draft_file": latest_draft,
            "latest_template": item["latest_template"],
            "latest_template_file": latest_template,
            "source_template": item.get("source_template", ""),
            "source_template_file": source_template,
            "latest_check_report": item.get("latest_check_report", ""),
            "latest_check_report_file": latest_check_report,
            "folders": folders,
            "updated_at": item["updated_at"][:10] if item["updated_at"] else "",
            "uploaded_at": item.get("uploaded_at", ""),
            "notes": item.get("notes", ""),
            "blocked_reason": item["blocked_reason"],
            "project_dir": str(project_dir),
            "relative_dir": _relative(project_dir),
            "next_step": _next_step(item),
            "has_template": _has_template(project_dir),
            "has_draft": _has_draft(project_dir),
            "health": health,
        })
    return {
        "root": str(ROOT),
        "projects_dir": str(PROJECTS_DIR),
        "outputs_dir": str(OUTPUTS_DIR),
        "success_templates_dir": str(SUCCESS_TEMPLATES_DIR),
        "project_folders": PROJECT_FOLDERS,
        "counts": counts,
        "projects": projects,
        "totals": {
            "projects": len(projects),
            "uploaded": counts.get("uploaded_success", 0),
            "ready": counts.get("ready_for_upload", 0),
            "needs_fix": counts.get("needs_manual_fix", 0),
            "blocked": counts.get("blocked", 0),
            "not_started": counts.get("not_started", 0),
        },
    }


def _workbench_payload():
    summary = _summary_payload()
    projects = []
    for item in summary["projects"]:
        project_dir = Path(item["project_dir"])
        projects.append(_workbench_project_payload(project_dir, item))
    return {
        "ok": True,
        "root": summary["root"],
        "projects_dir": summary["projects_dir"],
        "project_folders": summary["project_folders"],
        "projects": projects,
        "totals": summary["totals"],
    }


def _workbench_project_payload(project_dir, item):
    item = {**item, **load_project_status(project_dir)}
    status = item.get("status") or "not_started"
    product_name = item.get("product_name") or infer_product_name(project_dir, "")
    rows, draft_file, rows_error = _latest_intake_rows(project_dir, item)
    source_template = _find_source_template(project_dir)
    latest_template = _latest_filled_template(project_dir, item)
    latest_report = _latest_check_report(project_dir, item, latest_template)
    version_files = _template_version_files(project_dir)
    failure_reports = _folder_file_infos(project_dir, [FEEDBACK_REPORT_DIR], suffixes={".txt", ".md", ".csv", ".tsv", ".html", ".htm", ".xlsx", ".xlsm", ".pdf"})
    check_findings = _parse_check_report(str(latest_report)) if latest_report else []
    error_count = len(check_findings)
    next_version = _next_template_version(project_dir)
    uploaded = status == "uploaded_success"
    latest_template_info = _file_info(project_dir, latest_template) if latest_template else None
    latest_report_info = _file_info(project_dir, latest_report) if latest_report else None
    source_template_info = _file_info(project_dir, source_template) if source_template else None

    return {
        "id": _relative(project_dir),
        "project_dir": str(project_dir),
        "relative_dir": _relative(project_dir),
        "name": product_name,
        "folder": project_dir.name,
        "status": "已上传" if uploaded else "等待中",
        "status_code": status,
        "updated": _format_updated(item.get("updated_at")),
        "summary": _project_summary_text(status, latest_template, error_count, next_version, item),
        "steps": _project_steps(status, rows, latest_template, error_count),
        "intake": _intake_module_payload(project_dir, product_name, rows, draft_file, rows_error, item),
        "template": _template_module_payload(
            project_dir,
            product_name,
            rows,
            source_template_info,
            latest_template_info,
            latest_report_info,
            check_findings,
            next_version,
            item,
        ),
        "feedback": _feedback_module_payload(
            project_dir,
            product_name,
            status,
            latest_template_info,
            latest_report_info,
            version_files,
            failure_reports,
            next_version,
            item,
        ),
    }


def _intake_module_payload(project_dir, product_name, rows, draft_file, rows_error, item):
    first = rows[0] if rows else _blank_intake_row(project_dir, product_name)
    issues = _intake_issues(rows, rows_error)
    return {
        "productName": first.get("product_name") or product_name,
        "draftFile": _file_info(project_dir, draft_file) if draft_file else None,
        "reportFile": item.get("latest_draft_file"),
        "rows": _json_rows(rows or [first]),
        "files": [
            {
                "key": "price",
                "label": "最终价格表",
                "type": "Excel",
                "folder": PACKAGING_PRICING_DIR,
                "accept": ".xlsx,.xlsm,.xls,.csv",
                "files": _folder_file_infos(project_dir, [PACKAGING_PRICING_DIR], suffixes={".xlsx", ".xlsm", ".xls", ".csv"}),
            },
            {
                "key": "detail",
                "label": "1688详情页",
                "type": "HTML",
                "folder": PRODUCT_DETAIL_DIR,
                "accept": ".html,.htm,.txt",
                "files": _folder_file_infos(project_dir, [PRODUCT_DETAIL_DIR, PURCHASE_DIR], suffixes={".html", ".htm", ".txt"}),
            },
            {
                "key": "competitor",
                "label": "竞品详情页",
                "type": "多个 HTML",
                "folder": COMPETITOR_DIR,
                "accept": ".html,.htm,.txt",
                "files": _folder_file_infos(project_dir, [COMPETITOR_DIR, LEGACY_COMPETITOR_DIR], suffixes={".html", ".htm", ".txt"}),
            },
        ],
        "basic": [
            {"label": "路线", "field": "route", "value": first.get("route") or "Haul Generic", "source": "系统推断", "scope": "all"},
            {"label": "Brand", "field": "brand", "value": first.get("brand") or "Generic", "source": "路线", "scope": "all"},
            {"label": "Manufacturer", "field": "manufacturer", "value": first.get("manufacturer") or "Generic", "source": "路线", "scope": "all"},
            {"label": "品类", "field": "category", "value": first.get("category") or first.get("product_type") or "", "source": "系统推断", "scope": "all"},
            {"label": "Product Type", "field": "product_type", "value": first.get("product_type") or "", "source": "模板", "scope": "all"},
            {"label": "材质", "field": "material", "value": first.get("material") or "", "source": "1688", "scope": "all"},
            {"label": "包装内容", "field": "accessories", "value": first.get("accessories") or first.get("set_count") or "", "source": "价格表", "scope": "all"},
        ],
        "specs": [
            {"label": "单品尺寸 length", "field": "package_length_in", "value": first.get("package_length_in") or "", "source": "价格表", "scope": "all"},
            {"label": "单品尺寸 width", "field": "package_width_in", "value": first.get("package_width_in") or "", "source": "价格表", "scope": "all"},
            {"label": "单品尺寸 height", "field": "package_height_in", "value": first.get("package_height_in") or "", "source": "价格表", "scope": "all"},
            {"label": "重量 lb", "field": "package_weight_lb", "value": first.get("package_weight_lb") or "", "source": "价格表", "scope": "all"},
            {"label": "产地", "field": "country_of_origin", "value": first.get("country_of_origin") or "China", "source": "默认", "scope": "all"},
            {"label": "危险品", "field": "dangerous_goods", "value": first.get("dangerous_goods") or "No", "source": "人工确认", "scope": "all"},
        ],
        "title": first.get("title") or "",
        "selling": _selling_text(first),
        "keywords": _keywords_for_row(first),
        "issues": issues,
    }


def _template_module_payload(project_dir, product_name, rows, source_template, latest_template, latest_report, check_findings, next_version, item):
    first = rows[0] if rows else _blank_intake_row(project_dir, product_name)
    latest_name = latest_template["name"] if latest_template else "尚未生成"
    error_count = len(check_findings)
    missing = _missing_required_fields(rows)
    can_upload = bool(latest_template)
    source_name = source_template["name"] if source_template else "等待上传 Amazon 模板"
    source_path = source_template["folder_relative"] if source_template else f"{_relative(project_dir)}/{TEMPLATE_SOURCE_DIR}"
    if latest_template and error_count == 0:
        status = "自检通过"
    elif latest_template:
        status = f"有 {error_count} 项待修正"
    else:
        status = "等待模板填表"

    checks = []
    if check_findings:
        checks.extend([
            ["error", f"{item.get('field') or '字段'}", item.get("message") or "模板自检发现问题"]
            for item in check_findings[:6]
        ])
    elif latest_template:
        checks.append(["ok", "模板自检通过", "未发现当前自检规则覆盖的问题。"])
    else:
        checks.append(["warn", "待填表", f"上传 Amazon 模板后生成 {safe_name(product_name)}_{shared_version_label(next_version)}.xlsx。"])
    if missing:
        checks.append(["warn", "产品资料仍需确认", "缺少：" + "、".join(missing[:6])])

    generated = []
    if latest_template:
        generated.append({"label": "填好的表格", "file": latest_template})
    if latest_report:
        generated.append({"label": "自检报告", "file": latest_report})

    return {
        "ext": (source_template["name"].rsplit(".", 1)[-1].upper() if source_template and "." in source_template["name"] else "XLSX"),
        "source": source_name,
        "path": source_path,
        "sourceFile": source_template,
        "latestFile": latest_template,
        "reportFile": latest_report,
        "productType": first.get("product_type") or "-",
        "sheet": "Template",
        "skuCount": f"{len(rows)} 个" if rows else "0 个",
        "output": latest_name,
        "status": status,
        "metrics": [
            ["写入字段", str(item.get("written_field_count") or "-")],
            ["生成 SKU", str(len(rows) if rows else item.get("sku_count") or 0)],
            ["必须修正", str(error_count)],
            ["建议确认", str(len(missing))],
        ],
        "checks": checks,
        "mappings": _mapping_preview(first),
        "generated": generated,
        "next": _template_next_text(latest_template, error_count, next_version, product_name),
        "fillAction": f"生成 {safe_name(product_name)}_{shared_version_label(next_version)}",
        "canUpload": can_upload,
        "canFill": bool(rows and source_template),
    }


def _feedback_module_payload(project_dir, product_name, status, latest_template, latest_report, version_files, failure_reports, next_version, item):
    uploaded = status == "uploaded_success"
    pending_confirmation = status == "upload_failed_pending_confirmation"
    current_version = _version_label(latest_template["name"] if latest_template else "")
    if uploaded:
        version_status = "上传成功"
    elif pending_confirmation:
        version_status = f"{current_version or '当前版本'} 上传失败"
    elif latest_template:
        version_status = f"{current_version or '当前版本'} 待上传"
    else:
        version_status = "未开始"

    version_infos = [_file_info(project_dir, path) for path in version_files]
    version_infos = [item for item in version_infos if item]
    timeline = _version_timeline(version_infos, uploaded, pending_confirmation)
    pending_fix_report = _file_info(project_dir, _project_file_candidate(project_dir, item.get("latest_fix_report", ""))) if item.get("latest_fix_report") else None
    latest_failure_report = _file_info(project_dir, _project_file_candidate(project_dir, item.get("latest_failure_report", ""))) if item.get("latest_failure_report") else None
    report_name = (latest_failure_report or (failure_reports[0] if failure_reports else None) or {}).get("name", "无失败报告")
    latest_template_name = latest_template["name"] if latest_template else f"尚未生成 {safe_name(product_name)}_V"
    fix_rows = _learning_rows_for_product(product_name)
    success_learning = _success_learning_rows(item)
    planned_version = item.get("pending_next_version") or next_version
    can_upload_feedback = bool(latest_template and not uploaded and not pending_confirmation)

    return {
        "versionStatus": version_status,
        "currentFile": latest_template_name,
        "currentFileInfo": latest_template,
        "reportFile": pending_fix_report or latest_report,
        "fixReportFile": pending_fix_report,
        "note": _feedback_note(uploaded, latest_template, current_version, pending_confirmation, planned_version),
        "meta": [
            ["产品名", product_name],
            ["当前版本", current_version or "未生成"],
            ["下一版", "成功" if uploaded else shared_version_label(planned_version)],
        ],
        "report": report_name,
        "reportFiles": failure_reports,
        "plannedVersion": "成功" if uploaded else shared_version_label(planned_version),
        "reportNote": "看下方解析，确认后生成下一版。" if pending_confirmation else ("上传失败时选择 Amazon Processing Summary。" if latest_template else "先生成 V 模板。"),
        "analysis": _feedback_analysis_cards(project_dir, latest_failure_report, pending_fix_report, planned_version),
        "analysisText": _feedback_analysis_text(project_dir, latest_failure_report, pending_fix_report, planned_version),
        "failureSummary": [["失败报告", str(len(failure_reports))], ["学习记录", str(len(fix_rows))]],
        "generateAction": "已上传成功" if uploaded else ("确认并生成下一版" if pending_confirmation else ("上传失败报告" if latest_template else "等待 V 生成")),
        "fixStatus": "学习完成" if uploaded else ("待确认" if pending_confirmation else ("待上传" if latest_template else "未开始")),
        "fixMeta": [
            ["来源模板", latest_template_name],
            ["失败报告", report_name],
            ["新模板", "成功结束" if uploaded else (f"确认后生成 {safe_name(product_name)}_{shared_version_label(planned_version)}.xlsx" if pending_confirmation else f"{safe_name(product_name)}_{shared_version_label(planned_version)}.xlsx")],
        ],
        "timeline": timeline,
        "fixes": fix_rows or [["无", "尚未上传失败报告", "无", "无", "等待反馈"]],
        "successLearning": success_learning,
        "learning": _learning_cards(product_name, uploaded, fix_rows),
        "versionFiles": version_infos,
        "next": _feedback_next_text(uploaded, latest_template, current_version, pending_confirmation, planned_version),
        "download": "下载当前模板" if latest_template else "等待 V 生成",
        "success": "标记当前版本上传成功并学习" if latest_template else "等待上传反馈",
        "fail": "确认修正并生成下一版" if pending_confirmation else ("上传失败报告" if latest_template else "等待失败报告"),
        "canUploadFeedback": can_upload_feedback,
        "canConfirmFix": bool(latest_template and pending_confirmation and not uploaded),
        "canMarkSuccess": bool(latest_template and not uploaded and not pending_confirmation),
    }


def _form_files(form):
    if "files" not in form:
        return []
    files_field = form["files"]
    return files_field if isinstance(files_field, list) else [files_field]


def _success_learning_rows(item):
    rows = []
    version = item.get("success_version")
    template = item.get("success_template") or item.get("latest_template") or ""
    if item.get("status") == "uploaded_success":
        title = f"v{version} 上传成功" if version else "上传成功"
        body = item.get("notes") or "当前模板已确认上传成功。"
        rows.append(["success", title, body])
        if template:
            rows.append(["", "成功模板", template])
    learned_rules = item.get("learned_rules") if isinstance(item.get("learned_rules"), list) else []
    for rule in learned_rules[:6]:
        rows.append(["", "已学习规则", str(rule)])
    return rows


def _upload_files(project_dir, folder, file_items):
    project_path = _project_path({"project_dir": project_dir})
    if project_path == PROJECTS_DIR.resolve():
        raise ValueError("请选择具体项目。")
    if folder not in PROJECT_FOLDERS:
        raise ValueError("资料夹不正确。")
    real_files = [item for item in file_items if getattr(item, "filename", "") and getattr(item, "file", None)]
    if not real_files:
        raise ValueError("请选择文件。")

    target_dir = project_path / folder
    target_dir.mkdir(parents=True, exist_ok=True)
    saved = []
    for item in real_files:
        filename = _safe_upload_filename(item.filename)
        output_path = _unique_file_path(target_dir, filename)
        item.file.seek(0)
        with output_path.open("wb") as output:
            shutil.copyfileobj(item.file, output)
        saved.append({
            "name": output_path.name,
            "path": str(output_path),
            "relative_path": _relative(output_path),
        })

    if folder in {TEMPLATE_SOURCE_DIR, LEGACY_TEMPLATE_SOURCE_DIR}:
        template_files = [Path(file["path"]) for file in saved if Path(file["path"]).suffix.lower() in {".xlsx", ".xlsm"}]
        if template_files:
            save_project_status(project_path, {
                "source_template": _relative_to_project(project_path, template_files[-1]),
                "blocked_reason": None,
            })

    return {
        "ok": True,
        "message": f"已放入 {len(saved)} 个文件",
        "folder": folder,
        "files": saved,
    }


def _upload_feedback(project_dir, file_items, note=""):
    project_path = _project_path({"project_dir": project_dir})
    status = load_project_status(project_path)
    if status.get("status") == "uploaded_success":
        raise ValueError("该项目已标记上传成功，不能再上传失败报告。")
    latest_template = _latest_filled_template(project_path, status)
    if not latest_template:
        raise ValueError("请先生成 V 模板，再记录上传失败。")
    if status.get("status") == "upload_failed_pending_confirmation":
        raise ValueError("已有一份失败修正计划待确认，请先确认生成下一版。")
    real_files = [item for item in file_items if getattr(item, "filename", "") and getattr(item, "file", None)]
    if not real_files:
        raise ValueError("请选择失败分析报告。")

    target_dir = project_path / FEEDBACK_REPORT_DIR
    target_dir.mkdir(parents=True, exist_ok=True)
    saved_paths = []
    saved = []
    for item in real_files:
        filename = _safe_upload_filename(item.filename)
        output_path = _unique_file_path(target_dir, filename)
        item.file.seek(0)
        with output_path.open("wb") as output:
            shutil.copyfileobj(item.file, output)
        saved_paths.append(output_path)
        saved.append({
            "name": output_path.name,
            "path": str(output_path),
            "relative_path": _relative(output_path),
        })

    product_name = infer_product_name(project_path, "")
    added, learnings_path, report_path = learn_reports(saved_paths, product_name=product_name, note=note or "工作台上传失败报告")
    next_version = _next_template_version(project_path)
    fix_report = _write_feedback_fix_report(project_path, saved_paths, added, next_version=next_version)
    save_project_status(project_path, {
        "status": "upload_failed_pending_confirmation",
        "failed_template": _relative_to_project(project_path, latest_template),
        "latest_failure_report": _relative_to_project(project_path, saved_paths[-1]),
        "latest_fix_report": _relative_to_project(project_path, fix_report),
        "pending_next_version": next_version,
        "blocked_reason": "上传失败报告已解析，等待人工确认修正方案。",
    })
    return {
        "ok": True,
        "message": f"已解析失败报告并生成修正计划，识别 {len(added)} 条失败记录，请确认后生成下一版",
        "files": saved,
        "learned_count": len(added),
        "learnings_path": str(learnings_path),
        "learning_report_path": str(report_path),
        "fix_report": _file_info(project_path, fix_report),
        "next_version": next_version,
    }


def _confirm_feedback_fix(payload):
    project_dir = _project_path(payload)
    status = load_project_status(project_dir)
    if status.get("status") == "uploaded_success":
        raise ValueError("该项目已标记上传成功。")
    if status.get("status") != "upload_failed_pending_confirmation":
        raise ValueError("没有待确认的失败修正计划，请先上传失败报告。")

    version = int(status.get("pending_next_version") or _next_template_version(project_dir))
    fill_result = _fill_template_version({
        "project_dir": str(project_dir),
        "version": version,
        "reason": "confirmed_failure_fix",
    })
    fix_report = _project_file_candidate(project_dir, status.get("latest_fix_report", ""))
    if fix_report and fix_report.exists():
        _append_feedback_generation_result(fix_report, fill_result)
    save_project_status(project_dir, {
        "pending_next_version": None,
        "blocked_reason": None,
        "failed_template": None,
    })
    return {
        "ok": True,
        "message": f"已按修正计划生成 {shared_version_label(version)}，请人工上传 Amazon 验证",
        "fill": fill_result,
        "fix_report": _file_info(project_dir, fix_report) if fix_report else None,
    }


def _safe_upload_filename(filename):
    name = Path(str(filename).replace("\\", "/")).name
    cleaned = safe_name(name)
    if cleaned in {".", ".."}:
        cleaned = "uploaded_file"
    return cleaned


def _unique_file_path(folder, filename):
    path = folder / filename
    if not path.exists():
        return path
    original = Path(filename)
    stem = original.stem or "uploaded_file"
    suffix = original.suffix
    index = 2
    while True:
        candidate = folder / f"{stem}_{index}{suffix}"
        if not candidate.exists():
            return candidate
        index += 1

def _rules_payload():
    if not RULES_JSON.exists():
        return {
            "exists": False,
            "template_count": 0,
            "product_type_count": 0,
            "product_types": [],
        }
    rules = json.loads(RULES_JSON.read_text(encoding="utf-8"))
    product_types = []
    for product_type, data in sorted(rules.get("product_types", {}).items()):
        product_types.append({
            "product_type": product_type,
            "template_count": data.get("template_count", 0),
            "sku_count": data.get("sku_count", 0),
            "fixed_default_count": len(data.get("fixed_default_fields", [])),
            "unmapped_count": len(data.get("often_filled_unmapped_fields", [])),
        })
    return {
        "exists": True,
        "template_count": rules.get("template_count", 0),
        "product_type_count": rules.get("product_type_count", 0),
        "report_path": str(RULES_REPORT),
        "json_path": str(RULES_JSON),
        "product_types": product_types,
    }


def _report_payload():
    if not RULES_REPORT.exists():
        return {"exists": False, "preview": ""}
    lines = RULES_REPORT.read_text(encoding="utf-8").splitlines()
    return {
        "exists": True,
        "path": str(RULES_REPORT),
        "preview": "\n".join(lines[:80]),
    }


def _create_project(payload):
    name = str(payload.get("name", "")).strip()
    if not name:
        raise ValueError("请填写产品名。")
    project_dir, intake_path = create_project(name)
    save_project_status(project_dir, {
        "product_name": name,
        "status": "not_started",
        "latest_draft": _relative_to_project(project_dir, intake_path),
    })
    return {
        "ok": True,
        "message": "项目已创建",
        "project_dir": str(project_dir),
        "intake_path": str(intake_path),
    }


def _rename_project(payload):
    project_dir = _project_path(payload)
    name = str(payload.get("name", "")).strip()
    if not name:
        raise ValueError("请填写新的项目名。")
    if any(char in name for char in "/\\:"):
        raise ValueError("项目名不能包含 /、\\ 或 :。")

    status = load_project_status(project_dir)
    rows, draft_path, _rows_error = _latest_intake_rows(project_dir, status)
    if rows and draft_path and draft_path.suffix.lower() == ".xlsx":
        for row in rows:
            row["product_name"] = name
        write_intake_workbook(draft_path, rows)

    save_project_status(project_dir, {
        "product_name": name,
        "renamed_at": datetime.now().isoformat(timespec="seconds"),
    })
    return {
        "ok": True,
        "message": "项目名已更新",
        "product_name": name,
    }


def _analyze_project(payload):
    project_dir = _project_path(payload)
    route_mode = str(payload.get("route_mode") or "").strip()
    draft_path, report_path = analyze_project(project_dir, route_mode=route_mode)
    rows = _read_rows_clean(draft_path)
    product_name = rows[0].get("product_name") if rows else infer_product_name(project_dir, "")
    save_project_status(project_dir, {
        "product_name": product_name,
        "latest_draft": _relative_to_project(project_dir, draft_path),
        "latest_analysis_report": str(report_path),
        "selected_route_mode": route_mode,
        "sku_count": len(rows),
        "status": "not_started",
        "blocked_reason": None,
    })
    return {
        "ok": True,
        "message": "产品资料已提炼，等待人工确认",
        "draft_file": _file_info(project_dir, draft_path),
        "report_file": _file_info(project_dir, report_path),
        "rows": _json_rows(rows),
    }


def _save_intake(payload):
    project_dir = _project_path(payload)
    rows = payload.get("rows") or []
    product_name = str(payload.get("product_name") or "").strip() or infer_product_name(project_dir, "")
    normalized_rows = _normalize_intake_rows(rows, project_dir, product_name)
    if not normalized_rows:
        raise ValueError("没有可保存的产品资料行。")
    blocking_copy_issues = [issue for issue in _copy_quality_issues(normalized_rows) if issue and issue[0] == "error"]
    if blocking_copy_issues:
        first_issue = blocking_copy_issues[0]
        raise ValueError(f"{first_issue[1]}：{first_issue[2]} 宁可留空，也不要保存兜底或串项文案。")
    source_template = _find_source_template(project_dir)
    if source_template:
        _apply_template_product_type(normalized_rows, source_template)
    product_name = normalized_rows[0].get("product_name") or product_name
    output_path = project_dir / PRODUCT_DETAIL_DIR / f"{safe_name(product_name)}_产品资料.xlsx"
    write_intake_workbook(output_path, normalized_rows)
    save_project_status(project_dir, {
        "product_name": product_name,
        "latest_draft": _relative_to_project(project_dir, output_path),
        "sku_count": len(normalized_rows),
        "confirmed_at": datetime.now().isoformat(timespec="seconds"),
        "status": "not_started",
        "blocked_reason": None,
    })
    return {
        "ok": True,
        "message": "产品资料已保存并确认",
        "draft_file": _file_info(project_dir, output_path),
        "rows": _json_rows(normalized_rows),
    }


def _fill_template_version(payload):
    project_dir = _project_path(payload)
    status = load_project_status(project_dir)
    if status.get("status") == "uploaded_success" and not payload.get("force"):
        raise ValueError("该项目已标记上传成功。如需重新生成，请先明确 force。")

    product_name = infer_product_name(project_dir, str(payload.get("product_name", "")).strip())
    draft_path = _latest_draft_for_fill(project_dir, status)
    template_path = _source_template_for_fill(project_dir)
    rows = _read_rows_clean(draft_path)
    if not rows:
        raise ValueError("产品资料为空，请先完成资料提炼和人工确认。")
    if _apply_template_product_type(rows, template_path):
        write_intake_workbook(draft_path, rows)
    intake_errors = [item for item in validate_intake(draft_path) if item.get("severity") == "error"]
    if intake_errors:
        first_error = intake_errors[0]
        raise ValueError(f"产品资料自检未通过：{first_error.get('field')} - {first_error.get('message')} {first_error.get('fix')}")
    product_name = rows[0].get("product_name") or product_name
    version = int(payload.get("version") or _next_template_version(project_dir))
    output_path = _versioned_template_path(project_dir, product_name, version)

    result = auto_fill_project(
        project_dir,
        draft_path=draft_path,
        template_path=template_path,
        output_path=output_path,
        force=bool(payload.get("force")),
    )
    if result.get("blocked"):
        raise ValueError(result.get("reason") or "自动填表暂停。")

    filled_file = _file_info(project_dir, result.get("filled_path"))
    report_file = _file_info(project_dir, result.get("report_path"))
    return {
        "ok": True,
        "message": _auto_fill_message(result),
        "version": version,
        "status": result.get("status"),
        "paths": {key: str(value) for key, value in result.items() if key.endswith("_path") or key.endswith("_dir")},
        "filled_file": filled_file,
        "report_file": report_file,
        "sku_count": result.get("sku_count"),
        "written_field_count": result.get("written_field_count"),
        "error_count": result.get("error_count"),
    }


def _auto_fill(payload):
    project_dir = _project_path(payload)
    result = auto_fill_project(project_dir, force=bool(payload.get("force")))
    cleaned = {key: str(value) for key, value in result.items() if key.endswith("_path") or key.endswith("_dir")}
    filled_file = _file_info(project_dir, result.get("filled_path"))
    report_file = _file_info(project_dir, result.get("report_path"))
    return {
        "ok": True,
        "message": _auto_fill_message(result),
        "status": result.get("status"),
        "skipped": result.get("skipped", False),
        "blocked": result.get("blocked", False),
        "paths": cleaned,
        "filled_file": filled_file,
        "report_file": report_file,
        "sku_count": result.get("sku_count"),
        "error_count": result.get("error_count"),
    }


def _mark_uploaded(payload):
    project_dir = _project_path(payload)
    template_value = str(payload.get("template_path", "")).strip()
    template_path = _local_project_file(project_dir, template_value) if template_value else _latest_filled_template(project_dir)
    if template_path is None:
        raise ValueError("没有找到可标记的上传模板。")
    product_name = infer_product_name(project_dir, "")
    sku_count = infer_sku_count(project_dir)
    status_path = mark_uploaded_success(
        project_dir,
        product_name=product_name,
        latest_template=template_path,
        sku_count=sku_count,
        notes="工作台标记上传成功",
    )
    copied_template = _copy_success_template(project_dir, template_path)
    learn_warning = ""
    try:
        rules, json_path, report_path = learn_success_templates()
    except Exception as exc:
        rules, json_path, report_path = {}, "", ""
        learn_warning = f"成功状态已记录，但学习规则刷新失败：{exc}"
    return {
        "ok": True,
        "message": learn_warning or "已标记上传成功，并更新成功模板学习库",
        "status_path": str(status_path),
        "success_template": str(copied_template) if copied_template else "",
        "template_count": rules.get("template_count", 0) if isinstance(rules, dict) else 0,
        "rules_json": str(json_path) if json_path else "",
        "rules_report": str(report_path) if report_path else "",
    }


def _delete_project(payload):
    project_dir = _project_path(payload)
    deleted_path = delete_project(project_dir)
    return {
        "ok": True,
        "message": f"项目已删除：{deleted_path.name}",
        "project_dir": str(deleted_path),
    }


def _reveal_file(payload):
    path = _local_file_path(payload.get("path", ""))
    if not path.exists():
        raise ValueError("文件不存在。")
    if path.is_dir():
        subprocess.run(["open", str(path)], check=False)
    else:
        subprocess.run(["open", "-R", str(path)], check=False)
    return {
        "ok": True,
        "message": "已在 Finder 中定位",
        "path": str(path),
    }


def _learn_success():
    rules, json_path, report_path = learn_success_templates()
    return {
        "ok": True,
        "message": "成功规则已更新",
        "template_count": rules["template_count"],
        "product_type_count": rules["product_type_count"],
        "json_path": str(json_path),
        "report_path": str(report_path),
    }


def _latest_intake_rows(project_dir, item=None):
    item = item or {}
    candidates = []
    latest_value = item.get("latest_draft") or ""
    if latest_value:
        candidates.append(_project_file_candidate(project_dir, latest_value))
    for folder in DRAFT_DIRS:
        draft_dir = Path(project_dir) / folder
        if not draft_dir.exists():
            continue
        candidates.extend([
            path for path in draft_dir.glob("*.xlsx")
            if not path.name.startswith("~$")
            and ("产品资料" in path.name or "自动提炼草稿" in path.name)
        ])
    candidates = [path for path in candidates if path and path.exists()]
    candidates = sorted(set(candidates), key=lambda path: path.stat().st_mtime, reverse=True)
    last_error = ""
    for path in candidates:
        try:
            rows = _read_rows_clean(path)
        except Exception as exc:
            last_error = str(exc)
            continue
        if rows:
            return rows, path, ""
    return [], candidates[0] if candidates else None, last_error


def _latest_filled_template(project_dir, item=None):
    item = item or {}
    latest_value = item.get("latest_template") or item.get("verification_template") or ""
    if latest_value:
        candidate = _project_file_candidate(project_dir, latest_value)
        if candidate and candidate.exists() and candidate.suffix.lower() in {".xlsx", ".xlsm"}:
            return candidate
    version_files = _template_version_files(project_dir)
    if version_files:
        return version_files[-1]
    return None


def _latest_check_report(project_dir, item, latest_template):
    latest_value = item.get("latest_check_report") or ""
    if latest_value:
        candidate = _project_file_candidate(project_dir, latest_value)
        if candidate and candidate.exists():
            return candidate
    if latest_template:
        output_report = OUTPUTS_DIR / f"{Path(latest_template).stem}_模板自检报告.md"
        if output_report.exists():
            return output_report
    candidates = sorted(OUTPUTS_DIR.glob("*_模板自检报告.md"), key=lambda path: path.stat().st_mtime, reverse=True)
    template_stem = Path(latest_template).stem if latest_template else ""
    for path in candidates:
        if template_stem and template_stem in path.stem:
            return path
    return None


def _find_source_template(project_dir):
    project_dir = Path(project_dir)
    status = load_project_status(project_dir)
    for key in ("source_template", "verification_source_template"):
        value = status.get(key)
        if value:
            candidate = _project_file_candidate(project_dir, value)
            if candidate and candidate.exists():
                return candidate
    for folder in [TEMPLATE_SOURCE_DIR, LEGACY_TEMPLATE_SOURCE_DIR]:
        template_dir = project_dir / folder
        if not template_dir.exists():
            continue
        candidates = [
            path for path in template_dir.iterdir()
            if path.is_file()
            and not path.name.startswith("~$")
            and path.suffix.lower() in {".xlsx", ".xlsm"}
        ]
        if candidates:
            return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]
    return None


def _source_template_for_fill(project_dir):
    source_template = _find_source_template(project_dir)
    if source_template:
        return source_template
    return find_template(project_dir)


def _apply_template_product_type(rows, template_path):
    if all(str(row.get("product_type") or "").strip() for row in rows):
        return False
    product_type = extract_template_product_type(template_path)
    if not product_type:
        raise ValueError("产品资料缺少 Product Type，且当前 Amazon 模板没有可读取的 Product Type。请先确认模板是否正确。")
    changed = False
    for row in rows:
        if not str(row.get("product_type") or "").strip():
            row["product_type"] = product_type
            changed = True
    return changed


def _latest_draft_for_fill(project_dir, status):
    latest = status.get("latest_draft") or status.get("verification_draft") or ""
    if latest:
        candidate = _project_file_candidate(project_dir, latest)
        if candidate and candidate.exists():
            return candidate
    rows, path, _error = _latest_intake_rows(project_dir, status)
    if rows and path:
        return path
    raise FileNotFoundError("没有找到已确认的产品资料，请先提炼并确认。")


def _read_rows_clean(path):
    rows = read_intake_rows(path)
    cleaned = []
    for row in rows:
        normalized = {header: _cell_text(row.get(header)) for header in INTAKE_HEADERS}
        if normalized.get("sku", "").upper().startswith("DEMO"):
            continue
        if normalized.get("project_name") == "示例项目":
            continue
        if normalized.get("product_name") == "示例收纳袋":
            continue
        if not any(normalized.get(header) for header in INTAKE_HEADERS):
            continue
        cleaned.append(normalized)
    return cleaned


def _normalize_intake_rows(rows, project_dir, product_name):
    normalized_rows = []
    for index, raw_row in enumerate(rows, 1):
        row = {header: _cell_text(raw_row.get(header)) for header in INTAKE_HEADERS}
        if not any(row.values()):
            continue
        row["project_name"] = row.get("project_name") or Path(project_dir).name
        row["product_name"] = row.get("product_name") or product_name
        row["route"] = row.get("route") or "Haul Generic Variation"
        row["brand"] = row.get("brand") or "Generic"
        row["manufacturer"] = row.get("manufacturer") or row.get("brand") or "Generic"
        row["sku"] = row.get("sku") or _default_sku(row.get("product_name") or product_name, index)
        row["country_of_origin"] = row.get("country_of_origin") or "China"
        row["batteries_required"] = row.get("batteries_required") or "No"
        row["dangerous_goods"] = row.get("dangerous_goods") or "No"
        normalized_rows.append(row)
    if not normalized_rows and product_name:
        normalized_rows.append(_blank_intake_row(project_dir, product_name))
    return prepare_variation_rows(normalized_rows, project_dir)


def _json_rows(rows):
    return [
        {header: _cell_text(row.get(header)) for header in INTAKE_HEADERS}
        for row in rows
    ]


def _blank_intake_row(project_dir, product_name):
    return {
        "project_name": Path(project_dir).name,
        "product_name": product_name,
        "route": "Haul Generic Variation",
        "brand": "Generic",
        "manufacturer": "Generic",
        "parent_sku": "",
        "parentage_level": "",
        "variation_theme": "",
        "country_of_origin": "China",
        "batteries_required": "No",
        "dangerous_goods": "No",
    }


def _cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _default_sku(product_name, index):
    words = re.findall(r"[A-Za-z0-9]+", str(product_name).upper())
    if words:
        base = "-".join(words[:4])[:28]
    else:
        base = safe_name(str(product_name)).upper()[:28] or "PRODUCT"
    return f"{base}-{index:03d}"


def _folder_file_infos(project_dir, folders, suffixes=None):
    infos = []
    seen = set()
    for folder in folders:
        folder_path = Path(project_dir) / folder
        if not folder_path.exists():
            continue
        for path in sorted(folder_path.iterdir(), key=lambda item: item.stat().st_mtime if item.exists() else 0, reverse=True):
            if not path.is_file() or path.name.startswith("~$"):
                continue
            if suffixes and path.suffix.lower() not in suffixes:
                continue
            if path.resolve() in seen:
                continue
            info = _file_info(project_dir, path)
            if info:
                infos.append(info)
                seen.add(path.resolve())
    return infos


def _template_version_files(project_dir):
    version_dir = Path(project_dir) / LEGACY_FILLED_TEMPLATE_DIR
    if not version_dir.exists():
        return []
    candidates = [
        path for path in version_dir.iterdir()
        if path.is_file()
        and not path.name.startswith("~$")
        and path.suffix.lower() in {".xlsx", ".xlsm"}
        and "产品资料" not in path.name
        and "自动提炼草稿" not in path.name
    ]
    return sorted(candidates, key=lambda path: (_version_number(path.name), path.stat().st_mtime))


def _version_number(value):
    return shared_version_number(value)


def _version_label(value):
    number = _version_number(value)
    return shared_version_label(number) if number else ""


def _next_template_version(project_dir):
    return shared_next_template_version(project_dir)


def _versioned_template_path(project_dir, product_name, version):
    return shared_versioned_template_path(project_dir, product_name, version)


def _intake_issues(rows, rows_error):
    issues = []
    if rows_error:
        issues.append(["error", "产品资料读取失败", rows_error])
    if not rows:
        issues.append(["warn", "等待资料提炼", "上传最终价格表、1688详情页和竞品详情页后点击提炼。"])
        return issues
    missing = _missing_required_fields(rows)
    if missing:
        issues.append(["warn", "核心字段待确认", "缺少：" + "、".join(missing[:8])])
    if any(
        not row.get("list_price")
        for row in rows
        if str(row.get("parentage_level") or "").strip().lower() != "parent"
    ):
        issues.append(["warn", "售价需要确认", "List Price / Haul Price 为空会在模板自检中被拦截。"])
    cjk_fields = sorted({
        field
        for row in rows
        for field in COPY_FIELDS
        if CJK_RE.search(str(row.get(field) or ""))
    })
    if cjk_fields:
        issues.append(["error", "文案含中文", "请把 " + "、".join(cjk_fields) + " 改成自然的跨境英语表达。"])
    quality_issues = _copy_quality_issues(rows)
    issues.extend(quality_issues)
    if not missing:
        issues.append(["ok", "资料结构完整", "核心字段已具备，可以人工检查后确认。"])
    return issues


BAD_COPY_PATTERNS = [
    ("generic product", "标题或描述还是通用兜底商品名。"),
    ("designed for everyday use", "五点还是通用兜底场景。"),
    ("home, travel, office", "描述还是通用兜底场景。"),
    ("includes the product shown in the listing photos", "五点没有提炼真实包装内容。"),
    ("wide range of everyday needs", "五点过泛，没有商品特征。"),
]


def _copy_quality_issues(rows):
    issues = []
    for row in rows:
        copy = " ".join(str(row.get(field) or "") for field in COPY_FIELDS)
        lower_copy = copy.lower()
        for phrase, reason in BAD_COPY_PATTERNS:
            if phrase in lower_copy:
                issues.append(["error", "文案质量不合格", reason + " 请重新提炼或手动改写标题、五点和描述。"])
                break
        title = str(row.get("title") or "")
        if "videos" in title.lower():
            issues.append(["error", "标题混入网页噪音", "标题里出现 VIDEOS，通常是从网页脚本或图片字段误抓。"])

        product_context = " ".join(str(row.get(field) or "").lower() for field in [
            "product_name", "title", "item_type_keyword", "accessories", "notes"
        ])
        material = str(row.get("material") or "").lower()
        category = str(row.get("category") or "").lower()
        item_type = str(row.get("item_type_keyword") or "").lower()
        if any(keyword in product_context for keyword in ["cat", "猫", "q-tip", "catnip"]) and (
            "cookbook" in item_type or category == "home" or material == "metal"
        ):
            issues.append(["error", "商品理解疑似跑偏", "猫玩具不应识别成 home / cookbook-stands / Metal，请按宠物玩具方向重写。"])
    return issues


def _missing_required_fields(rows):
    if not rows:
        return list(REQUIRED_CORE_FIELDS)
    missing = []
    for field in REQUIRED_CORE_FIELDS:
        if any(
            not row.get(field)
            for row in rows
            if not (
                str(row.get("parentage_level") or "").strip().lower() == "parent"
                and field in PARENT_OPTIONAL_CORE_FIELDS
            )
        ):
            missing.append(field)
    return missing


def _selling_text(row):
    bullets = [row.get(f"bullet_{index}") for index in range(1, 6)]
    text = "\n".join(_cell_text(value) for value in bullets if _cell_text(value))
    description = _cell_text(row.get("description"))
    return f"{text}\n\n{description}".strip() if description else text


def _keywords_for_row(row):
    candidates = [
        row.get("item_type_keyword"),
        row.get("product_type"),
        row.get("category"),
        row.get("material"),
    ]
    title_words = re.findall(r"[A-Za-z][A-Za-z0-9-]{2,}", row.get("title") or "")
    candidates.extend(title_words[:6])
    keywords = []
    for value in candidates:
        value = _cell_text(value)
        if value and value not in keywords:
            keywords.append(value)
    return keywords[:8]


def _mapping_preview(row):
    rows = [
        ("产品名", row.get("title") or row.get("product_name"), "item_name", "已写入" if row.get("title") else "待写入"),
        ("SKU", row.get("sku"), "contribution_sku / model_number", "已写入" if row.get("sku") else "待写入"),
        ("Product Type", row.get("product_type"), "product_type", "已写入" if row.get("product_type") else "需确认"),
        ("颜色", row.get("color"), "color", "已写入" if row.get("color") else "待写入"),
        ("材质", row.get("material"), "material", "已写入" if row.get("material") else "待写入"),
        ("售价", row.get("list_price") or row.get("haul_price"), "list_price / BZR price", "已写入" if row.get("list_price") or row.get("haul_price") else "需确认"),
        ("尺寸重量", _dimension_text(row), "package dimensions / weight", "已写入" if row.get("package_weight_lb") else "需确认"),
    ]
    return [[_cell_text(value) for value in item] for item in rows]


def _dimension_text(row):
    dims = [row.get("package_length_in"), row.get("package_width_in"), row.get("package_height_in")]
    dim_text = " x ".join(_cell_text(value) or "-" for value in dims)
    weight = _cell_text(row.get("package_weight_lb")) or "-"
    return f"{dim_text} in / {weight} lb"


def _template_next_text(latest_template, error_count, next_version, product_name):
    if not latest_template:
        return f"生成 {safe_name(product_name)}_{shared_version_label(next_version)}.xlsx 后，下载并人工上传 Amazon。"
    if error_count:
        return "先查看自检报告并修正；必要时重新生成下一版。"
    return "人工上传当前模板。成功则标记上传成功，失败则上传失败报告生成下一版。"


def _feedback_note(uploaded, latest_template, current_version, pending_confirmation=False, planned_version=None):
    if uploaded:
        return "当前版本已确认上传成功，成功模板已进入学习流程。"
    if pending_confirmation:
        return f"{current_version or '当前版本'} 上传失败，已生成失败原因分析和修正计划；人工确认后生成 {shared_version_label(planned_version)}。"
    if latest_template:
        return f"{current_version or '当前版本'} 已生成，等待人工上传 Amazon 后记录成功或失败。"
    return "等待 Amazon 模板填表后生成 V。"


def _feedback_next_text(uploaded, latest_template, current_version, pending_confirmation=False, planned_version=None):
    if uploaded:
        return "已上传成功，可用于后续同类模板学习。"
    if pending_confirmation:
        return f"先查看失败原因分析和准备修改的字段；确认无误后生成 {shared_version_label(planned_version)}。"
    if latest_template:
        return f"人工上传 {current_version or '当前版本'}。成功则标记上传成功；失败则上传失败报告，先确认修正计划再生成下一版。"
    return "先完成填表，生成 V 后再进入上传反馈循环。"


def _version_timeline(version_infos, uploaded, pending_confirmation=False):
    if not version_infos:
        return [["", "V", "等待生成", "先填入 Amazon 模板。"]]
    timeline = []
    for index, info in enumerate(version_infos):
        version = _version_label(info["name"]) or shared_version_label(index + 1)
        is_last = index == len(version_infos) - 1
        state = "success" if uploaded and is_last else ("current" if is_last else "")
        title = "上传成功" if uploaded and is_last else ("当前待上传版本" if is_last else "历史版本")
        body = "最终成功模板已收录学习库。" if uploaded and is_last else f"{info['name']} 已生成。"
        timeline.append([state, version, title, body])
    if not uploaded:
        if pending_confirmation:
            timeline.append(["current", "失败报告", "等待人工确认", "先看失败原因分析和准备修改的字段，再生成下一版。"])
        else:
            timeline.append(["", "成功/失败", "等待人工上传反馈", "成功则结束并学习；失败则上传失败报告进入修正确认。"])
    return timeline


def _learning_rows_for_product(product_name):
    path = ROOT / "data" / "error_learnings.json"
    if not path.exists():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []
    rows = []
    for record in reversed(data.get("records", [])):
        if product_name and record.get("product_name") and product_name not in record.get("product_name", ""):
            continue
        rows.append([
            record.get("field") or record.get("code") or "未识别字段",
            record.get("message") or "报告已记录",
            record.get("field") or "待判断",
            record.get("rule_hint") or "待归纳",
            "加入失败复盘库",
        ])
        if len(rows) >= 8:
            break
    return rows


def _feedback_analysis_cards(project_dir, failure_report, fix_report, planned_version):
    cards = []
    report_path = Path(failure_report["path"]) if failure_report and failure_report.get("path") else None
    if report_path and report_path.exists() and report_path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            extracted = extract_processing_summary(report_path)
        except Exception as exc:
            extracted = {"summary": {}, "by_code": [], "by_sku": [], "error": str(exc)}
        summary = extracted.get("summary") or {}
        if summary:
            cards.append([
                "warn",
                "上传结果",
                f"处理 {summary.get('processed', '-')} 个 SKU，成功 {summary.get('successful', '-')} 个，失败 {summary.get('unsuccessful', '-')} 个，警告 {summary.get('warnings', '-')} 个。",
            ])
        for item in (extracted.get("by_sku") or [])[:6]:
            cards.append([
                "error",
                f"{item.get('sku') or 'SKU'} / {item.get('field') or '未识别字段'}",
                f"错误码 {item.get('code') or '-'}：{item.get('message') or '报告未提供原因'}",
            ])
        if not cards:
            cards.append([
                "warn",
                "报告未能结构化识别",
                "系统没有读到标准错误行，请打开失败报告核对具体字段；确认后再生成下一版。",
            ])

    plan_lines = _markdown_section_lines(fix_report, "准备修改")
    if plan_lines:
        cards.append([
            "warn",
            "准备修改",
            "；".join(plan_lines[:5]),
        ])
    elif planned_version:
        cards.append([
            "warn",
            "准备修改",
            f"确认失败原因后生成 {shared_version_label(planned_version)}，不自动越过人工确认。",
        ])

    confirm_lines = _markdown_section_lines(fix_report, "人工确认")
    if confirm_lines:
        cards.append(["ok", "人工确认", "；".join(confirm_lines[:3])])
    return cards


def _feedback_analysis_text(project_dir, failure_report, fix_report, planned_version):
    project_dir = Path(project_dir)
    report_path = Path(failure_report["path"]) if failure_report and failure_report.get("path") else None
    extracted = {"summary": {}, "by_sku": []}
    if report_path and report_path.exists() and report_path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            extracted = extract_processing_summary(report_path)
        except Exception:
            extracted = {"summary": {}, "by_sku": []}

    errors = extracted.get("by_sku") or []
    primary = _primary_feedback_error(errors)
    context = _feedback_error_context(project_dir, primary)
    product_context = _feedback_product_context(project_dir)
    if primary:
        sku = primary.get("sku") or "当前 SKU"
        field = primary.get("field") or "未识别字段"
        message = primary.get("message") or "Amazon 报告未提供具体原因。"
        why = _plain_error_reason(message)
        where = f"{sku} 的 {field}"
    elif report_path:
        why = "系统没有从失败报告中识别到明确字段错误。"
        where = f"失败报告：{report_path.name}"
    else:
        why = "还没有上传失败报告。"
        where = "等待 Amazon processing summary。"

    plan_lines = _markdown_section_lines(fix_report, "准备修改")
    if plan_lines:
        fix = _clean_inline_text(plan_lines[0])
    elif primary and "product_type" in (primary.get("message", "") + primary.get("field", "")).lower():
        fix = "补齐正确的 product_type / 产品类型后生成下一版。"
    elif planned_version:
        fix = f"确认失败原因后生成 {shared_version_label(planned_version)}。"
    else:
        fix = "等待失败原因确认后再生成下一版。"

    details = []
    if primary:
        details.extend([
            ["SKU", primary.get("sku") or "-"],
            ["报错码", primary.get("code") or "-"],
            ["报错信息", primary.get("message") or "-"],
            ["报告定位", context.get("reported_cell") or primary.get("field") or "-"],
            ["当前值", context.get("current_value") or "空"],
            ["对应字段", context.get("field_name") or _field_name_from_error(primary) or "-"],
        ])
    if context.get("actual_cell") and context.get("actual_cell") != context.get("reported_cell"):
        details.append(["当前模板实际字段位置", context["actual_cell"]])

    notes = []
    valid_type = context.get("template_product_type")
    if valid_type:
        notes.append(f"这个模板里的有效产品类型目前看是 `{valid_type}`。")
    title = product_context.get("title")
    item_type = product_context.get("item_type_keyword")
    if title or item_type:
        notes.append(f"商品资料显示商品是 `{title or '未识别标题'}`，产品类型关键词填了 `{item_type or '空'}`。")
    if valid_type and (title or item_type):
        notes.append(f"需要判断 `{valid_type}` 这个模板/类目是否真的适合这个商品；如果类目不匹配，补完产品类型后还可能继续报类目或属性不匹配。")

    conclusion = why
    if valid_type:
        conclusion = f"直接原因是产品类型为空；更大的风险是当前模板是 `{valid_type}`，可能和商品类目不匹配。"

    return {
        "why": why,
        "where": where,
        "fix": fix,
        "headline": _feedback_headline(context, primary),
        "details": details,
        "notes": notes,
        "conclusion": conclusion,
    }


def _feedback_headline(context, primary):
    field_label = context.get("field_label") or "产品类型"
    cell = context.get("reported_cell") or context.get("actual_cell") or ""
    if cell:
        return f"报错核心是：{cell} 缺少必填字段 {field_label}。"
    if primary:
        return f"报错核心是：缺少必填字段 {field_label}。"
    return "等待失败报告后生成解析。"


def _feedback_error_context(project_dir, error):
    context = {}
    if not error:
        return context
    reported_cell = _cell_ref_from_text(error.get("field", ""))
    if reported_cell:
        context["reported_cell"] = f"模板!{reported_cell}"

    field_name = _field_name_from_error(error)
    context["field_name"] = field_name
    context["field_label"] = _field_label_from_name(field_name) if field_name else _field_label_from_error(error)

    template_path = _latest_filled_template(project_dir, load_project_status(project_dir))
    if not template_path or not Path(template_path).exists():
        return context
    try:
        wb = load_workbook(template_path, data_only=True, read_only=True, keep_vba=Path(template_path).suffix.lower() == ".xlsm")
        ws = find_template_sheet(wb)
        if ws is None:
            return context
        context["template_product_type"] = _template_product_type(ws)
        if field_name:
            for col in range(1, ws.max_column + 1):
                if str(ws.cell(5, col).value or "").strip() == field_name:
                    cell = f"模板!{get_column_letter(col)}7"
                    value = ws.cell(7, col).value
                    context["actual_cell"] = cell
                    context["current_value"] = _cell_text(value)
                    label = _cell_text(ws.cell(4, col).value)
                    if label:
                        context["field_label"] = label
                    break
        if "current_value" not in context and reported_cell:
            value = ws[reported_cell].value
            context["current_value"] = _cell_text(value)
    except Exception:
        return context
    return context


def _feedback_product_context(project_dir):
    rows, _path, _error = _latest_intake_rows(project_dir, load_project_status(project_dir))
    row = rows[0] if rows else {}
    return {
        "title": row.get("title") or row.get("product_name") or "",
        "item_type_keyword": row.get("item_type_keyword") or "",
        "product_type": row.get("product_type") or "",
    }


def _field_name_from_error(error):
    text = f"{error.get('message', '')} {error.get('field', '')}"
    match = re.search(r"([a-z0-9_]+(?:#[0-9]+)?\.value)", text, re.I)
    if match:
        return match.group(1)
    if "产品类型" in text or "product_type" in text.lower():
        return "product_type#1.value"
    return ""


def _field_label_from_error(error):
    text = f"{error.get('field', '')} {error.get('message', '')}"
    if "产品类型" in text or "product_type" in text.lower():
        return "产品类型"
    return "未识别字段"


def _field_label_from_name(field_name):
    labels = {
        "product_type#1.value": "产品类型",
    }
    return labels.get(field_name, field_name)


def _cell_ref_from_text(value):
    matches = re.findall(r"\(([A-Z]{1,3}[0-9]+)\)", str(value or ""))
    return matches[-1] if matches else ""


def _template_product_type(ws):
    raw_settings = str(ws.cell(1, 1).value or "")
    params = parse_qs(raw_settings)
    ptds = params.get("ptds", [""])[0]
    if ptds:
        try:
            return base64.b64decode(ptds + "=" * (-len(ptds) % 4)).decode("utf-8")
        except Exception:
            pass
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(5, col).value or "").strip() == "product_type#1.value":
            sample = _cell_text(ws.cell(6, col).value)
            if sample:
                return sample
    return ""


def _primary_feedback_error(errors):
    for item in errors:
        text = f"{item.get('field', '')} {item.get('message', '')}".lower()
        if "product_type" in text or "产品类型" in text:
            return item
    return errors[0] if errors else None


def _plain_error_reason(message):
    text = str(message or "")
    if "product_type" in text and "required" in text:
        return "Amazon 要求 product_type / 产品类型必填，v1 没有填。"
    if "missing required fields" in text.lower():
        return "Amazon 认为模板缺少必填字段。"
    return text


def _clean_inline_text(value):
    return str(value or "").strip()


def _markdown_section_lines(file_info, heading):
    if not file_info or not file_info.get("path"):
        return []
    path = Path(file_info["path"])
    if not path.exists() or path.suffix.lower() not in {".md", ".txt"}:
        return []
    lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    captured = []
    in_section = False
    target = f"## {heading}"
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("## "):
            if in_section:
                break
            in_section = stripped == target
            continue
        if in_section and stripped:
            captured.append(stripped.lstrip("- ").strip())
    return captured


def _learning_cards(product_name, uploaded, fix_rows):
    cards = [
        ["成功模板样本", "上传成功后写入成功模板库" if not uploaded else f"{product_name} 已收录"],
        ["失败规则", f"{len(fix_rows)} 条失败记录可用于后续自检"],
        ["下一次动作", "继续上传验证" if not uploaded else "后续同类模板可复用规则"],
    ]
    return cards


def _project_steps(status, rows, latest_template, error_count):
    if rows:
        intake = "已确认"
    else:
        intake = "未开始"
    if latest_template and error_count == 0:
        template = "已通过"
    elif latest_template:
        template = "需修正"
    else:
        template = "待填表"
    if status == "uploaded_success":
        feedback = "已上传"
    elif status == "upload_failed_pending_confirmation":
        feedback = "待确认修正"
    elif latest_template:
        feedback = "等待上传"
    else:
        feedback = "未开始"
    return [intake, template, feedback]


def _project_summary_text(status, latest_template, error_count, next_version, item):
    if status == "uploaded_success":
        uploaded_at = item.get("uploaded_at") or ""
        return f"已上传成功{('：' + uploaded_at) if uploaded_at else ''}"
    if status == "upload_failed_pending_confirmation":
        version = _version_label(Path(latest_template).name) if latest_template else "当前版本"
        pending_next = item.get("pending_next_version") or next_version
        return f"{version or '当前版本'} 上传失败，待确认修正后生成 {shared_version_label(pending_next)}"
    if latest_template:
        version = _version_label(Path(latest_template).name)
        if error_count:
            return f"{version or '模板'} 已生成，自检有 {error_count} 项需处理"
        return f"{version or '模板'} 已生成，等待人工上传 Amazon"
    return f"等待生成 {safe_name(item.get('product_name') or '产品')}_{shared_version_label(next_version)}"


def _format_updated(value):
    if not value:
        return "未更新"
    text = str(value)
    try:
        dt = datetime.fromisoformat(text)
    except ValueError:
        return text[:16]
    return dt.strftime("%Y-%m-%d %H:%M")


def _write_feedback_fix_report(project_dir, report_paths, added_records, next_version=None):
    product_name = safe_name(infer_product_name(project_dir, ""))
    version_label = f"_{shared_version_label(next_version)}" if next_version else ""
    filled_name = f"{product_name}{version_label}_修正计划"
    output_dir = Path(project_dir) / LEGACY_FILLED_TEMPLATE_DIR
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / f"{Path(filled_name).stem}_修正说明.md"
    lines = [
        f"# {Path(filled_name).stem} 修正说明",
        "",
        f"- 生成时间：{datetime.now().isoformat(timespec='seconds')}",
        f"- 失败报告：{', '.join(path.name for path in report_paths)}",
        f"- 当前状态：等待人工确认",
        f"- 准备生成：{product_name}_{shared_version_label(next_version)}.xlsx" if next_version else "- 准备生成：下一版模板",
        f"- 学习记录：{len(added_records)} 条",
        "",
        "## 失败报告识别",
        "",
    ]
    if not added_records:
        lines.append("- 未从报告中识别到结构化错误，已按确认产品资料重新生成下一版模板。")
    else:
        for record in added_records:
            lines.extend([
                f"- 字段：{record.get('field') or record.get('code') or '未识别'}",
                f"  - 报错：{record.get('message') or '无'}",
                f"  - 规则提示：{record.get('rule_hint') or '待归纳'}",
            ])
    lines.extend([
        "",
        "## 准备修改",
        "",
    ])
    if not added_records:
        lines.append("- 未识别到结构化字段错误；确认后会基于已确认产品资料和当前规则重新生成下一版模板。")
    else:
        for record in added_records:
            lines.append(f"- {record.get('field') or record.get('code') or '未识别字段'}：{record.get('rule_hint') or '按失败报告修正'}")
    lines.extend([
        "",
        "## 人工确认",
        "",
        "- 确认以上失败原因和准备修改方向后，在工作台点击“确认修正并生成下一版”。",
    ])
    report_path.write_text("\n".join(lines), encoding="utf-8")
    return report_path


def _append_feedback_generation_result(report_path, fill_result):
    report_path = Path(report_path)
    filled_file = fill_result.get("filled_file") or {}
    report_file = fill_result.get("report_file") or {}
    lines = [
        "",
        "## 已生成下一版",
        "",
        f"- 生成时间：{datetime.now().isoformat(timespec='seconds')}",
        f"- 新模板：{filled_file.get('name') or '未生成'}",
        f"- 自检报告：{report_file.get('name') or '未生成'}",
        f"- 自检错误数：{fill_result.get('error_count', 0)}",
    ]
    with report_path.open("a", encoding="utf-8") as output:
        output.write("\n".join(lines))


def _copy_success_template(project_dir, template_path):
    template_path = Path(template_path)
    if not template_path.exists() or template_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return None
    SUCCESS_TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    product_name = safe_name(infer_product_name(project_dir, ""))
    target_name = f"{product_name}_{template_path.name}" if product_name not in template_path.stem else template_path.name
    target_path = _unique_file_path(SUCCESS_TEMPLATES_DIR, target_name)
    shutil.copy2(template_path, target_path)
    return target_path


def _local_project_file(project_dir, path_value):
    path = Path(path_value)
    if not path.is_absolute():
        path = (ROOT / unquote(str(path_value))).resolve()
        if not path.exists():
            path = (Path(project_dir) / unquote(str(path_value))).resolve()
    else:
        path = path.resolve()
    try:
        path.relative_to(Path(project_dir).resolve())
    except ValueError as exc:
        raise ValueError("文件必须在当前项目内。") from exc
    if not path.exists():
        raise ValueError("文件不存在。")
    return path


def _project_file_candidate(project_dir, path_value):
    if not path_value:
        return None
    raw = unquote(str(path_value))
    path = Path(raw)
    candidates = []
    if path.is_absolute():
        candidates.append(path)
    else:
        candidates.append(Path(project_dir) / path)
        candidates.append(ROOT / path)
    for candidate in candidates:
        candidate = candidate.resolve()
        if candidate.exists():
            return candidate
    return candidates[0].resolve() if candidates else None


def _relative_to_project(project_dir, path):
    project_dir = Path(project_dir).resolve()
    path = Path(path).resolve()
    try:
        return str(path.relative_to(project_dir))
    except ValueError:
        return str(path)


def _project_path(payload):
    value = str(payload.get("project_dir", "")).strip()
    if not value:
        raise ValueError("缺少项目路径。")
    path = Path(value)
    if not path.is_absolute():
        path = ROOT / path
    path = path.resolve()
    try:
        path.relative_to(PROJECTS_DIR.resolve())
    except ValueError as exc:
        raise ValueError("项目路径必须在 data/projects 下。") from exc
    if not path.exists():
        raise ValueError("项目不存在。")
    return path


def _auto_fill_message(result):
    if result.get("skipped"):
        return "项目已上传，已跳过"
    if result.get("blocked"):
        return f"自动填表暂停：{result.get('reason')}"
    if result.get("status") == "ready_for_upload":
        return "自动填表完成，可进入人工上传"
    return "自动填表完成，需要查看自检报告"


def _next_step(item):
    status = item["status"]
    if status == "uploaded_success":
        return "归档"
    if status == "ready_for_upload":
        return "上传 Amazon 后标记成功"
    if status == "upload_failed_pending_confirmation":
        return "确认失败修正计划并生成下一版"
    if status == "needs_manual_fix":
        return "查看自检报告并修正"
    if status == "blocked":
        return item["blocked_reason"] or "补齐资料"
    return "放入资料并自动填表"


def _workflow_health(project_dir, item, latest_draft, latest_template, latest_check_report, folders):
    project_dir = Path(project_dir)
    check_findings = _parse_check_report(latest_check_report["path"] if latest_check_report else "")
    stale_sources = _sources_newer_than(project_dir, latest_draft["path"] if latest_draft else "")
    legacy_folders = [
        folder for folder in folders
        if folder.get("legacy") and folder.get("file_count") and folder.get("name") in LEGACY_SOURCE_FOLDERS
    ]
    suspicious_files = _suspicious_files(project_dir, item.get("product_name", ""))
    blockers = []
    warnings = []

    is_uploaded = item.get("status") == "uploaded_success"
    if is_uploaded:
        stale_sources = []
        suspicious_files = []
    if not latest_draft and not is_uploaded:
        blockers.append("还没有自动提炼草稿")
    if not latest_template and not is_uploaded:
        blockers.append("还没有填表版本")
    if item.get("status") == "needs_manual_fix" and check_findings:
        blockers.append(f"模板自检还有 {len(check_findings)} 个错误")
    if stale_sources and not is_uploaded:
        warnings.append(f"有 {len(stale_sources)} 个资料晚于草稿，建议重新自动填表")
    if legacy_folders:
        warnings.append("存在旧目录资料，资料架已合并显示")
    if suspicious_files and not is_uploaded:
        warnings.append(f"发现 {len(suspicious_files)} 个疑似错放资料")

    next_action = _health_next_action(item, blockers, warnings, stale_sources, check_findings)
    return {
        "level": "error" if blockers else ("warning" if warnings else "ok"),
        "blockers": blockers,
        "warnings": warnings,
        "next_action": next_action,
        "check_findings": check_findings[:8],
        "stale_sources": stale_sources[:8],
        "legacy_folder_count": len(legacy_folders),
        "suspicious_files": suspicious_files[:8],
    }


def _health_next_action(item, blockers, warnings, stale_sources, check_findings):
    status = item.get("status")
    if status == "uploaded_success":
        return "已上传成功，除非要复盘，不建议重新生成"
    if stale_sources:
        return "资料更新晚于草稿，先重新自动填表"
    if check_findings:
        first = check_findings[0]
        return f"先修 {first.get('field')}: {first.get('message')}"
    if blockers:
        return blockers[0]
    if status == "ready_for_upload":
        return "上传 Amazon，上传成功后标记成功"
    if status == "upload_failed_pending_confirmation":
        return "确认失败原因和修改计划后生成下一版"
    return item.get("blocked_reason") or item.get("next_step") or "补齐资料并自动填表"


def _parse_check_report(path_value):
    if not path_value:
        return []
    path = Path(path_value)
    if not path.exists() or not path.is_file():
        return []
    text = path.read_text(encoding="utf-8", errors="ignore")
    pattern = re.compile(
        r"### \[ERROR\] 行 (?P<row>.*?) - (?P<field>.*?)\n\n"
        r"- 问题：(?P<message>.*?)\n"
        r"- 建议：(?P<fix>.*?)(?:\n\n|$)",
        re.S,
    )
    findings = []
    for match in pattern.finditer(text):
        findings.append({
            "row": match.group("row").strip(),
            "field": match.group("field").strip(),
            "message": _single_line(match.group("message")),
            "fix": _single_line(match.group("fix")),
        })
    return findings


def _sources_newer_than(project_dir, draft_path_value):
    if not draft_path_value:
        return []
    draft_path = Path(draft_path_value)
    if not draft_path.exists():
        return []
    draft_mtime = draft_path.stat().st_mtime
    newer = []
    for path in _iter_project_files(project_dir):
        if path.parent.name not in SOURCE_REVIEW_FOLDERS:
            continue
        if path.suffix.lower() not in READABLE_SUFFIXES:
            continue
        if path.stat().st_mtime <= draft_mtime:
            continue
        if _is_generated_file(path):
            continue
        info = _file_info(project_dir, path)
        if info:
            newer.append(info)
    return sorted(newer, key=lambda item: item["path"], reverse=True)


def _suspicious_files(project_dir, product_name):
    product_tokens = _name_tokens(product_name)
    if not product_tokens:
        return []
    suspicious = []
    for path in _iter_project_files(project_dir):
        if path.parent.name not in SOURCE_REVIEW_FOLDERS:
            continue
        if path.suffix.lower() not in {".html", ".htm", ".xlsx", ".xlsm", ".xls", ".pdf", ".png", ".jpg", ".jpeg"}:
            continue
        haystack = path.name.lower()
        if any(token in haystack for token in product_tokens):
            continue
        mismatch_tokens = ["spray", "bottle", "玻璃", "洒水", "喷壶", "yoga", "瑜伽", "glove", "手套", "floral", "花束"]
        if any(token in haystack for token in mismatch_tokens):
            info = _file_info(project_dir, path)
            if info:
                suspicious.append(info)
    return suspicious


def _name_tokens(value):
    lowered = str(value or "").lower()
    tokens = [token for token in re.split(r"[^a-z0-9\u4e00-\u9fff]+", lowered) if len(token) >= 2]
    if "瑜伽砖" in lowered:
        tokens.extend(["瑜伽", "yoga"])
    if "花束卡片夹" in lowered:
        tokens.extend(["花束", "floral"])
    if "手套" in lowered:
        tokens.extend(["手套", "glove"])
    if "洒水壶" in lowered or "喷壶" in lowered:
        tokens.extend(["洒水", "喷壶", "bottle"])
    return sorted(set(tokens))


def _iter_project_files(project_dir):
    for path in Path(project_dir).rglob("*"):
        if path.is_file() and not path.name.startswith("~$") and path.name != ".DS_Store":
            yield path


def _is_generated_file(path):
    generated_names = ["自动提炼草稿", "资料提炼报告", "模板自检报告", "写入报告"]
    if any(name in path.name for name in generated_names):
        return True
    return path.parent.name in {"04_模板原件", "05_填表版本", "05_模版原件", "06_处理报告"}


def _single_line(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _has_template(project_dir):
    for folder in TEMPLATE_DIRS:
        path = Path(project_dir) / folder
        if path.exists() and any(item.suffix.lower() in {".xlsx", ".xlsm"} for item in path.iterdir() if item.is_file()):
            return True
    return False


def _has_draft(project_dir):
    for folder in DRAFT_DIRS:
        path = Path(project_dir) / folder
        if path.exists() and any("自动提炼草稿" in item.name for item in path.iterdir() if item.is_file()):
            return True
    return False


def _infer_latest_draft(project_dir):
    candidates = []
    for folder in DRAFT_DIRS:
        path = Path(project_dir) / folder
        if not path.exists():
            continue
        candidates.extend([
            item for item in path.iterdir()
            if item.is_file() and "自动提炼草稿" in item.name and item.suffix.lower() == ".xlsx"
        ])
    if not candidates:
        return ""
    return str(sorted(candidates, key=lambda item: item.stat().st_mtime, reverse=True)[0])


def _project_folder_infos(project_dir):
    project_dir = Path(project_dir)
    folders = []
    display_folders = list(PROJECT_FOLDERS)
    for folder in LEGACY_DISPLAY_FOLDERS:
        if folder not in display_folders and (project_dir / folder).exists():
            display_folders.append(folder)
    for folder in display_folders:
        folder_path = project_dir / folder
        files = []
        if folder_path.exists():
            for path in sorted(folder_path.iterdir(), key=lambda item: item.stat().st_mtime if item.exists() else 0, reverse=True):
                if not path.is_file() or path.name.startswith("~$"):
                    continue
                info = _file_info(project_dir, path)
                if info:
                    files.append(info)
        folders.append({
            "name": folder,
            "path": str(folder_path),
            "relative_path": _relative(folder_path),
            "file_count": len(files),
            "files": files,
            "legacy": folder not in PROJECT_FOLDERS,
        })
    return folders


def _file_info(project_dir, path_value):
    if not path_value:
        return None
    project_dir = Path(project_dir)
    path = Path(path_value)
    if not path.is_absolute():
        path = project_dir / path
    path = path.resolve()
    try:
        relative_to_root = path.relative_to(ROOT.resolve())
    except ValueError:
        return None
    exists = path.exists() and path.is_file()
    size_bytes = path.stat().st_size if exists else 0
    return {
        "name": path.name,
        "path": str(path),
        "relative_path": str(relative_to_root),
        "folder": str(path.parent),
        "folder_relative": str(path.parent.relative_to(ROOT.resolve())),
        "download_url": f"/file?path={quote(str(relative_to_root))}",
        "size_bytes": size_bytes,
        "exists": exists,
    }


def _local_file_path(path_value):
    raw_path = str(path_value or "")
    if not raw_path:
        return ROOT / "__missing__"
    path = (ROOT / unquote(raw_path)).resolve()
    allowed_roots = [PROJECTS_DIR.resolve(), OUTPUTS_DIR.resolve()]
    if not any(path == root or root in path.parents for root in allowed_roots):
        return ROOT / "__forbidden__"
    return path


def _served_file_path(query):
    params = parse_qs(query)
    return _local_file_path(params.get("path", [""])[0])


def _download_content_type(path):
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if suffix == ".md":
        return "text/markdown; charset=utf-8"
    if suffix == ".json":
        return "application/json; charset=utf-8"
    return "application/octet-stream"


def _relative(path):
    try:
        return str(Path(path).relative_to(ROOT))
    except ValueError:
        return str(path)


def _html():
    return (Path(__file__).with_name("workbench_frontend.html")).read_text(encoding="utf-8")
