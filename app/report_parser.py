import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook

from .paths import CONFIG_DIR


CODE_RE = re.compile(r"\b\d{4,6}\b")


def load_error_codes():
    return json.loads((CONFIG_DIR / "error_codes.json").read_text(encoding="utf-8"))


def parse_processing_summary(path):
    path = Path(path)
    text_chunks = []

    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, data_only=True, read_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                values = [str(value) for value in row if value not in (None, "")]
                if values:
                    text_chunks.append(" | ".join(values))
    elif path.suffix.lower() in {".csv", ".txt", ".tsv"}:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", errors="ignore") as handle:
            if path.suffix.lower() == ".txt":
                text_chunks.extend(handle.readlines())
            else:
                reader = csv.reader(handle, delimiter=delimiter)
                text_chunks.extend(" | ".join(row) for row in reader)
    else:
        text_chunks.append(path.read_text(encoding="utf-8", errors="ignore"))

    error_codes = load_error_codes()
    findings = []
    seen = set()
    for line in text_chunks:
        for code in CODE_RE.findall(line):
            key = (code, line.strip())
            if key in seen:
                continue
            seen.add(key)
            info = error_codes.get(code, {})
            findings.append({
                "code": code,
                "type": info.get("type", "未知错误码"),
                "meaning": info.get("meaning", "配置里还没有这个错误码的解释。"),
                "fix": info.get("fix", "先查看 processing summary 原文定位字段。"),
                "source": line.strip()
            })
    return findings


def extract_processing_summary(path):
    path = Path(path)
    workbook = load_workbook(path, data_only=True, read_only=True, keep_vba=path.suffix.lower() == ".xlsm")
    if "Feed Processing Summary" not in workbook.sheetnames:
        return {
            "path": str(path),
            "summary": {},
            "by_code": [],
            "by_sku": []
        }

    ws = workbook["Feed Processing Summary"]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(value).strip() if value is not None else "" for value in row])

    summary = {}
    by_code = []
    by_sku = []
    section = None
    code_header = None
    sku_header = None

    for row in rows:
        compact = [value for value in row if value]
        if not compact:
            continue

        joined = " | ".join(compact)
        if "Errors and Warnings per Error Code" in joined:
            section = "by_code"
            continue
        if "Errors and Warnings per SKU" in joined:
            section = "by_sku"
            continue

        if "Number of SKUs processed" in compact:
            summary["processed"] = compact[-1]
        elif "Number of SKUs successful" in compact and "other errors" not in joined:
            summary["successful"] = compact[-1]
        elif "Number of SKUs unsuccessful due to errors" in compact:
            summary["unsuccessful"] = compact[-1]
        elif "Number of SKUs with warnings" in compact:
            summary["warnings"] = compact[-1]

        if section == "by_code":
            if "Error code" in compact and "Error message" in compact:
                code_header = compact
                continue
            if code_header and len(compact) >= 5 and compact[1].isdigit():
                item = dict(zip(code_header, compact))
                by_code.append({
                    "code": item.get("Error code", ""),
                    "category": item.get("Category of error", ""),
                    "message": item.get("Error message", ""),
                    "field": item.get("Affected field (Impacted column)", ""),
                    "count": item.get("Number of errors", "")
                })

        if section == "by_sku":
            if "Error code" in compact and "SKU" in compact:
                sku_header = compact
                continue
            if sku_header and len(compact) >= 6 and compact[1].isdigit():
                item = dict(zip(sku_header, compact))
                by_sku.append({
                    "code": item.get("Error code", ""),
                    "category": item.get("Category of error", ""),
                    "message": item.get("Error message", ""),
                    "field": item.get("Affected field (Impacted cell)", ""),
                    "sku": item.get("SKU", "")
                })

    return {
        "path": str(path),
        "summary": summary,
        "by_code": by_code,
        "by_sku": by_sku
    }
