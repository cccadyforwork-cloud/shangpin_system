import re
from copy import deepcopy
from pathlib import Path
from shutil import copyfile

from openpyxl import load_workbook

from .paths import DRAFT_DIRS, TEMPLATE_DIRS, safe_name
from .success_rule_defaults import load_safe_defaults
from .template_sheet import find_template_sheet, template_sheet_names_text
from .template_validator import PRODUCT_TYPE_CONDITIONAL_FIELDS
from .versioning import versioned_template_path
from .workbook_io import read_intake_rows


DRAFT_PATTERN = "*_自动提炼草稿.xlsx"
TEMPLATE_EXAMPLE_ROW = 6
TEMPLATE_DATA_START_ROW = 7


FIELD_MAP = {
    "contribution_sku#1.value": "sku",
    "product_type#1.value": "product_type",
    "::record_action": "__listing_action",
    "item_name[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value": "title",
    "brand[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value": "brand",
    "item_type_keyword[marketplace_id=ATVPDKIKX0DER]#1.value": "item_type_keyword",
    "parentage_level[marketplace_id=ATVPDKIKX0DER]#1.value": "parentage_level",
    "child_parent_sku_relationship[marketplace_id=ATVPDKIKX0DER]#1.parent_sku": "parent_sku",
    "child_parent_sku_relationship[marketplace_id=ATVPDKIKX0DER]#1.child_relationship_type": "__variation_relationship",
    "variation_theme#1.name": "variation_theme",
    "model_number[marketplace_id=ATVPDKIKX0DER]#1.value": "sku",
    "manufacturer[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value": "manufacturer",
    "main_product_image_locator[marketplace_id=ATVPDKIKX0DER]#1.media_location": "main_image_url",
    "product_description[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value": "description",
    "bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value": "bullet_1",
    "bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#2.value": "bullet_2",
    "bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#3.value": "bullet_3",
    "bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#4.value": "bullet_4",
    "bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#5.value": "bullet_5",
    "material[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value": "material",
    "number_of_items[marketplace_id=ATVPDKIKX0DER]#1.value": "set_count",
    "item_package_quantity[marketplace_id=ATVPDKIKX0DER]#1.value": "__package_quantity",
    "color[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value": "color",
    "size[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value": "size",
    "part_number[marketplace_id=ATVPDKIKX0DER]#1.value": "sku",
    "list_price[marketplace_id=ATVPDKIKX0DER]#1.value": "list_price",
    "purchasable_offer[marketplace_id=ATVPDKIKX0DER][audience=BZR]#1.our_price#1.schedule#1.value_with_tax": "haul_price",
    "item_depth_width_height[marketplace_id=ATVPDKIKX0DER]#1.depth.value": "package_height_in",
    "item_depth_width_height[marketplace_id=ATVPDKIKX0DER]#1.depth.unit": "__title_inches",
    "item_depth_width_height[marketplace_id=ATVPDKIKX0DER]#1.height.value": "package_length_in",
    "item_depth_width_height[marketplace_id=ATVPDKIKX0DER]#1.height.unit": "__title_inches",
    "item_depth_width_height[marketplace_id=ATVPDKIKX0DER]#1.width.value": "package_width_in",
    "item_depth_width_height[marketplace_id=ATVPDKIKX0DER]#1.width.unit": "__title_inches",
    "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.value": "package_length_in",
    "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.unit": "__inches",
    "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.value": "package_width_in",
    "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.unit": "__inches",
    "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.value": "package_height_in",
    "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.unit": "__inches",
    "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.value": "package_weight_lb",
    "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.unit": "__pounds",
    "country_of_origin[marketplace_id=ATVPDKIKX0DER]#1.value": "country_of_origin",
    "batteries_required[marketplace_id=ATVPDKIKX0DER]#1.value": "batteries_required",
    "batteries_included[marketplace_id=ATVPDKIKX0DER]#1.value": "__no",
    "supplier_declared_dg_hz_regulation[marketplace_id=ATVPDKIKX0DER]#1.value": "__dg_not_applicable",
    "condition_type[marketplace_id=ATVPDKIKX0DER]#1.value": "__new_condition",
    "glove_or_mitt[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value": "__glove",
    "glove_liner_material[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value": "__cotton",
    "base_coating_material[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value": "__latex",
    "palm_style[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value": "__coated_grip",
}


ANIMAL_COLLAR_DEFAULT_FIELDS = {
    "item_type_keyword": "item_type_keyword[marketplace_id=ATVPDKIKX0DER]#1.value",
    "color_map": "color[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.standardized_values#1",
    "model_name": "model_name[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    "style": "style[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    "number_of_items": "number_of_items[marketplace_id=ATVPDKIKX0DER]#1.value",
    "unit_count_value": "unit_count[marketplace_id=ATVPDKIKX0DER]#1.value",
    "unit_count_type": "unit_count[marketplace_id=ATVPDKIKX0DER]#1.type[language_tag=en_US].value",
    "included_components": "included_components[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    "specific_use_1": "specific_uses_for_product[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    "specific_use_2": "specific_uses_for_product[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#2.value",
    "care_instructions": "care_instructions[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    "pattern": "pattern[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    "closure": "closure[marketplace_id=ATVPDKIKX0DER]#1.type[language_tag=en_US]#1.value",
    "item_length": "item_length[marketplace_id=ATVPDKIKX0DER]#1.value",
    "item_length_unit": "item_length[marketplace_id=ATVPDKIKX0DER]#1.unit",
    "item_width": "item_width[marketplace_id=ATVPDKIKX0DER]#1.value",
    "item_width_unit": "item_width[marketplace_id=ATVPDKIKX0DER]#1.unit",
    "item_length_width_length": "item_length_width[marketplace_id=ATVPDKIKX0DER]#1.length.value",
    "item_length_width_length_unit": "item_length_width[marketplace_id=ATVPDKIKX0DER]#1.length.unit",
    "item_length_width_width": "item_length_width[marketplace_id=ATVPDKIKX0DER]#1.width.value",
    "item_length_width_width_unit": "item_length_width[marketplace_id=ATVPDKIKX0DER]#1.width.unit",
    "number_of_packs": "number_of_packs[marketplace_id=ATVPDKIKX0DER]#1.value",
    "pet_type_1": "pet_type[marketplace_id=ATVPDKIKX0DER]#1.value",
    "pet_type_2": "pet_type[marketplace_id=ATVPDKIKX0DER]#2.value",
    "animal_collar_type": "animal_collar_type[marketplace_id=ATVPDKIKX0DER]#1.value",
    "number_of_boxes": "number_of_boxes[marketplace_id=ATVPDKIKX0DER]#1.value",
    "directions": "directions[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
}

COLOR_MAP_VALUES = {
    "black": "Black",
    "white": "White",
    "red": "Red",
    "blue": "Blue",
    "green": "Green",
    "pink": "Pink",
    "gray": "Gray",
    "grey": "Gray",
    "yellow": "Yellow",
    "purple": "Purple",
    "orange": "Orange",
    "transparent": "Transparent",
    "黑": "Black",
    "白": "White",
    "红": "Red",
    "蓝": "Blue",
    "绿": "Green",
    "粉": "Pink",
    "灰": "Gray",
    "黄": "Yellow",
    "紫": "Purple",
    "橙": "Orange",
    "透明": "Transparent",
}

EXPLICIT_SINGLE_ROUTE_KEYWORDS = (
    "single",
    "single link",
    "single-link",
    "单链接",
    "独立",
    "single_child",
)
EXPLICIT_BUNDLE_ROUTE_KEYWORDS = (
    "bundle",
    "set bundle",
    "套装售卖",
    "组合售卖",
)
PARENT_CLEAR_FIELDS = {
    "parent_sku",
    "color",
    "size",
    "cost",
    "target_price",
    "list_price",
    "haul_price",
    "package_length_in",
    "package_width_in",
    "package_height_in",
    "package_weight_lb",
    "main_image_url",
}


def find_latest_draft(project_dir):
    project_dir = Path(project_dir)
    candidates = []
    for folder in DRAFT_DIRS:
        draft_dir = project_dir / folder
        if draft_dir.exists():
            candidates.extend(draft_dir.glob(DRAFT_PATTERN))
    candidates = sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError(f"没有找到自动提炼草稿：{project_dir}")
    return candidates[0]


def find_template(project_dir):
    project_dir = Path(project_dir)
    candidates = []
    for folder in TEMPLATE_DIRS:
        template_dir = project_dir / folder
        if not template_dir.exists():
            continue
        for path in template_dir.glob("*"):
            if path.name.startswith("~$"):
                continue
            if path.suffix.lower() in {".xlsx", ".xlsm"} and "自动提炼草稿" not in path.name and "产品资料" not in path.name:
                candidates.append(path)
    if not candidates:
        raise FileNotFoundError(f"没有找到亚马逊模板，请放到 05_模版原件：{project_dir}")
    return sorted(candidates, key=lambda p: (TEMPLATE_DIRS.index(p.parent.name), p.name))[0]


def extract_template_product_type(template_path):
    template_path = Path(template_path)
    wb = load_workbook(template_path, data_only=True, read_only=True, keep_vba=template_path.suffix.lower() == ".xlsm")
    if "Valid Values" in wb.sheetnames:
        ws = wb["Valid Values"]
        for row in ws.iter_rows(values_only=True):
            values = [str(value).strip() if value is not None else "" for value in row]
            for idx, value in enumerate(values):
                if value.startswith("Product Type"):
                    for candidate in values[idx + 1:]:
                        if candidate and not candidate.startswith("["):
                            return candidate
    return ""


def _value_for(row, source):
    is_parent = str(row.get("parentage_level") or "").strip().lower() == "parent"
    if is_parent and source in {"__package_quantity", "__inches", "__title_inches", "__pounds"}:
        return ""
    if source == "__listing_action":
        return "Create or Replace (Full Update)"
    if source == "__variation_relationship":
        return "Variation" if row.get("parent_sku") else ""
    if source == "__package_quantity":
        return row.get("set_count") or 1
    if source == "__inches":
        return "Inches"
    if source == "__title_inches":
        return "Inches"
    if source == "__pounds":
        return "Pounds"
    if source == "__no":
        return "No"
    if source == "__dg_not_applicable":
        return "Not Applicable"
    if source == "__new_condition":
        return "New"
    if source == "__glove":
        return "Glove"
    if source == "__cotton":
        return "Cotton"
    if source == "__latex":
        return "Latex"
    if source == "__coated_grip":
        return "Coated with Added Grip"

    value = row.get(source)
    if value is None:
        return ""
    return value


def _text(value):
    return str(value or "").strip()


def _is_parent_row(row):
    return _text(row.get("parentage_level")).lower() == "parent"


def _is_child_row(row):
    return _text(row.get("parentage_level")).lower() == "child"


def _field_texts(rows, field):
    values = []
    for row in rows:
        if _is_parent_row(row):
            continue
        value = _text(row.get(field))
        if value:
            values.append(value)
    return values


def _infer_variation_theme(rows):
    colors = set(_field_texts(rows, "color"))
    sizes = set(_field_texts(rows, "size"))
    if len(colors) > 1 and len(sizes) > 1:
        return "ColorSize"
    if len(sizes) > 1 and len(colors) <= 1:
        return "Size"
    if len(colors) > 1:
        return "Color"
    if sizes:
        return "Size"
    return "Color"


def _sku_base(value):
    words = re.findall(r"[A-Za-z0-9]+", str(value or "").upper())
    if words:
        return "-".join(words[:4])[:24]
    return safe_name(str(value or "PRODUCT")).upper()[:24] or "PRODUCT"


def _parent_sku_for_rows(rows, project_dir):
    for row in rows:
        parent_sku = _text(row.get("parent_sku"))
        if parent_sku:
            return parent_sku
    for row in rows:
        if _is_parent_row(row) and _text(row.get("sku")):
            return _text(row.get("sku"))
    product_name = next((_text(row.get("product_name")) for row in rows if _text(row.get("product_name"))), "")
    return f"{_sku_base(product_name or Path(project_dir).name)}-PARENT"


def _should_default_variation(rows):
    route_text = "\n".join(_text(row.get("route")).lower() for row in rows)
    if any(keyword in route_text for keyword in EXPLICIT_BUNDLE_ROUTE_KEYWORDS):
        return False
    if any(keyword in route_text for keyword in EXPLICIT_SINGLE_ROUTE_KEYWORDS):
        return False
    return True


def prepare_variation_rows(rows, project_dir):
    prepared = [dict(row) for row in rows]
    if not prepared:
        return prepared

    theme = next((_text(row.get("variation_theme")) for row in prepared if _text(row.get("variation_theme"))), "")
    theme = theme or _infer_variation_theme(prepared)
    parent_sku = _parent_sku_for_rows(prepared, project_dir)

    if any(_is_parent_row(row) for row in prepared):
        for row in prepared:
            row["variation_theme"] = _text(row.get("variation_theme")) or theme
            if _is_child_row(row) and not _text(row.get("parent_sku")):
                row["parent_sku"] = parent_sku
            if _text(row.get("parent_sku")) and not _text(row.get("parentage_level")):
                row["parentage_level"] = "Child"
        return prepared

    if not _should_default_variation(prepared):
        return prepared

    parent = deepcopy(prepared[0])
    parent["sku"] = parent_sku
    parent["parent_sku"] = ""
    parent["parentage_level"] = "Parent"
    parent["variation_theme"] = theme
    parent["route"] = "Haul Generic Variation" if "generic" in _text(parent.get("route")).lower() or not _text(parent.get("route")) else parent.get("route")
    for field in PARENT_CLEAR_FIELDS:
        parent[field] = ""

    child_rows = []
    for index, row in enumerate(prepared, 1):
        row["parent_sku"] = parent_sku
        row["parentage_level"] = "Child"
        row["variation_theme"] = theme
        if "generic" in _text(row.get("route")).lower() or not _text(row.get("route")):
            row["route"] = "Haul Generic Variation"
        if not _text(row.get("sku")):
            row["sku"] = f"{_sku_base(row.get('product_name') or Path(project_dir).name)}-{index:03d}"
        child_rows.append(row)
    return [parent] + child_rows


def _row_text(row):
    return "\n".join(str(value or "") for value in row.values())


def _normalize_color_map(value):
    text = str(value or "").strip()
    if not text:
        return ""
    lowered = text.lower()
    for key, color in COLOR_MAP_VALUES.items():
        if key.lower() in lowered or key in text:
            return color
    return text


def _first_positive_int(value, fallback=1):
    match = re.search(r"\d+", str(value or ""))
    if not match:
        return fallback
    number = int(match.group(0))
    return number if number > 0 else fallback


def _animal_collar_dimensions(row):
    text = _row_text(row)
    matches = list(re.finditer(
        r"([0-9]+(?:\.[0-9]+)?)\s*(cm|mm|in|inch|inches)?\s*[xX*×]\s*([0-9]+(?:\.[0-9]+)?)\s*(cm|mm|in|inch|inches)\b",
        text,
        re.I,
    ))
    if not matches:
        return "", ""

    def to_inches(number, unit):
        value = float(number)
        unit = unit.lower()
        if unit == "cm":
            value /= 2.54
        elif unit == "mm":
            value /= 25.4
        return f"{value:.2f}".rstrip("0").rstrip(".")

    match = max(matches, key=lambda item: float(item.group(1)) * float(item.group(3)))
    first_unit = match.group(2) or match.group(4)
    return to_inches(match.group(1), first_unit), to_inches(match.group(3), match.group(4))


def _animal_collar_rule_defaults(row):
    if str(row.get("product_type") or "").strip().upper() != "ANIMAL_COLLAR":
        return {}, set()

    text = _row_text(row)
    lowered = text.lower()
    count = _first_positive_int(row.get("set_count"), 1)
    color_map = _normalize_color_map(row.get("color"))
    is_airtag = "airtag" in lowered
    is_glow = any(keyword in lowered for keyword in ["glow", "luminous", "夜光", "发光", "reflective"])
    is_cat = any(keyword in lowered for keyword in ["cat", "cats", "猫", "猫咪"])
    is_dog = any(keyword in lowered for keyword in ["dog", "dogs", "狗", "狗狗"])
    has_heart = any(keyword in lowered for keyword in ["heart", "hearts", "爱心"])
    length, width = _animal_collar_dimensions(row)

    model_name = "AirTag Pet Collar" if is_airtag else "Pet Collar"
    if is_airtag and is_glow:
        style = "Glow in the Dark AirTag Holder Collar"
    elif is_airtag:
        style = "AirTag Holder Collar"
    elif is_glow:
        style = "Glow in the Dark Pet Collar"
    else:
        style = "Pet Collar"

    directions = "Adjust collar to fit the pet comfortably before use. Check the fit regularly and remove the collar if it becomes damaged. AirTag is not included." if is_airtag else "Adjust collar to fit the pet comfortably before use. Check the fit regularly and remove the collar if it becomes damaged."

    defaults = {
        ANIMAL_COLLAR_DEFAULT_FIELDS["item_type_keyword"]: "pet-collars",
        ANIMAL_COLLAR_DEFAULT_FIELDS["model_name"]: model_name,
        ANIMAL_COLLAR_DEFAULT_FIELDS["style"]: style,
        ANIMAL_COLLAR_DEFAULT_FIELDS["number_of_items"]: count,
        ANIMAL_COLLAR_DEFAULT_FIELDS["unit_count_value"]: count,
        ANIMAL_COLLAR_DEFAULT_FIELDS["unit_count_type"]: "Count",
        ANIMAL_COLLAR_DEFAULT_FIELDS["included_components"]: "Safety Clip" if is_cat or "breakaway" in lowered else "1 x Pet Collar",
        ANIMAL_COLLAR_DEFAULT_FIELDS["specific_use_1"]: "Outdoor" if is_airtag or is_glow else "Active",
        ANIMAL_COLLAR_DEFAULT_FIELDS["specific_use_2"]: "Active" if is_airtag or is_glow else "",
        ANIMAL_COLLAR_DEFAULT_FIELDS["care_instructions"]: "Spot Clean",
        ANIMAL_COLLAR_DEFAULT_FIELDS["pattern"]: "Hearts" if has_heart else "Solid",
        ANIMAL_COLLAR_DEFAULT_FIELDS["closure"]: "Break Away" if is_cat or "breakaway" in lowered else "Buckle",
        ANIMAL_COLLAR_DEFAULT_FIELDS["number_of_packs"]: 1,
        ANIMAL_COLLAR_DEFAULT_FIELDS["pet_type_1"]: "Cat" if is_cat or not is_dog else "Dog",
        ANIMAL_COLLAR_DEFAULT_FIELDS["pet_type_2"]: "Dog" if is_cat and is_dog else "",
        ANIMAL_COLLAR_DEFAULT_FIELDS["animal_collar_type"]: "Basic Animal Collar",
        ANIMAL_COLLAR_DEFAULT_FIELDS["number_of_boxes"]: 1,
        ANIMAL_COLLAR_DEFAULT_FIELDS["directions"]: directions,
    }
    if color_map:
        defaults[ANIMAL_COLLAR_DEFAULT_FIELDS["color_map"]] = color_map
    if length and width:
        defaults.update({
            ANIMAL_COLLAR_DEFAULT_FIELDS["item_length"]: length,
            ANIMAL_COLLAR_DEFAULT_FIELDS["item_length_unit"]: "Inches",
            ANIMAL_COLLAR_DEFAULT_FIELDS["item_width"]: width,
            ANIMAL_COLLAR_DEFAULT_FIELDS["item_width_unit"]: "Inches",
            ANIMAL_COLLAR_DEFAULT_FIELDS["item_length_width_length"]: length,
            ANIMAL_COLLAR_DEFAULT_FIELDS["item_length_width_length_unit"]: "Inches",
            ANIMAL_COLLAR_DEFAULT_FIELDS["item_length_width_width"]: width,
            ANIMAL_COLLAR_DEFAULT_FIELDS["item_length_width_width_unit"]: "Inches",
        })

    defaults = {field_name: value for field_name, value in defaults.items() if value not in (None, "")}
    override_fields = {
        ANIMAL_COLLAR_DEFAULT_FIELDS["item_type_keyword"],
        ANIMAL_COLLAR_DEFAULT_FIELDS["closure"],
        ANIMAL_COLLAR_DEFAULT_FIELDS["pet_type_1"],
        ANIMAL_COLLAR_DEFAULT_FIELDS["pet_type_2"],
    }
    return defaults, override_fields


def _required_fields_from_data_definitions(wb):
    if "Data Definitions" not in wb.sheetnames:
        return {}
    ws = wb["Data Definitions"]
    required = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        values = [_text(value) for value in row]
        if len(values) < 6:
            continue
        field_name = values[1]
        label = values[2] or field_name
        if field_name and values[5].lower() == "required":
            required[field_name] = label
    return required


def _dimension_value_for_field(field_name, row):
    if ".unit" in field_name:
        return "Inches"
    package_length = row.get("package_length_in")
    package_width = row.get("package_width_in")
    package_height = row.get("package_height_in")
    lowered = field_name.lower()
    if ".length.value" in lowered or "length_width" in lowered and ".length." in lowered:
        return package_length
    if ".width.value" in lowered:
        return package_width
    if ".height.value" in lowered:
        return package_height
    if ".depth.value" in lowered:
        return package_height
    return ""


def _stable_field_default(field_name, row):
    field = field_name.lower()
    product_name = row.get("product_name") or row.get("title") or "Product"
    count = _first_positive_int(row.get("set_count"), 1)

    if "model_name" in field:
        return product_name
    if "model_number" in field or "part_number" in field:
        return row.get("sku")
    if "manufacturer" in field:
        return row.get("manufacturer") or row.get("brand") or "Generic"
    if "brand" in field:
        return row.get("brand") or "Generic"
    if "material" in field and row.get("material"):
        return row.get("material")
    if "color" in field and "standardized_values" in field:
        return _normalize_color_map(row.get("color"))
    if field.startswith("number_of_items") or field.startswith("item_package_quantity") or field.startswith("unit_count") and "#1.value" in field:
        return count
    if field.startswith("number_of_packs") or field.startswith("number_of_boxes"):
        return 1
    if "unit_count" in field and ".type" in field:
        return "Count"
    if "included_components" in field:
        return row.get("accessories") or f"{count} Count"
    if "specific_uses_for_product" in field:
        return "Outdoor" if str(row.get("category") or "").lower() in {"garden", "patio", "sports"} else "Everyday Use"
    if "recommended_uses_for_product" in field:
        return "Chewing" if str(row.get("product_type") or "").upper() == "PET_TOY" else "Everyday Use"
    if "breed_recommendation" in field:
        return "All Breed Sizes"
    if "pet_toy_type" in field:
        return "Chew Toy"
    if "pet_type" in field:
        return "Cat" if "cat" in _row_text(row).lower() or "猫" in _row_text(row) else "Dog"
    if "theme" in field:
        return "Animals" if str(row.get("product_type") or "").upper() == "PET_TOY" else product_name
    if "indoor_outdoor_usage" in field:
        return "Indoor"
    if "directions" in field:
        return "For supervised pet play only." if str(row.get("product_type") or "").upper() == "PET_TOY" else "Use as directed."
    if "care_instructions" in field:
        return "Spot Clean"
    if "pattern" in field:
        return "Solid"
    if "style" in field:
        return row.get("size") or row.get("color") or product_name
    if "country_of_origin" in field:
        return row.get("country_of_origin") or "China"
    if "batteries_required" in field or "batteries_included" in field:
        return "No"
    if "supplier_declared_dg_hz_regulation" in field:
        return "Not Applicable"
    if "contains_liquid_contents" in field:
        return "No"
    if "condition_type" in field:
        return "New"
    if "required_product_compliance_certificate" in field:
        return "Not Applicable"
    if "product_tax_code" in field:
        return "A_GEN_NOTAX"
    if "item_package_weight" in field or "item_weight" in field or "display_weight" in field:
        return "Pounds" if ".unit" in field else row.get("package_weight_lb")
    if any(token in field for token in [
        "item_package_dimensions",
        "item_dimensions",
        "item_length_width_height",
        "item_length_width",
        "item_width_height",
        "item_depth_width_height",
    ]):
        return _dimension_value_for_field(field_name, row)
    return ""


def _write_required_defaults(ws, row_index, row_data, field_to_col, required_fields):
    written = []
    is_parent = _is_parent_row(row_data)
    for field_name in required_fields:
        col = field_to_col.get(field_name)
        if not col or ws.cell(row_index, col).value not in (None, ""):
            continue
        if is_parent and any(token in field_name for token in [
            "list_price",
            "purchasable_offer",
            "item_package_dimensions",
            "item_package_weight",
            "main_product_image",
        ]):
            continue
        value = _stable_field_default(field_name, row_data)
        if value in (None, ""):
            continue
        ws.cell(row_index, col).value = value
        written.append(field_name)
    return written


def fill_template(project_dir, draft_path=None, template_path=None, output_path=None, write_report=False):
    project_dir = Path(project_dir)
    draft_path = Path(draft_path) if draft_path else find_latest_draft(project_dir)
    template_path = Path(template_path) if template_path else find_template(project_dir)

    rows = read_intake_rows(draft_path)
    if not rows:
        raise ValueError(f"草稿没有可写入的数据行：{draft_path}")
    rows = prepare_variation_rows(rows, project_dir)

    if output_path is None:
        product_name = rows[0].get("product_name") or project_dir.name
        output_path = versioned_template_path(project_dir, product_name)
    else:
        output_path = Path(output_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    copyfile(template_path, output_path)

    wb = load_workbook(output_path, keep_vba=output_path.suffix.lower() == ".xlsm")
    ws = find_template_sheet(wb)
    if ws is None:
        raise ValueError(f"模板里没有 {template_sheet_names_text()} 页：{template_path}")

    field_to_col = {}
    for col in range(1, ws.max_column + 1):
        field_name = ws.cell(5, col).value
        if field_name:
            field_to_col[str(field_name).strip()] = col

    required_fields = _required_fields_from_data_definitions(wb)
    clear_template_data(ws)

    start_row = TEMPLATE_DATA_START_ROW
    written_fields = []
    rule_written_fields = []
    category_rule_written_fields = []
    required_default_fields = []
    for row_index, row_data in enumerate(rows, start_row):
        for field_name, source in FIELD_MAP.items():
            col = field_to_col.get(field_name)
            if not col:
                continue
            value = _value_for(row_data, source)
            if value in (None, ""):
                continue
            ws.cell(row_index, col).value = value
            written_fields.append(field_name)

        rule_defaults = load_safe_defaults(
            row_data.get("product_type"),
            existing_fields=FIELD_MAP.keys(),
        )
        for field_name, value in rule_defaults.items():
            col = field_to_col.get(field_name)
            if not col:
                continue
            if ws.cell(row_index, col).value not in (None, ""):
                continue
            ws.cell(row_index, col).value = value
            rule_written_fields.append(field_name)

        category_defaults, category_override_fields = _animal_collar_rule_defaults(row_data)
        for field_name, value in category_defaults.items():
            col = field_to_col.get(field_name)
            if not col:
                continue
            current = ws.cell(row_index, col).value
            if current not in (None, "") and field_name not in category_override_fields:
                continue
            if current == value:
                continue
            ws.cell(row_index, col).value = value
            category_rule_written_fields.append(field_name)

        default_targets = dict(required_fields)
        product_type = str(row_data.get("product_type") or "")
        for label, field_name in PRODUCT_TYPE_CONDITIONAL_FIELDS.get(product_type, {}).items():
            default_targets.setdefault(field_name, label)
        required_default_fields.extend(_write_required_defaults(ws, row_index, row_data, field_to_col, default_targets))

    wb.save(output_path)
    if write_report:
        report_path = output_path.with_name(f"{output_path.stem}_写入报告.md")
        write_fill_report(report_path, output_path, draft_path, template_path, rows, written_fields, rule_written_fields, category_rule_written_fields, required_default_fields)
    return output_path, draft_path, template_path, sorted(set(written_fields + rule_written_fields + category_rule_written_fields + required_default_fields))


def clear_template_data(ws):
    for row in range(TEMPLATE_EXAMPLE_ROW + 1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None


def write_fill_report(path, output_path, draft_path, template_path, rows, written_fields, rule_written_fields=None, category_rule_written_fields=None, required_default_fields=None):
    rule_written_fields = rule_written_fields or []
    category_rule_written_fields = category_rule_written_fields or []
    required_default_fields = required_default_fields or []
    first = rows[0]
    lines = [
        f"# {Path(output_path).name} 写入报告",
        "",
        f"- 输出文件：{output_path}",
        f"- 草稿来源：{draft_path}",
        f"- 模板来源：{template_path}",
        f"- 写入 SKU 数：{len(rows)}",
        f"- 写入字段数：{len(set(written_fields))}",
        f"- 成功规则补字段数：{len(set(rule_written_fields))}",
        f"- 类目规则补字段数：{len(set(category_rule_written_fields))}",
        f"- 必填字段兜底补字段数：{len(set(required_default_fields))}",
        "",
        "## 关键字段",
        "",
        f"- SKU：{first.get('sku')}",
        f"- Product Type：{first.get('product_type')}",
        f"- Brand：{first.get('brand')}",
        f"- Manufacturer：{first.get('manufacturer')}",
        f"- Item Type Keyword：{first.get('item_type_keyword')}",
        f"- List Price：{first.get('list_price') or '留空，等待人工填写'}",
        f"- Haul Price：{first.get('haul_price') or '留空，后续与 List Price 保持一致'}",
        f"- Main Image URL：留空",
        "",
        "## 已写入字段",
        ""
    ]
    for field_name in sorted(set(written_fields)):
        lines.append(f"- `{field_name}`")
    if rule_written_fields:
        lines.extend(["", "## 成功规则补写字段", ""])
        for field_name in sorted(set(rule_written_fields)):
            lines.append(f"- `{field_name}`")
    if category_rule_written_fields:
        lines.extend(["", "## 类目规则补写字段", ""])
        for field_name in sorted(set(category_rule_written_fields)):
            lines.append(f"- `{field_name}`")
    if required_default_fields:
        lines.extend(["", "## 必填字段兜底补写字段", ""])
        for field_name in sorted(set(required_default_fields)):
            lines.append(f"- `{field_name}`")
    Path(path).write_text("\n".join(lines), encoding="utf-8")
