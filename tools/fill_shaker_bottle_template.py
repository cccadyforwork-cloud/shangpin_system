from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName


BASE_DIR = Path("/Users/cc/Desktop/摇摇杯资料")
TEMPLATE = BASE_DIR / "BOTTLE.xlsm"
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT = OUTPUT_DIR / "健身摇摇杯_v2.xlsm"
DESKTOP_OUTPUT = Path("/Users/cc/Desktop/健身摇摇杯_v2.xlsm")


ROWS = [
    {
        "sku": "CA-Shakerbottle-Black",
        "color": "Black",
        "title_color": "Black",
        "source_color": "黑色",
        "image": "https://cbu01.alicdn.com/img/ibank/O1CN01PRN4sR29lyjaOrsI7_!!2219354758109-0-cib.jpg",
    },
    {
        "sku": "CA-Shakerbottle-Pink",
        "color": "Light Pink",
        "title_color": "Pink",
        "source_color": "粉色",
        "image": "https://cbu01.alicdn.com/img/ibank/O1CN01acwAVj29lyjeAm5ta_!!2219354758109-0-cib.jpg",
    },
    {
        "sku": "CA-Shakerbottle-Green",
        "color": "Green Machine",
        "title_color": "Green",
        "source_color": "绿色",
        "image": "https://cbu01.alicdn.com/img/ibank/O1CN01QiwRKy29lyjdjlh0A_!!2219354758109-0-cib.jpg",
    },
]


COMMON_BULLETS = [
    "350 ml shaker bottle is sized for protein powder, meal replacement shakes, smoothies, soy milk, and daily hydration at the gym or office.",
    "Leak-proof flip top design helps reduce spills in a gym bag, backpack, or cup holder when the lid is closed properly.",
    "Built-in mixing structure helps blend powder and liquid more evenly for workout drinks, milkshakes, and supplement mixes.",
    "Lightweight plastic body is easy to carry for fitness training, running, travel, school, work, and meal prep routines.",
    "Wide-mouth cup opening makes it easier to add powder, pour liquid, drink, and clean after each use.",
]


def build_col_map(ws):
    cols = {}
    seen = {}
    for col in range(1, ws.max_column + 1):
        label = ws.cell(4, col).value
        if not label:
            continue
        seen[label] = seen.get(label, 0) + 1
        key = label if seen[label] == 1 else f"{label}.{seen[label] - 1}"
        cols[key] = col
    return cols


def setv(ws, row, cols, label, value):
    col = cols.get(label)
    if col and value not in (None, ""):
        ws.cell(row, col).value = value


def clearv(ws, row, cols, label):
    col = cols.get(label)
    if col:
        ws.cell(row, col).value = None


def copy_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


def allow_generic_brand(wb):
    dropdown = wb["Dropdown Lists"]
    dropdown["G5"] = "Generic"
    name = "BOTTLEbrandmarketplace_idATVPDKIKX0DERlanguage_tagen_US1.value"
    wb.defined_names[name] = DefinedName(
        name=name,
        attr_text="'Dropdown Lists'!$G$4:$G$5",
    )


def fill_row(ws, row, cols, item):
    sku = item["sku"]
    color = item["color"]
    title_color = item["title_color"]
    title = f"Protein Shaker Bottle 350 ml, Leak Proof Plastic Fitness Mixer Cup for Gym Workout Smoothies and Meal Prep, {title_color}"
    description = (
        f"This 350 ml protein shaker bottle is a compact fitness mixer cup for protein powder, smoothies, milkshakes, soy milk, supplements, and daily water. "
        f"The lightweight plastic body is easy to carry to the gym, office, school, or travel bag, while the leak-proof flip top and mixing structure help support everyday workout drink preparation. "
        f"Color: {title_color}."
    )

    setv(ws, row, cols, "SKU", sku)
    setv(ws, row, cols, "Product Type", "BOTTLE")
    setv(ws, row, cols, "Listing Action", "Create or Replace (Full Update)")
    clearv(ws, row, cols, "Parentage Level")
    clearv(ws, row, cols, "Parent SKU")
    clearv(ws, row, cols, "Variation Theme Name")

    setv(ws, row, cols, "Item Name", title)
    setv(ws, row, cols, "Item Highlight", f"{title_color}, 350 ml")
    setv(ws, row, cols, "Brand Name", "Generic")
    setv(ws, row, cols, "Item Type Keyword", "家居、厨具、家装 > 厨房和餐厅 > 旅行和外出用杯子 > 蛋白粉摇摇杯健身运动水瓶 (sports-nutrition-shaker-bottles)")
    setv(ws, row, cols, "Package Level", "Unit")
    setv(ws, row, cols, "Model Number", sku)
    setv(ws, row, cols, "Model Name", "Protein Shaker Bottle")
    setv(ws, row, cols, "Manufacturer", "Generic")
    setv(ws, row, cols, "Main Image URL", item["image"])
    setv(ws, row, cols, "Swatch Image URL", item["image"])

    setv(ws, row, cols, "Product Description", description)
    for idx, bullet in enumerate(COMMON_BULLETS, start=1):
        label = "Bullet Point" if idx == 1 else f"Bullet Point.{idx - 1}"
        setv(ws, row, cols, label, bullet)
    setv(ws, row, cols, "Generic Keyword", "protein shaker bottle; fitness shaker cup; workout mixer cup; gym bottle; smoothie shaker; meal prep cup")
    for idx, feature in enumerate(["Leak Proof", "Lightweight", "Flip Top", "Volume Marking", "Wide Mouth"], start=1):
        label = "Special Features" if idx == 1 else f"Special Features.{idx - 1}"
        setv(ws, row, cols, label, feature)
    setv(ws, row, cols, "Style", "Modern")
    setv(ws, row, cols, "Material", "Plastic")
    setv(ws, row, cols, "Number of Items", 1)
    setv(ws, row, cols, "Item Package Quantity", 1)
    setv(ws, row, cols, "Color", color)
    setv(ws, row, cols, "Size", "350 ml")
    setv(ws, row, cols, "Part Number", sku)
    setv(ws, row, cols, "Item Shape", "Round")
    setv(ws, row, cols, "Theme", "Sport")
    setv(ws, row, cols, "Care Instructions", "Hand Wash Only")
    setv(ws, row, cols, "Material Type Free", "Food grade plastic")
    setv(ws, row, cols, "Capacity", 350)
    setv(ws, row, cols, "Capacity Unit", "Milliliters")
    setv(ws, row, cols, "Volume Capacity", 350)
    setv(ws, row, cols, "Volume Capacity Unit", "Milliliters")
    setv(ws, row, cols, "Pattern", "Solid")
    setv(ws, row, cols, "Unit Count", 1)
    setv(ws, row, cols, "Unit Count Type", "Count")
    setv(ws, row, cols, "Included Components", "1 x shaker bottle")
    setv(ws, row, cols, "Specific Uses For Product", "Water")
    setv(ws, row, cols, "Recommended Uses For Product", "Gym")
    setv(ws, row, cols, "Recommended Uses For Product.1", "Running")
    setv(ws, row, cols, "Recommended Uses For Product.2", "Travel")
    setv(ws, row, cols, "Bottle Color", color)
    setv(ws, row, cols, "Reusability", "Reusable")
    setv(ws, row, cols, "Number of Packs", 1)
    setv(ws, row, cols, "Item Weight", 0.249)
    setv(ws, row, cols, "Item Weight Unit", "Pounds")

    # 1688 product size is 8.5 x 7.5 x 15.5 cm.
    setv(ws, row, cols, "Item Height Base to Top", 6.10)
    setv(ws, row, cols, "Item Height Unit", "Inches")
    setv(ws, row, cols, "Item Width Top", 2.95)
    setv(ws, row, cols, "Item Width Unit", "Inches")
    setv(ws, row, cols, "Item Length", 3.35)
    setv(ws, row, cols, "Item Length Unit", "Inches")
    setv(ws, row, cols, "Item Width", 2.95)
    setv(ws, row, cols, "Item Width Unit.1", "Inches")
    setv(ws, row, cols, "Item Height", 6.10)
    setv(ws, row, cols, "Item Height Unit.1", "Inches")

    setv(ws, row, cols, "Item Condition", "New")
    setv(ws, row, cols, "List Price", 3.59)
    setv(ws, row, cols, "Product Tax Code", "A_GEN_NOTAX")
    setv(ws, row, cols, "Main Image Location", item["image"])
    setv(ws, row, cols, "Your Price USD (Low-Cost Store, US)", 3.59)
    setv(ws, row, cols, "Item Package Length", 3.15)
    setv(ws, row, cols, "Package Length Unit", "Inches")
    setv(ws, row, cols, "Item Package Width", 3.94)
    setv(ws, row, cols, "Package Width Unit", "Inches")
    setv(ws, row, cols, "Item Package Height", 6.3)
    setv(ws, row, cols, "Package Height Unit", "Inches")
    setv(ws, row, cols, "Package Weight", 0.249)
    setv(ws, row, cols, "Package Weight Unit", "Pounds")
    setv(ws, row, cols, "Item Display Weight", 0.249)
    setv(ws, row, cols, "Item Display Weight Unit", "Pounds")
    setv(ws, row, cols, "Country of Origin", "China")
    setv(ws, row, cols, "Are batteries required?", "No")
    setv(ws, row, cols, "Are batteries included?", "No")
    setv(ws, row, cols, "Dangerous Goods Regulations", "Not Applicable")


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(TEMPLATE, keep_vba=True)
    allow_generic_brand(wb)
    ws = wb["Template"]
    cols = build_col_map(ws)

    for row in range(7, 10):
        copy_row_style(ws, 6, row)

    for offset, item in enumerate(ROWS):
        fill_row(ws, 7 + offset, cols, item)

    wb.save(OUTPUT)
    wb.save(DESKTOP_OUTPUT)
    print(OUTPUT)
    print(DESKTOP_OUTPUT)


if __name__ == "__main__":
    main()
