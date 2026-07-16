import re
import unicodedata
import xml.etree.ElementTree as ET
from html import unescape
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from .paths import (
    COMPETITOR_DIR,
    LEGACY_COMPETITOR_DIR,
    LEGACY_REMARKS_DIR,
    OUTPUTS_DIR,
    PACKAGING_PRICING_DIR,
    PRODUCT_DETAIL_DIR,
    PURCHASE_DIR,
    TEMPLATE_DIRS,
    safe_name,
)
from .template_writer import extract_template_product_type
from .workbook_io import write_intake_workbook


SUPPORTED_TEXT_EXT = {".txt", ".md", ".csv", ".tsv", ".html", ".htm"}
SUPPORTED_EXCEL_EXT = {".xlsx", ".xlsm", ".xls"}
SUPPORTED_PDF_EXT = {".pdf"}
URL_RE = re.compile(r"https?://[^\s,，;；)）]+", re.I)
PRICE_RE = re.compile(r"(?:¥|￥|\$|usd|rmb|价格|售价|成本|单价)[:：\s]*([0-9]+(?:\.[0-9]+)?)", re.I)
UNIT_PRICE_RE = re.compile(r"([0-9]+(?:\.[0-9]+)?)\s*元\s*/\s*个", re.I)
WEIGHT_RE = re.compile(r"([0-9]+(?:\.[0-9]+)?)\s*(kg|kgs|g|gram|grams|lb|lbs|pound|pounds)\b", re.I)
SIZE_RE = re.compile(r"([0-9]+(?:\.[0-9]+)?)\s*[xX*×]\s*([0-9]+(?:\.[0-9]+)?)\s*(?:[xX*×]\s*([0-9]+(?:\.[0-9]+)?))?\s*(cm|mm|in|inch|inches)?", re.I)
SKU_RE = re.compile(r"\b[A-Z0-9][A-Z0-9_-]{4,}\b")
ORDER_ITEM_CODE_RE = re.compile(r"(?m)^\s*\d+\s+([A-Z0-9-]{3,})\s*$")
CHINESE_COLOR_RE = re.compile(
    r"(?:颜色|顏色)\s*[:：]\s*(?:[0-9]+(?:\.[0-9]+)?\s*g\s*)?(?:瑜\s*伽\s*砖\s*)?-\s*([^\s,，;；]+)"
)


COLOR_WORDS = {
    "black": "Black",
    "white": "White",
    "red": "Red",
    "blue": "Blue",
    "green": "Green",
    "pink": "Pink",
    "gray": "Gray",
    "grey": "Gray",
    "yellow": "Yellow",
    "purple": "Purple",
    "orange": "Orange",
    "transparent": "Transparent",
    "黑": "Black",
    "黑色": "Black",
    "白": "White",
    "白色": "White",
    "红": "Red",
    "红色": "Red",
    "蓝": "Blue",
    "蓝色": "Blue",
    "绿": "Green",
    "绿色": "Green",
    "粉": "Pink",
    "粉色": "Pink",
    "灰": "Gray",
    "灰色": "Gray",
    "黄": "Yellow",
    "黄色": "Yellow",
    "紫": "Purple",
    "紫色": "Purple",
    "橙": "Orange",
    "橙色": "Orange",
    "透明": "Transparent",
    "透明色": "Transparent"
}


MATERIAL_WORDS = {
    "metal": "Metal",
    "silicone": "Silicone",
    "plastic": "Plastic",
    "stainless steel": "Stainless Steel",
    "wood": "Wood",
    "cotton": "Cotton",
    "polyester": "Polyester",
    "nylon": "Nylon",
    "abs": "ABS",
    "硅胶": "Silicone",
    "塑料": "Plastic",
    "不锈钢": "Stainless Steel",
    "木": "Wood",
    "棉": "Cotton",
    "涤纶": "Polyester",
    "尼龙": "Nylon"
}


RISK_WORDS = [
    "food",
    "食品",
    "skin",
    "皮肤",
    "medical",
    "medicine",
    "医疗",
    "治疗",
    "cure",
    "treat",
    "pain relief",
    "brand",
    "logo",
    "trademark",
    "品牌",
    "商标"
]


CJK_RE = re.compile(r"[\u4e00-\u9fff]")


ENGLISH_PRODUCT_NAMES = [
    ("宠物airtag", "AirTag Pet Collar"),
    ("AirTag硅胶护套宠物项圈", "AirTag Pet Collar"),
    ("Airtag硅胶护套宠物项圈", "AirTag Pet Collar"),
    ("宠物项圈", "Pet Collar"),
    ("项圈", "Pet Collar"),
    ("猫玩具三支大棉签", "Cat Q-Tip Toy Set"),
    ("猫玩具", "Cat Toy"),
    ("大棉签", "Cat Q-Tip Toy"),
    ("逗猫棒", "Interactive Cat Toy"),
    ("猫薄荷", "Catnip Cat Toy"),
    ("防磨贴", "Heel Blister Pads"),
    ("水泡贴", "Blister Cushions"),
    ("后跟贴", "Heel Cushion Pads"),
    ("瑜伽砖", "Yoga Block"),
    ("花束卡片夹", "Floral Card Holder Picks"),
]


SOURCE_FOLDERS = {
    PURCHASE_DIR,
    PACKAGING_PRICING_DIR,
    PRODUCT_DETAIL_DIR,
    COMPETITOR_DIR,
    LEGACY_COMPETITOR_DIR,
    LEGACY_REMARKS_DIR,
}

TEMPLATE_FOLDERS = set(TEMPLATE_DIRS)


def _read_text_file(path):
    return Path(path).read_text(encoding="utf-8-sig", errors="ignore")


def _read_excel_xml_file(path):
    text = _read_text_file(path)
    root = ET.fromstring(text)
    ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
    ss_name = "{urn:schemas-microsoft-com:office:spreadsheet}Name"
    ss_index = "{urn:schemas-microsoft-com:office:spreadsheet}Index"
    chunks = []
    for ws in root.findall(".//ss:Worksheet", ns):
        sheet_name = ws.attrib.get(ss_name, "")
        chunks.append(f"Sheet: {sheet_name}")
        for row in ws.findall(".//ss:Row", ns):
            values = []
            col_index = 1
            for cell in row.findall("ss:Cell", ns):
                explicit_index = cell.attrib.get(ss_index)
                if explicit_index:
                    while col_index < int(explicit_index):
                        values.append("")
                        col_index += 1
                data = cell.find("ss:Data", ns)
                values.append((data.text or "").strip() if data is not None else "")
                col_index += 1
            if any(values):
                chunks.append(" | ".join(values))
    return "\n".join(chunks)


def _read_excel_file(path):
    if path.suffix.lower() == ".xls":
        return _read_excel_xml_file(path)

    chunks = []
    wb = load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        chunks.append(f"Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value not in (None, "")]
            if values:
                chunks.append(" | ".join(values))
    return "\n".join(chunks)


def _read_pdf_file(path):
    reader = PdfReader(str(path))
    chunks = []
    for page in reader.pages:
        chunks.append(page.extract_text() or "")
    return unicodedata.normalize("NFKC", "\n".join(chunks))


def _read_html_file(path):
    text = _read_text_file(path)
    title_match = re.search(r"(?is)<title[^>]*>(.*?)</title>", text)
    text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", text)
    text = re.sub(r"(?is)<br\s*/?>", "\n", text)
    text = re.sub(r"(?is)</(div|p|li|tr|h[1-6])>", "\n", text)
    text = re.sub(r"(?is)<[^>]+>", " ", text)
    text = unescape(text)
    if title_match:
        title = unescape(re.sub(r"\s+", " ", title_match.group(1))).strip()
        text = f"HTML_TITLE: {title}\n{text}"
    return re.sub(r"[ \t]+", " ", text)


def _collect_sources(project_dir):
    project_dir = Path(project_dir)
    sources = []
    for path in project_dir.rglob("*"):
        if not path.is_file():
            continue
        if path.name.startswith("~$") or path.name == ".DS_Store":
            continue
        if path.parent.name not in SOURCE_FOLDERS:
            continue
        if path.suffix.lower() in SUPPORTED_EXCEL_EXT and (
            "产品资料" in path.name or "自动提炼草稿" in path.name
        ):
            continue
        if "自动提炼草稿" in path.name or "资料提炼报告" in path.name:
            continue
        suffix = path.suffix.lower()
        try:
            if suffix in {".html", ".htm"}:
                text = _read_html_file(path)
            elif suffix in SUPPORTED_TEXT_EXT:
                text = _read_text_file(path)
            elif suffix in SUPPORTED_EXCEL_EXT:
                text = _read_excel_file(path)
            elif suffix in SUPPORTED_PDF_EXT:
                text = _read_pdf_file(path)
            else:
                continue
        except Exception as exc:
            sources.append((path, f"[读取失败: {exc}]"))
            continue
        sources.append((path, text))
    return sources


def _first_match(pattern, text, group=1):
    match = pattern.search(text)
    if not match:
        return ""
    if match.lastindex:
        return match.group(group).strip()
    return match.group(0).strip()


def _extract_named_value(labels, text):
    for label in labels:
        pattern = re.compile(rf"{re.escape(label)}\s*[:：]\s*([^\n\r|,，;；]+)", re.I)
        match = pattern.search(text)
        if match:
            return match.group(1).strip()
    return ""


def _normalize_table_header(value):
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    return re.sub(r"[\s_（）()]+", "", text)


def _cell_by_headers(row, headers, candidates):
    normalized_candidates = [_normalize_table_header(item) for item in candidates]
    for idx, header in enumerate(headers):
        if any(candidate and candidate in header for candidate in normalized_candidates):
            return row[idx].strip() if idx < len(row) else ""
    return ""


def _number_text(value):
    match = re.search(r"-?\d+(?:\.\d+)?", str(value or ""))
    return match.group(0) if match else ""


def _extract_count_from_style_title(value):
    text = str(value or "")
    chinese_numbers = {
        "一": 1,
        "二": 2,
        "两": 2,
        "三": 3,
        "四": 4,
        "五": 5,
        "六": 6,
        "七": 7,
        "八": 8,
        "九": 9,
        "十": 10,
    }
    chinese_match = re.search(r"([一二两三四五六七八九十])\s*(?:支|只|个|件|条|片)", text)
    if chinese_match:
        return str(chinese_numbers.get(chinese_match.group(1), ""))
    match = re.search(r"(\d+)\s*款\s*各\s*(\d+)\s*(?:片|pcs?|pieces?)?", text, re.I)
    if match:
        return str(int(match.group(1)) * int(match.group(2)))
    match = re.search(r"(\d+)\s*(?:片|pcs?|pieces?|pack|count)\b", text, re.I)
    return match.group(1) if match else ""


def _extract_price_table_data(sources):
    for source_path, text in sources:
        if source_path.parent.name != PACKAGING_PRICING_DIR:
            continue
        rows = []
        for line in text.splitlines():
            if "|" not in line:
                continue
            values = [item.strip() for item in line.split("|")]
            if len(values) >= 2:
                rows.append(values)
        for index, header_row in enumerate(rows):
            headers = [_normalize_table_header(item) for item in header_row]
            if not _cell_by_headers(header_row, headers, ["SKU"]) or not any(
                keyword in "".join(headers)
                for keyword in ["定价", "售价", "长in", "宽in", "重lb"]
            ):
                continue
            data_row = next((row for row in rows[index + 1:] if any(cell.strip() for cell in row)), [])
            if not data_row:
                continue
            style_title = _cell_by_headers(data_row, headers, ["款式标题", "标题", "产品名称", "品名"])
            price = _number_text(_cell_by_headers(data_row, headers, ["定价", "售价", "listprice", "price"]))
            return {
                "sku": _cell_by_headers(data_row, headers, ["SKU", "货号"]),
                "style_title": style_title,
                "set_count": _extract_count_from_style_title(style_title),
                "cost": _number_text(_cell_by_headers(data_row, headers, ["采购价", "成本", "单价", "cost"])),
                "package_length_in": _number_text(_cell_by_headers(data_row, headers, ["长(in)", "长in", "length"])),
                "package_width_in": _number_text(_cell_by_headers(data_row, headers, ["宽(in)", "宽in", "width"])),
                "package_height_in": _number_text(_cell_by_headers(data_row, headers, ["高(in)", "高in", "height"])),
                "package_weight_lb": _number_text(_cell_by_headers(data_row, headers, ["重(lb)", "重量(lb)", "重lb", "weight"])),
                "list_price": price,
                "haul_price": price,
            }
    return {}


def _convert_size_to_inches(match):
    values = [float(match.group(i)) for i in [1, 2, 3] if match.group(i)]
    unit = (match.group(4) or "").lower()
    if unit in {"cm", ""}:
        values = [round(value / 2.54, 2) for value in values]
    elif unit == "mm":
        values = [round(value / 25.4, 2) for value in values]
    else:
        values = [round(value, 2) for value in values]
    while len(values) < 3:
        values.append("")
    return values[:3]


def _convert_weight_to_lb(match):
    value = float(match.group(1))
    unit = match.group(2).lower()
    if unit in {"kg", "kgs"}:
        return round(value * 2.20462, 2)
    if unit in {"g", "gram", "grams"}:
        return round(value / 453.59237, 2)
    return round(value, 2)


def _guess_from_words(words, text):
    lowered = text.lower()
    for key, value in words.items():
        if key.lower() in lowered:
            return value
    return ""


def _normalize_color_value(value):
    text = str(value or "").strip()
    if not text:
        return ""
    if text in COLOR_WORDS:
        return COLOR_WORDS[text]
    compact = text.replace("色", "")
    if compact in COLOR_WORDS:
        return COLOR_WORDS[compact]
    for key, color in COLOR_WORDS.items():
        if key and key in text:
            return color
    return text


def _has_cjk(value):
    return bool(CJK_RE.search(str(value or "")))


def _english_product_name(value, context=""):
    text = str(value or "")
    combined = f"{text}\n{context}"
    lowered = combined.lower()
    if "airtag" in lowered and any(keyword in combined for keyword in ["项圈", "宠物", "猫咪", "狗狗", "collar"]):
        return "AirTag Pet Collar"
    for keyword, english_name in ENGLISH_PRODUCT_NAMES:
        if keyword in combined:
            return english_name
    return text if text and not _has_cjk(text) else ""


def _english_style_phrase(value):
    text = str(value or "")
    match = re.search(r"(\d+)\s*款\s*各\s*(\d+)\s*(片|个|只|双|pcs?|pieces?)?", text, re.I)
    if match:
        total = int(match.group(1)) * int(match.group(2))
        return f"{match.group(1)} Styles, {total} Count"
    match = re.search(r"(\d+)\s*(片|个|只|双|pcs?|pieces?|pack|count)", text, re.I)
    if match:
        return f"{match.group(1)} Count"
    return ""


def _guess_product_name(project_dir, text):
    project_name = Path(project_dir).name.split("_", 1)[-1]
    if "瑜伽砖" in project_name:
        return "Yoga Block"
    if "花束卡片夹" in project_name:
        return "Floral Card Holder Picks"
    if project_name and not project_name.startswith("20"):
        return project_name

    named = _extract_named_value(["产品名", "品名", "商品名称", "product name", "item name", "title"], text)
    if named and "amazon haul" not in named.lower():
        return named[:120]

    lines = [line.strip(" -\t") for line in text.splitlines() if len(line.strip()) >= 4]
    for line in lines[:20]:
        if not URL_RE.search(line) and not PRICE_RE.search(line):
            return line[:120]
    return Path(project_dir).name.split("_", 1)[-1]


def _guess_category(text):
    lowered = text.lower()
    if any(word in lowered for word in ["cat toy", "q tip for cats", "q-tip", "catnip", "猫玩具", "逗猫", "猫薄荷", "宠物"]):
        return "pet"
    if any(word in lowered for word in ["toy", "玩具"]):
        return "toy"
    if any(word in lowered for word in ["floral", "bouquet", "flower", "card holder", "花束", "鲜花", "卡片夹", "卡插"]):
        return "home"
    if any(word in lowered for word in ["yoga", "pilates", "fitness", "exercise", "瑜伽", "健身", "运动"]):
        return "sports"
    if any(word in lowered for word in ["kitchen", "厨房", "cookware", "utensil"]):
        return "kitchen"
    if any(word in lowered for word in ["pet", "宠物"]):
        return "pet"
    if any(word in lowered for word in ["beauty", "makeup", "cosmetic", "美妆"]):
        return "beauty"
    if any(word in lowered for word in ["home", "storage", "收纳", "家居"]):
        return "home"
    return ""


def _guess_route(text, price, brand_hint=""):
    if brand_hint and brand_hint.lower() != "generic":
        return "Brand"
    if price and price <= 20:
        return "Haul Generic"
    return "Haul Generic"


def _normalize_route_strategy(route_mode):
    value = str(route_mode or "").strip().lower().replace("-", "_")
    aliases = {
        "single": "single_child",
        "single_child": "single_child",
        "single_link": "single_child",
        "variation": "variation",
        "variant": "variation",
        "parent_child": "variation",
        "set": "set_bundle",
        "bundle": "set_bundle",
        "set_bundle": "set_bundle",
    }
    return aliases.get(value, "")


def _route_strategy(text, route, row_count, set_count, route_mode=None):
    selected_strategy = _normalize_route_strategy(route_mode)
    if selected_strategy:
        return selected_strategy

    route_text = f"{route}\n{text}".lower()
    if (
        any(keyword in route_text for keyword in ["父子", "变体", "多链接"])
        or re.search(r"\b(variation|variant|parent|child)\b", route_text)
        or "parent/child" in route_text
        or "parent-child" in route_text
    ):
        return "variation"
    if any(keyword in route_text for keyword in ["set bundle", "bundle", "套装售卖", "组合售卖"]):
        return "set_bundle"
    return "variation"


def _route_name_for_strategy(route, strategy):
    if route == "Brand":
        return route
    if strategy == "variation":
        return "Haul Generic Variation"
    if strategy == "set_bundle":
        return "Haul Generic Set Bundle"
    return route


def _make_sku(product_name):
    words = re.findall(r"[A-Za-z0-9]+", product_name.upper())
    base = "-".join(words[:4]) if words else safe_name(product_name).upper()
    return f"{base[:32]}-001"


def _sku_base(product_name):
    words = re.findall(r"[A-Za-z0-9]+", product_name.upper())
    if words:
        return "-".join(words[:4])[:24]
    if "瑜伽" in product_name:
        return "YOGA-BLOCK"
    return "PRODUCT"


def _single_link_sku(product_name):
    return f"{_sku_base(product_name)}-001"


def _parent_sku(product_name):
    return f"{_sku_base(product_name)}-PARENT"


def _sku_for_color(product_name, color, index):
    if "yoga" in product_name.lower() or "瑜伽" in product_name:
        return f"YOGA-Block-{color}"
    base = _sku_base(product_name)
    color_part = re.sub(r"[^A-Z0-9]+", "", str(color).upper()) or f"{index:02d}"
    return f"{base}-{color_part}-{index:02d}"


def _make_copy(product_name, color, size, material, set_count):
    subject = _english_product_name(product_name)
    if not subject:
        return "", ["", "", "", "", ""], ""
    set_text = f", {set_count} Pack" if set_count else ""
    tail = ", ".join(value for value in [color, size] if value)
    title = f"{subject}{set_text}" + (f", {tail}" if tail else "")
    combined = f"{product_name}\n{subject}".lower()
    is_yoga = "yoga" in subject.lower() or "瑜伽" in subject
    is_floral_card_holder = "floral card holder" in subject.lower() or "花束卡片夹" in subject
    is_cat_qtip = any(keyword in combined for keyword in ["cat q-tip", "cat toy", "猫玩具", "大棉签", "逗猫", "猫薄荷"])
    is_airtag_pet_collar = "airtag" in combined and ("collar" in combined or "项圈" in combined or "宠物" in combined)
    if is_airtag_pet_collar:
        title = "AirTag Cat Collar with Silicone Holder, Breakaway Pet Collar for Cats and Small Dogs"
        if color:
            title += f", {color}"
        title += ", AirTag Not Included"
        bullet_1 = "Built-in silicone holder helps keep an Apple AirTag attached to the collar without a separate case."
        bullet_2 = f"Adjustable collar size {size} is suitable for cats and small pets." if size else "Adjustable collar is suitable for cats and small pets."
        bullet_3 = f"Made with {material or 'polyester and silicone'} for lightweight daily pet wear."
        bullet_4 = "Breakaway-style collar design helps cats release from pressure during normal use and play."
        bullet_5 = f"{color} collar is made for cats and small dogs; AirTag device is not included." if color else "Made for cats and small dogs; AirTag device is not included."
        desc = "This AirTag pet collar combines a lightweight collar strap with a silicone holder designed for Apple AirTag. It helps pet owners keep a tracker attached during daily wear while keeping the AirTag device separate from the listing."
    elif is_cat_qtip:
        pack_text = f"{set_count} Pack" if set_count else "3 Pack"
        title = f"Cat Q-Tip Toy Set, {pack_text} Giant Cotton Swab Cat Toys with Catnip for Indoor Interactive Play and Chewing"
        bullet_1 = "Q-tip shaped cat toys encourage batting, chasing, tossing, and independent indoor play."
        bullet_2 = f"Made with {material or 'soft cotton and felt'} for a lightweight texture cats can paw and carry."
        bullet_3 = "Catnip scent helps attract attention and adds variety to daily play sessions."
        bullet_4 = f"Includes {pack_text.lower()} for rotation, replacement, or use in multi-cat homes."
        bullet_5 = "Long swab shape is easy for cats to grab, kick, chew, and roll across floors."
        desc = f"This {pack_text.lower()} cat Q-tip toy set gives indoor cats a simple way to bat, chase, chew, and kick during daily play. The giant cotton swab shape is lightweight and easy to carry, while the catnip scent helps draw interest for solo play or owner-led interaction."
    elif is_floral_card_holder:
        title = f"Floral Card Holder Picks, {color}, Metal Flower Bouquet Card Sticks for Gift Cards and Table Centerpieces"
        bullet_1 = "Designed for flower bouquets, gift cards, table centerpieces, party signs, and photo notes."
        bullet_2 = f"Made with {material or 'metal'} for a clean decorative look in floral arrangements."
        bullet_3 = "Slim pick design slides into bouquets, foam, vases, baskets, and display arrangements."
        bullet_4 = "Useful for weddings, birthdays, flower shops, gift wrapping, and event table decor."
        bullet_5 = "Lightweight pieces are easy to place, remove, store, and reuse for seasonal displays."
        desc = "These floral card holder picks help display cards, photos, notes, and small signs in bouquets, centerpieces, baskets, and party arrangements. The slim metal design is suitable for flower shops, weddings, birthdays, gift displays, and home decoration."
    elif is_yoga:
        bullet_1 = "Supports yoga poses, balance practice, stretching routines, and studio sessions."
        bullet_2 = f"Made with lightweight {material} for steady everyday practice." if material else "Made with lightweight foam for steady everyday practice."
        bullet_3 = "Beveled edges provide a comfortable grip during floor and standing poses."
        bullet_4 = f"Includes {set_count} piece set for home, gym, or studio use." if set_count else "Works well for home, gym, or studio use."
        bullet_5 = "Compact shape is easy to carry, stack, and store between sessions."
        desc = f"{subject} helps add lift, reach, and stability during yoga, stretching, and balance practice. The lightweight block is easy to carry and store for home, gym, and studio routines."
    else:
        bullet_1 = "Designed for everyday use in practical home, travel, or work settings."
        bullet_2 = f"Made with {material} for simple daily handling." if material else "Made for simple daily handling and repeated use."
        bullet_3 = "Compact design helps keep items organized without taking up extra space."
        bullet_4 = f"Includes {set_count} piece set for convenient replacement or sharing." if set_count else "Includes the product shown in the listing photos."
        bullet_5 = "Easy to carry, store, and use for a wide range of everyday needs."
        desc = f"{subject} is designed for everyday use with a simple, practical structure. It works well for home, travel, office, and general organization needs."
    return title, [bullet_1, bullet_2, bullet_3, bullet_4, bullet_5], desc


def _make_parent_copy(product_name, material, set_count):
    title, bullets, description = _make_copy(product_name, "", "", material, set_count)
    title = title.replace(", ,", ",").strip(" ,")
    return title, bullets, description


def _apply_cross_border_copy(row, product_name, price_table=None):
    price_table = price_table or {}
    english_name = _english_product_name(product_name, price_table.get("style_title", ""))
    style_phrase = _english_style_phrase(price_table.get("style_title", ""))
    if _has_cjk(row.get("title")):
        parts = [english_name] if english_name else []
        if style_phrase:
            parts.append(style_phrase)
        if row.get("color") and row.get("color") != "Assorted":
            parts.append(str(row.get("color")))
        row["title"] = ", ".join(parts)
    if _has_cjk(row.get("description")):
        row["description"] = "" if not english_name else f"{english_name} details need manual confirmation before Amazon copy is generated."
    return row


def _notes_for_strategy(strategy, source_colors, parent_sku=""):
    base = "List Price 留空人工填写；Haul Price 后续与 List Price 保持一致。"
    if strategy == "variation":
        return f"多链接变体无品牌路线：创建 Parent/Child 关系，Parent SKU：{parent_sku}，子体按款式分别建 SKU。采购单款式：{source_colors}。{base}"
    if strategy == "set_bundle":
        return f"单链接套装售卖无品牌路线：只创建 Child SKU，不设置 Parent/Child 变体；set_count/包装内容按套装数量填写。采购单款式：{source_colors}。{base}"
    return f"独立链接路线：每个款式单独 SKU，不设置 Parent/Child 变体。采购单款式：{source_colors}。{base}"


def _extract_color_variants(text):
    searchable = re.sub(r"\s+", " ", text)
    colors = []
    for color in CHINESE_COLOR_RE.findall(searchable):
        normalized = COLOR_WORDS.get(color, COLOR_WORDS.get(color.replace("色", ""), color))
        if normalized not in colors:
            colors.append(normalized)
    return colors


def _extract_floral_card_holder_variants(text):
    normalized_text = text.replace("⻆", "角")
    searchable = re.sub(r"\s+", "", normalized_text)
    variant_map = {
        "圆形卡插": ("Round", 1.5),
        "五角星卡插": ("Star", 1.5),
        "镀金小熊": ("Gold Bear", 3.0),
        "心形卡插": ("Heart", 1.5),
        "圆形粉色": ("Pink Round", 3.0),
    }
    variants = []
    for chinese_name, (english_name, fallback_price) in variant_map.items():
        if chinese_name in searchable:
            variants.append({
                "name": english_name,
                "cost": fallback_price,
            })
    return variants


def _guess_cost(text):
    unit_price = _first_match(UNIT_PRICE_RE, text)
    if unit_price:
        return float(unit_price)
    price = _first_match(PRICE_RE, text)
    return float(price) if price else ""


def _guess_main_image_url(urls):
    return ""


def _guess_item_type_keyword(project_dir, product_name, text):
    template_keyword = _extract_browse_node_from_templates(project_dir)
    if template_keyword:
        return template_keyword

    lowered = f"{product_name}\n{text}".lower()
    if "airtag" in lowered and any(keyword in lowered for keyword in ["collar", "项圈", "宠物", "猫咪", "狗狗"]):
        return "pet-collars"
    if any(keyword in lowered for keyword in ["cat toy", "q tip for cats", "q-tip", "catnip", "猫玩具", "逗猫", "猫薄荷", "磨牙"]):
        return "cat-toys"
    if "floral" in lowered or "bouquet" in lowered or "flower" in lowered or "花束" in lowered or "卡插" in lowered:
        return template_keyword or "cookbook-stands"
    if "yoga" in lowered or "瑜伽砖" in lowered:
        return "yoga-blocks"
    if "storage" in lowered or "收纳" in lowered:
        return "storage-bags"
    return "-".join(re.findall(r"[a-z0-9]+", product_name.lower())[:4])


def _extract_browse_node_from_templates(project_dir):
    project_dir = Path(project_dir)
    for folder in TEMPLATE_FOLDERS:
        template_dir = project_dir / folder
        if not template_dir.exists():
            continue
        for path in template_dir.glob("*"):
            if path.suffix.lower() not in SUPPORTED_EXCEL_EXT or path.name.startswith("~$"):
                continue
            try:
                wb = load_workbook(path, data_only=True, read_only=True, keep_vba=path.suffix.lower() == ".xlsm")
            except Exception:
                continue
            if "Browse Data" not in wb.sheetnames:
                continue
            ws = wb["Browse Data"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                value = row[0] if row and row[0] else None
                if value:
                    return str(value).strip()
    return ""


def _extract_product_type_from_templates(project_dir):
    project_dir = Path(project_dir)
    for folder in TEMPLATE_FOLDERS:
        template_dir = project_dir / folder
        if not template_dir.exists():
            continue
        for path in template_dir.glob("*"):
            if path.suffix.lower() not in SUPPORTED_EXCEL_EXT or path.name.startswith("~$"):
                continue
            if "EXERCISE_BLOCK" in path.stem.upper():
                return "EXERCISE_BLOCK"
            product_type = extract_template_product_type(path)
            if product_type:
                return product_type
    return ""


def analyze_project(project_dir, output_path=None, route_mode=None):
    project_dir = Path(project_dir)
    sources = _collect_sources(project_dir)
    combined = "\n".join(text for _path, text in sources)
    price_table = _extract_price_table_data(sources)

    urls = URL_RE.findall(combined)
    target_price = _guess_cost(combined)

    size_match = SIZE_RE.search(combined)
    length, width, height = _convert_size_to_inches(size_match) if size_match else ("", "", "")
    length = price_table.get("package_length_in") or length
    width = price_table.get("package_width_in") or width
    height = price_table.get("package_height_in") or height

    weight_match = WEIGHT_RE.search(combined)
    weight = _convert_weight_to_lb(weight_match) if weight_match else ""
    weight = price_table.get("package_weight_lb") or weight

    product_name = _guess_product_name(project_dir, combined)
    variant_colors = _extract_color_variants(combined)
    floral_variants = _extract_floral_card_holder_variants(combined)
    color = variant_colors[0] if variant_colors else (_extract_named_value(["颜色", "color"], combined) or _guess_from_words(COLOR_WORDS, combined))
    color = _normalize_color_value(color)
    size = _extract_named_value(["尺寸", "规格", "size"], combined)
    material = _extract_named_value(["材质", "material"], combined) or _guess_from_words(MATERIAL_WORDS, combined)
    set_count = price_table.get("set_count") or _extract_named_value(["套装数", "数量", "pack", "set count"], combined)
    brand_hint = _extract_named_value(["品牌", "brand"], combined)
    category = _guess_category(combined)
    product_type = _extract_product_type_from_templates(project_dir)
    route = _guess_route(combined, target_price if isinstance(target_price, float) else None, brand_hint)
    brand = brand_hint if route == "Brand" and brand_hint else "Generic"
    manufacturer = brand
    base_sku = price_table.get("sku") or _extract_named_value(["SKU", "sku", "货号"], combined)
    supplier_link = next((url for url in urls if "1688.com" in url or "alibaba" in url), urls[0] if urls else "")
    competitor_links = "\n".join(url for url in urls if "amazon." in url)
    main_image_url = _guess_main_image_url(urls)
    item_type_keyword = _guess_item_type_keyword(project_dir, product_name, combined)
    airtag_pet_collar = "airtag" in f"{product_name}\n{combined}".lower() and any(
        keyword in f"{product_name}\n{combined}"
        for keyword in ["项圈", "宠物", "猫咪", "狗狗", "collar"]
    )
    cat_qtip_product = (not airtag_pet_collar) and any(
        keyword in f"{product_name}\n{combined}".lower()
        for keyword in ["q tip for cats", "q-tip", "catnip", "猫玩具", "大棉签", "逗猫", "猫薄荷"]
    )
    if airtag_pet_collar:
        category = "pet"
        item_type_keyword = "pet-collars"
        if "涤纶" in combined and "硅胶" in combined:
            material = "Polyester and Silicone"
        elif not material or material in {"Cotton", "Cotton and Felt"}:
            material = "Polyester and Silicone"
        elif "silicone" not in material.lower():
            material = f"{material} and Silicone"
    if cat_qtip_product:
        category = "pet"
        item_type_keyword = "cat-toys"
        material = "Cotton and Felt"
        color = "Beige"
        size = size or "5.9 in"
        set_count = set_count or "3"
    source_colors = ", ".join(variant_colors) if variant_colors else color
    if floral_variants:
        colors_for_rows = [item["name"] for item in floral_variants]
    else:
        colors_for_rows = variant_colors or ([color] if color else [""])
    strategy = _route_strategy(combined, route, len(colors_for_rows), set_count, route_mode=route_mode)
    route = _route_name_for_strategy(route, strategy)
    if strategy == "set_bundle":
        colors_for_rows = ["Assorted"]
    parent_sku = _parent_sku(product_name) if strategy == "variation" else ""
    variation_theme = "Color" if strategy == "variation" else ""

    rows = []
    if strategy == "variation":
        parent_title, parent_bullets, parent_description = _make_parent_copy(product_name, material, set_count)
        rows.append(_apply_cross_border_copy({
            "project_name": project_dir.name,
            "product_name": product_name,
            "route": route,
            "brand": brand,
            "manufacturer": manufacturer,
            "category": category,
            "product_type": product_type,
            "item_type_keyword": item_type_keyword,
            "sku": parent_sku,
            "parent_sku": "",
            "parentage_level": "Parent",
            "variation_theme": variation_theme,
            "color": "",
            "size": "",
            "set_count": set_count or "",
            "material": material,
            "accessories": price_table.get("style_title") or _extract_named_value(["配件", "accessories"], combined),
            "cost": "",
            "target_price": "",
            "list_price": "",
            "haul_price": "",
            "package_length_in": "",
            "package_width_in": "",
            "package_height_in": "",
            "package_weight_lb": "",
            "country_of_origin": "China",
            "batteries_required": "No",
            "dangerous_goods": "No",
            "main_image_url": main_image_url,
            "supplier_link": supplier_link,
            "competitor_links": competitor_links,
            "title": parent_title,
            "bullet_1": parent_bullets[0],
            "bullet_2": parent_bullets[1],
            "bullet_3": parent_bullets[2],
            "bullet_4": parent_bullets[3],
            "bullet_5": parent_bullets[4],
            "description": parent_description,
            "notes": _notes_for_strategy(strategy, source_colors, parent_sku),
        }, product_name, price_table))

    for index, row_color in enumerate(colors_for_rows, 1):
        floral_variant = floral_variants[index - 1] if floral_variants else None
        if floral_variant:
            row_sku = f"FLORAL-CARD-HOLDER-{re.sub(r'[^A-Z0-9]+', '-', row_color.upper()).strip('-')}"
        elif len(colors_for_rows) > 1:
            row_sku = f"{base_sku}-{row_color}" if base_sku else _sku_for_color(product_name, row_color, index)
        else:
            row_sku = base_sku or _single_link_sku(product_name)
        row_color = _normalize_color_value(row_color)
        title, bullets, description = _make_copy(product_name, row_color, size, material, set_count)
        rows.append(_apply_cross_border_copy({
            "project_name": project_dir.name,
            "product_name": product_name,
            "route": route,
            "brand": brand,
            "manufacturer": manufacturer,
            "category": category,
            "product_type": product_type,
            "item_type_keyword": item_type_keyword,
            "sku": row_sku,
            "parent_sku": parent_sku if strategy == "variation" else "",
            "parentage_level": "Child" if strategy == "variation" else "",
            "variation_theme": variation_theme if strategy == "variation" else "",
            "color": row_color,
            "size": size,
            "set_count": set_count or (10 if floral_variant else ""),
            "material": material,
            "accessories": price_table.get("style_title") or _extract_named_value(["配件", "accessories"], combined),
            "cost": floral_variant["cost"] if floral_variant else (price_table.get("cost") or ("" if price_table else target_price)),
            "target_price": "",
            "list_price": price_table.get("list_price") or "",
            "haul_price": price_table.get("haul_price") or "",
            "package_length_in": length,
            "package_width_in": width,
            "package_height_in": height,
            "package_weight_lb": weight,
            "country_of_origin": "China",
            "batteries_required": "No",
            "dangerous_goods": "No",
            "main_image_url": main_image_url,
            "supplier_link": supplier_link,
            "competitor_links": competitor_links,
            "title": title,
            "bullet_1": bullets[0],
            "bullet_2": bullets[1],
            "bullet_3": bullets[2],
            "bullet_4": bullets[3],
            "bullet_5": bullets[4],
            "description": description,
            "notes": _notes_for_strategy(strategy, source_colors, parent_sku),
        }, product_name, price_table))

    if output_path is None:
        output_path = project_dir / PRODUCT_DETAIL_DIR / f"{safe_name(project_dir.name)}_自动提炼草稿.xlsx"
    write_intake_workbook(output_path, rows)

    report_path = OUTPUTS_DIR / f"{safe_name(project_dir.name)}_资料提炼报告.md"
    _write_analysis_report(report_path, project_dir, sources, rows, combined)
    return output_path, report_path


def _write_analysis_report(path, project_dir, sources, rows, combined):
    row = rows[0] if rows else {}
    risk_text = "\n".join(text for source_path, text in sources if source_path.suffix.lower() not in {".html", ".htm"})
    risk_hits = [word for word in RISK_WORDS if word.lower() in risk_text.lower()]
    source_lines = [f"- {source_path}" for source_path, _text in sources]
    lines = [
        f"# {Path(project_dir).name} 资料提炼报告",
        "",
        "## 已读取资料",
        "",
        *(source_lines or ["- 未找到可读取的 txt/md/csv/tsv/xlsx/xlsm/pdf/html 资料。"]),
        "",
        "## 提炼结果",
        "",
        f"- 产品名：{row.get('product_name')}",
        f"- 路线建议：{row.get('route')}",
        f"- Brand：{row.get('brand')}",
        f"- 类目猜测：{row.get('category')}",
        f"- SKU：{row.get('sku')}",
        f"- Parent SKU：{row.get('sku') if row.get('parentage_level') == 'Parent' else row.get('parent_sku') or '无'}",
        f"- 父子层级：{', '.join(str(item.get('parentage_level')) for item in rows if item.get('parentage_level')) or '独立 SKU'}",
        f"- 颜色：{', '.join(str(item.get('color')) for item in rows if item.get('color'))}",
        f"- 草稿 SKU 数：{len(rows)}",
        f"- 尺寸 inch：{row.get('package_length_in')} x {row.get('package_width_in')} x {row.get('package_height_in')}",
        f"- 重量 lb：{row.get('package_weight_lb')}",
        f"- 链接数量：{len(URL_RE.findall(combined))}",
        "",
        "## 需要人工确认",
        "",
        "- item_type_keyword",
        "- List Price / Haul Price",
        "- 主图 URL 是否是可公开访问的图片直链",
        "- 是否真的适合 Generic 或 Haul Generic",
        "- 亚马逊模板里的条件必填字段",
        ""
    ]
    if risk_hits:
        lines.extend([
            "## 风险词命中",
            "",
            ", ".join(sorted(set(risk_hits))),
            ""
        ])
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")
